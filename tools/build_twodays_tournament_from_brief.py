from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from build_twodays_tournament_template import (
    BORDER,
    HEADER_FILL,
    OLD_MONTH_PREFIX,
    OLD_PREFIX,
    OLD_QUEST_GROUP_FOLDER,
    ROOT,
    SOURCE_ROOT,
    compact_json,
    copy_cell_style,
    deep_get,
    load_proto,
    new_conf_workbook,
    source_path,
    write_table,
)


DEFAULT_BRIEF = ROOT / "input" / "TwoDaysTournament_master_brief.xlsx"
OUT_ROOT = ROOT / "output"

OLD_RESOURCE = f"{OLD_PREFIX}_Resource_Competition"
OLD_PET = f"Pet19_{OLD_PREFIX}_1"
OLD_TROPHY = f"TrophyPet19_{OLD_PREFIX}_1"

FAMILY_ORDER = [
    "quest_item",
    "collection_item",
    "furniture",
    "pet",
    "mystery_box",
    "competition",
    "magazine",
    "recipe",
    "quest",
    "quest_action",
    "quest_group",
    "trophy",
]


def clean(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def text(value: Any) -> str:
    return "" if value is None else str(value)


def sheet_records(ws) -> list[dict[str, Any]]:
    headers = [text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    records: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        first = clean(ws.cell(row, 1).value)
        if first is None:
            continue
        records.append({headers[col - 1]: clean(ws.cell(row, col).value) for col in range(1, len(headers) + 1)})
    return records


def parse_brief(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    event = {record["field"]: record.get("value") for record in sheet_records(wb["01_Event"])}
    objects = {
        (record["object_key"], record["field"]): record.get("new_value")
        for record in sheet_records(wb["02_Objects_Text"])
    }
    magazine = {record["key"]: record.get("new_text") for record in sheet_records(wb["03_Magazine_Text"])}
    return {"event": event, "objects": objects, "magazine": magazine}


def brief_value(data: dict[str, Any], key: str, default: Any = None) -> Any:
    value = data["event"].get(key)
    return default if value in (None, "") else value


def object_value(data: dict[str, Any], object_key: str, field: str, default: Any = None) -> Any:
    value = data["objects"].get((object_key, field))
    return default if value in (None, "") else value


def magazine_value(data: dict[str, Any], key: str, default: Any = None) -> Any:
    value = data["magazine"].get(key)
    return default if value in (None, "") else value


class ProtoIdAllocator:
    def __init__(self, first_free: Any):
        self.current: int | None = None
        if isinstance(first_free, int):
            self.current = first_free - 1
        elif isinstance(first_free, float) and first_free.is_integer():
            self.current = int(first_free) - 1
        elif isinstance(first_free, str) and re.fullmatch(r"\d+", first_free.strip()):
            self.current = int(first_free.strip()) - 1
        self.first_free = self.current + 1 if self.current is not None else None

    @property
    def enabled(self) -> bool:
        return self.current is not None

    def next(self) -> int | None:
        if self.current is None:
            return None
        self.current += 1
        return self.current


class TaskIdAllocator:
    def __init__(self, first_free: Any):
        self.prefix = ""
        self.width = 0
        self.current: int | None = None
        if first_free in (None, ""):
            self.first_free = None
            return
        match = re.fullmatch(r"([A-Za-z]*)(\d+)", str(first_free).strip())
        if not match:
            self.first_free = None
            return
        self.prefix = match.group(1)
        numeric = match.group(2)
        self.width = len(numeric)
        self.current = int(numeric) - 1
        self.first_free = str(first_free).strip()

    @property
    def enabled(self) -> bool:
        return self.current is not None

    def next(self) -> str | None:
        if self.current is None:
            return None
        self.current += 1
        return f"{self.prefix}{self.current:0{self.width}d}"

    def last_used(self) -> str | None:
        if self.current is None:
            return None
        return f"{self.prefix}{self.current:0{self.width}d}"

    def next_free(self) -> str | None:
        if self.current is None:
            return None
        return f"{self.prefix}{self.current + 1:0{self.width}d}"


def replace_string(value: str, mapping: dict[str, str]) -> str:
    result = value
    for old in sorted(mapping, key=len, reverse=True):
        result = result.replace(old, mapping[old])
    return result


def replace_deep(value: Any, mapping: dict[str, str]) -> Any:
    if isinstance(value, str):
        return replace_string(value, mapping)
    if isinstance(value, list):
        return [replace_deep(item, mapping) for item in value]
    if isinstance(value, dict):
        return {key: replace_deep(item, mapping) for key, item in value.items()}
    return value


def after_condition(condition: Any) -> str | None:
    raw = clean(condition)
    if raw is None:
        return None
    raw_text = str(raw)
    if raw_text.startswith("time<"):
        return "time>" + raw_text[len("time<") :]
    return raw_text


def context(brief: dict[str, Any]) -> dict[str, Any]:
    new_prefix = str(brief_value(brief, "new_prefix"))
    new_month_prefix = str(brief_value(brief, "new_month_prefix", new_prefix.rsplit("_", 1)[0]))
    new_quest_group_folder = str(
        brief_value(brief, "new_quest_group_folder", new_prefix.replace("TwoDaysTournament_", ""))
    )
    resource = str(brief_value(brief, "resource_classname", f"{new_prefix}_Resource_Competition"))
    pet = str(brief_value(brief, "pet_classname", f"Pet19_{new_prefix}_1"))
    trophy = str(brief_value(brief, "trophy_classname", f"TrophyPet19_{new_prefix}_1"))
    mapping = {
        OLD_TROPHY: trophy,
        OLD_PET: pet,
        OLD_RESOURCE: resource,
        OLD_PREFIX: new_prefix,
        OLD_MONTH_PREFIX: new_month_prefix,
        OLD_QUEST_GROUP_FOLDER: new_quest_group_folder,
        "Майского турнира 2026": str(brief_value(brief, "event_title_genitive_ru", "Августовского турнира 2026")),
        "Майский турнир": str(brief_value(brief, "event_title_ru", "Августовский турнир")),
    }
    return {
        "new_prefix": new_prefix,
        "new_month_prefix": new_month_prefix,
        "new_quest_group_folder": new_quest_group_folder,
        "resource": resource,
        "pet": pet,
        "trophy": trophy,
        "mapping": mapping,
        "event_title": brief_value(brief, "event_title_ru", "Августовский турнир"),
        "event_title_genitive": brief_value(brief, "event_title_genitive_ru", "Августовского турнира 2026"),
        "resource_title": brief_value(brief, "resource_title", "Конфета"),
        "pet_title": object_value(brief, "pet", "title", brief_value(brief, "pet_title", "")),
        "pet_description": object_value(brief, "pet", "description", brief_value(brief, "pet_description", "")),
        "pet_default_name": object_value(brief, "pet", "default_name", brief_value(brief, "pet_default_name", "")),
        "pet_title_attention": object_value(brief, "pet", "title_attention", brief_value(brief, "pet_title_attention", "")),
        "competition_open_time": brief_value(brief, "competition_open_time"),
        "competition_close_time": brief_value(brief, "competition_close_time"),
        "competition_finish_condition": brief_value(brief, "competition_finish_condition"),
        "event_active_until": brief_value(brief, "event_active_until"),
        "tech_start_condition": brief_value(brief, "tech_start_condition"),
        "pet_encyclopedia_condition": brief_value(brief, "pet_encyclopedia_condition"),
        "title_date_short": brief_value(brief, "title_date_short"),
        "title_date_full": brief_value(brief, "title_date_full"),
    }


def target_path(proto_path: Path, ctx: dict[str, Any]) -> str:
    return replace_string(source_path(proto_path), ctx["mapping"])


def row_for(proto_path: Path, ctx: dict[str, Any], ids: ProtoIdAllocator, extra: dict[str, Any]) -> dict[str, Any]:
    return {
        "input": source_path(proto_path),
        "output": target_path(proto_path, ctx),
        "replace": (OLD_PREFIX, ctx["new_prefix"]),
        "id": ids.next(),
        **extra,
    }


def build_task_ids(ids: TaskIdAllocator) -> dict[str, str | None]:
    result = {
        "trophy": ids.next(),
        "activate": ids.next(),
        "counter": ids.next(),
        "task1": ids.next(),
        "task1_restarter": ids.next(),
        "task2": ids.next(),
        "task2_restarter": ids.next(),
        "task3": ids.next(),
        "task3_restarter": ids.next(),
        "tech1": ids.next(),
        "tech2": ids.next(),
        "tech_magazine": ids.next(),
        "tech_magazine_shop": ids.next(),
    }
    return result


def quest_task_id_for(filename: str, task_ids: dict[str, str | None]) -> str | None:
    mapping = {
        "TwoDaysTournament_2026_05_01_Competition_1_Activate.proto.js": "activate",
        "TwoDaysTournament_2026_05_01_Competition_1_Counter.proto.js": "counter",
        "TwoDaysTournament_2026_05_01_Competition_1_Task1.proto.js": "task1",
        "TwoDaysTournament_2026_05_01_Competition_1_Task1_restarter.proto.js": "task1_restarter",
        "TwoDaysTournament_2026_05_01_Competition_1_Task2.proto.js": "task2",
        "TwoDaysTournament_2026_05_01_Competition_1_Task2_restarter.proto.js": "task2_restarter",
        "TwoDaysTournament_2026_05_01_Competition_1_Task3.proto.js": "task3",
        "TwoDaysTournament_2026_05_01_Competition_1_Task3_restarter.proto.js": "task3_restarter",
        "TwoDaysTournament_2026_05_01_Tech1.proto.js": "tech1",
        "TwoDaysTournament_2026_05_01_Tech2.proto.js": "tech2",
        "TwoDaysTournament_2026_05_01_Tech_Magazine.proto.js": "tech_magazine",
        "TwoDaysTournament_2026_05_01_Tech_MagazineShopAvailable.proto.js": "tech_magazine_shop",
    }
    key = mapping.get(filename)
    return task_ids.get(key) if key else None


def build_rows(source_root: Path, brief: dict[str, Any]) -> tuple[dict[str, tuple[list[tuple[str, str]], list[dict[str, Any]]]], dict[str, Any]]:
    ctx = context(brief)
    proto_ids = ProtoIdAllocator(brief_value(brief, "last_proto_id"))
    task_allocator = TaskIdAllocator(brief_value(brief, "last_task_id"))
    task_ids = build_task_ids(task_allocator)
    families: dict[str, tuple[list[tuple[str, str]], list[dict[str, Any]]]] = {}

    qi_path = next((source_root / "quest_item").glob("*.proto.js"))
    qi = load_proto(qi_path)
    families["quest_item"] = (
        [("string", "classname"), ("string", "title"), ("string", "subgroup"), ("string", "inventory_section"), ("int", "id")],
        [row_for(qi_path, ctx, proto_ids, {
            "classname": ctx["resource"],
            "title": ctx["resource_title"],
            "subgroup": qi.get("subgroup"),
            "inventory_section": qi.get("inventory_section"),
        })],
    )

    ci_path = next((source_root / "collection_item").glob("*.proto.js"))
    ci = load_proto(ci_path)
    families["collection_item"] = (
        [("string", "classname"), ("string", "title"), ("string", "description"), ("object", "rand_reward"), ("int", "id")],
        [row_for(ci_path, ctx, proto_ids, {
            "classname": ctx["trophy"],
            "title": object_value(brief, "collection_item", "title", f"Медалька \"{ctx['pet_title']}\""),
            "description": ci.get("description"),
            "rand_reward": replace_deep(ci.get("rand_reward"), ctx["mapping"]),
        })],
    )

    furniture_rows: list[dict[str, Any]] = []
    cup_key_by_name = {
        "Cup_Gold": "cup_gold",
        "Cup_Silver": "cup_silver",
        "Cup_Bronze": "cup_bronze",
        "Cup_Iron": "cup_iron",
    }
    for proto_path in sorted((source_root / "furniture").glob("*.proto.js")):
        data = load_proto(proto_path)
        cup_key = next((key for marker, key in cup_key_by_name.items() if marker in proto_path.name), "")
        title = object_value(brief, cup_key, "title", replace_string(data.get("title", ""), ctx["mapping"]))
        furniture_rows.append(row_for(proto_path, ctx, proto_ids, {
            "classname": replace_string(data.get("classname"), ctx["mapping"]),
            "title": title,
            "subgroup": data.get("subgroup"),
            "price": data.get("price"),
            "currency": data.get("currency"),
            "meta_info": replace_string(data.get("meta_info", ""), ctx["mapping"]),
            "extra.wdr_type": deep_get(data, "extra.wdr_type"),
        }))
    families["furniture"] = (
        [("string", "classname"), ("string", "title"), ("string", "subgroup"), ("int", "price"), ("string", "currency"), ("string", "meta_info"), ("string", "extra.wdr_type"), ("int", "id")],
        furniture_rows,
    )

    pet_path = next((source_root / "pet").glob("*.proto.js"))
    pet = load_proto(pet_path)
    families["pet"] = (
        [
            ("string", "classname"),
            ("string", "title"),
            ("string", "description"),
            ("string", "subgroup"),
            ("string", "extra.default_name"),
            ("string", "extra.event_frame"),
            ("string", "extra.feed_price"),
            ("string", "extra.title_attention"),
            ("string", "extra.encyclopedia_conditions"),
            ("int", "id"),
        ],
        [row_for(pet_path, ctx, proto_ids, {
            "classname": ctx["pet"],
            "title": ctx["pet_title"],
            "description": ctx["pet_description"],
            "subgroup": pet.get("subgroup"),
            "extra.default_name": ctx["pet_default_name"],
            "extra.event_frame": deep_get(pet, "extra.event_frame"),
            "extra.feed_price": deep_get(pet, "extra.feed_price"),
            "extra.title_attention": ctx["pet_title_attention"],
            "extra.encyclopedia_conditions": ctx["pet_encyclopedia_condition"],
        })],
    )

    mb_rows: list[dict[str, Any]] = []
    for proto_path in sorted((source_root / "mystery_box").glob("*.proto.js")):
        data = load_proto(proto_path)
        mb_rows.append(row_for(proto_path, ctx, proto_ids, {
            "classname": replace_string(data.get("classname"), ctx["mapping"]),
            "title": data.get("title"),
            "description": data.get("description"),
            "price": data.get("price"),
            "currency": data.get("currency"),
            "meta_info": data.get("meta_info"),
            "rand_reward": replace_deep(data.get("rand_reward"), ctx["mapping"]),
        }))
    families["mystery_box"] = (
        [("string", "classname"), ("string", "title"), ("string", "description"), ("int", "price"), ("string", "currency"), ("string", "meta_info"), ("object", "rand_reward"), ("int", "id")],
        mb_rows,
    )

    comp_path = next((source_root / "competition").glob("*.proto.js"))
    comp = load_proto(comp_path)
    competition_id = proto_ids.next()
    comp_extra = replace_deep(comp.get("extra", {}), ctx["mapping"])
    comp_extra["title_date_short"] = ctx["title_date_short"]
    comp_extra["title_date_full"] = ctx["title_date_full"]
    families["competition"] = (
        [
            ("string", "identifier"),
            ("string", "quest_task_identifier"),
            ("string", "quest_group_identifier"),
            ("string", "update_top_classname"),
            ("string", "finish_conditions"),
            ("string", "active_conditions"),
            ("int", "leaders_top_amount"),
            ("int", "leaders_top_amount_my_mail"),
            ("int", "leaders_top_amount_odnoklassniki"),
            ("string", "icon"),
            ("object", "extra"),
            ("object", "rewards"),
            ("object", "rewards_my_mail"),
            ("object", "rewards_odnoklassniki"),
            ("int", "id"),
        ],
        [{
            "input": source_path(comp_path),
            "output": target_path(comp_path, ctx),
            "identifier": f"{ctx['new_prefix']}_Competition_1",
            "quest_task_identifier": task_ids["counter"],
            "quest_group_identifier": f"{ctx['new_prefix']}_Competition_1",
            "update_top_classname": ctx["resource"],
            "finish_conditions": ctx["competition_finish_condition"],
            "active_conditions": ctx["event_active_until"],
            "leaders_top_amount": comp.get("leaders_top_amount"),
            "leaders_top_amount_my_mail": comp.get("leaders_top_amount_my_mail"),
            "leaders_top_amount_odnoklassniki": comp.get("leaders_top_amount_odnoklassniki"),
            "icon": replace_string(comp.get("icon"), ctx["mapping"]),
            "extra": comp_extra,
            "rewards": replace_deep(comp.get("rewards"), ctx["mapping"]),
            "rewards_my_mail": replace_deep(comp.get("rewards_my_mail"), ctx["mapping"]),
            "rewards_odnoklassniki": replace_deep(comp.get("rewards_odnoklassniki"), ctx["mapping"]),
            "id": competition_id,
        }],
    )

    magazine_path = next((source_root / "magazine").glob("*.proto.js"))
    magazine = load_proto(magazine_path)
    magazine_fields = [
        "identifier",
        "title",
        "view",
        "pages.0.label_title",
        "pages.0.text_fields.0.text",
        "pages.0.text_fields.1.text",
        "pages.1.label_title",
        "pages.1.text_fields.1.text",
        "pages.2.label_title",
        "pages.2.text_fields.0.text",
        "pages.2.text_fields.1.text",
        "pages.2.text_fields.2.text",
        "pages.2.text_fields.3.text",
        "pages.3.label_title",
        "pages.3.text_fields.0.text",
        "pages.3.text_fields.1.text",
        "pages.3.main_resource",
    ]
    magazine_extra = {key: magazine_value(brief, key if key != "title" else "magazine.title", replace_string(text(deep_get(magazine, key)), ctx["mapping"])) for key in magazine_fields}
    magazine_extra["identifier"] = f"{ctx['new_prefix']}_Magazine"
    magazine_extra["title"] = magazine_value(brief, "magazine.title", ctx["event_title"])
    magazine_extra["view"] = replace_string(magazine.get("view"), ctx["mapping"])
    magazine_extra["pages.3.main_resource"] = ctx["resource"]
    magazine_extra["replace"] = (OLD_PREFIX, ctx["new_prefix"])
    families["magazine"] = (
        [("string", key) for key in magazine_fields] + [("replace", "replace"), ("int", "id")],
        [row_for(magazine_path, ctx, proto_ids, magazine_extra)],
    )

    recipe_rows: list[dict[str, Any]] = []
    for proto_path in sorted((source_root / "recipe").glob("*.proto.js")):
        data = load_proto(proto_path)
        recipe_rows.append(row_for(proto_path, ctx, proto_ids, {
            "identifier": replace_string(data.get("identifier"), ctx["mapping"]),
            "reward": replace_string(data.get("reward"), ctx["mapping"]),
            "ingredients": replace_string(data.get("ingredients"), ctx["mapping"]),
        }))
    families["recipe"] = ([("string", "identifier"), ("string", "reward"), ("string", "ingredients"), ("int", "id")], recipe_rows)

    quest_rows: list[dict[str, Any]] = []
    quest_paths = sorted((source_root / "quest").glob("*.proto.js")) + sorted((source_root / "quest" / "competition").glob("*.proto.js"))
    for proto_path in quest_paths:
        data = load_proto(proto_path)
        filename = proto_path.name
        conditions = replace_string(text(data.get("conditions")), ctx["mapping"]) if data.get("conditions") else None
        if filename.endswith("_Competition_1_Activate.proto.js"):
            conditions = f"time>{ctx['competition_open_time']}+time<{ctx['competition_close_time']}"
        elif filename.endswith("_Tech1.proto.js"):
            conditions = f"{ctx['tech_start_condition']}+{ctx['event_active_until']}"

        title = replace_string(text(data.get("title")), ctx["mapping"]) if data.get("title") else None
        if filename.endswith("_Tech_Magazine.proto.js"):
            title = f"{ctx['event_title']}!"

        on_accomplish_cid = competition_id if filename.endswith(("_Task1.proto.js", "_Task2.proto.js", "_Task3.proto.js")) else None
        quest_rows.append(row_for(proto_path, ctx, proto_ids, {
            "identifier": replace_string(data.get("identifier"), ctx["mapping"]),
            "conditions": conditions,
            "group_identifier": replace_string(text(data.get("group_identifier")), ctx["mapping"]) if data.get("group_identifier") else None,
            "progress_arrow_id": replace_string(text(data.get("progress_arrow_id")), ctx["mapping"]) if data.get("progress_arrow_id") else None,
            "title": title,
            "congratulation": replace_string(text(data.get("congratulation")), ctx["mapping"]) if data.get("congratulation") else None,
            "tasks.0.identifier": quest_task_id_for(filename, task_ids),
            "tasks.0.type": deep_get(data, "tasks.0.type"),
            "tasks.0.generator": deep_get(data, "tasks.0.generator"),
            "tasks.0.reward_classname": replace_string(text(deep_get(data, "tasks.0.reward_classname")), ctx["mapping"]) if deep_get(data, "tasks.0.reward_classname") else None,
            "tasks.0.classname": replace_string(text(deep_get(data, "tasks.0.classname")), ctx["mapping"]) if deep_get(data, "tasks.0.classname") else None,
            "tasks.0.title": replace_string(text(deep_get(data, "tasks.0.title")), ctx["mapping"]) if deep_get(data, "tasks.0.title") else None,
            "tasks.0.amount": deep_get(data, "tasks.0.amount"),
            "on_accomplish.0.cid": on_accomplish_cid,
            "replace": (OLD_PREFIX, ctx["new_prefix"]),
        }))
    families["quest"] = (
        [
            ("string", "identifier"),
            ("string", "conditions"),
            ("string", "group_identifier"),
            ("string", "progress_arrow_id"),
            ("string", "title"),
            ("string", "congratulation"),
            ("string", "tasks.0.identifier"),
            ("string", "tasks.0.type"),
            ("string", "tasks.0.generator"),
            ("string", "tasks.0.reward_classname"),
            ("string", "tasks.0.classname"),
            ("string", "tasks.0.title"),
            ("int", "tasks.0.amount"),
            ("int", "on_accomplish.0.cid"),
            ("replace", "replace"),
            ("int", "id"),
        ],
        quest_rows,
    )

    action_rows: list[dict[str, Any]] = []
    for proto_path in sorted((source_root / "quest_action").glob("*.proto.js")):
        data = load_proto(proto_path)
        action_rows.append(row_for(proto_path, ctx, proto_ids, {
            "identifier": replace_string(data.get("identifier"), ctx["mapping"]),
            "conditions": replace_string(text(data.get("conditions")), ctx["mapping"]),
            "open_price": data.get("open_price"),
            "replace": (OLD_PREFIX, ctx["new_prefix"]),
        }))
    families["quest_action"] = ([("string", "identifier"), ("string", "conditions"), ("string", "open_price"), ("replace", "replace"), ("int", "id")], action_rows)

    group_rows: list[dict[str, Any]] = []
    for proto_path in sorted((source_root / "quest_group").glob("*.proto.js")):
        data = load_proto(proto_path)
        spoil_conditions = replace_string(text(data.get("spoil_conditions")), ctx["mapping"]) if data.get("spoil_conditions") else None
        if proto_path.name.endswith("_Competition_1.proto.js"):
            spoil_conditions = f"time>{ctx['competition_close_time']}"
        elif proto_path.name.endswith("_Tech.proto.js"):
            spoil_conditions = after_condition(ctx["event_active_until"])
        title = replace_string(text(data.get("title")), ctx["mapping"]) if data.get("title") else None
        if proto_path.name.endswith("_Competition_1.proto.js"):
            title = ctx["event_title"]
        group_rows.append(row_for(proto_path, ctx, proto_ids, {
            "identifier": replace_string(data.get("identifier"), ctx["mapping"]),
            "title": title,
            "description": data.get("description"),
            "description_spoil": data.get("description_spoil"),
            "spoil_conditions": spoil_conditions,
            "first_quest": replace_string(text(data.get("first_quest")), ctx["mapping"]),
            "last_quest": replace_string(text(data.get("last_quest")), ctx["mapping"]),
            "extra": replace_deep(data.get("extra"), ctx["mapping"]),
            "replace": (OLD_PREFIX, ctx["new_prefix"]),
        }))
    families["quest_group"] = (
        [("string", "identifier"), ("string", "title"), ("string", "description"), ("string", "description_spoil"), ("string", "spoil_conditions"), ("string", "first_quest"), ("string", "last_quest"), ("object", "extra"), ("replace", "replace"), ("int", "id")],
        group_rows,
    )

    trophy_path = next((source_root / "trophy").glob("*.proto.js"))
    trophy = load_proto(trophy_path)
    families["trophy"] = (
        [("string", "identifier"), ("string", "title"), ("string", "conditions"), ("string", "pass_conditions"), ("string", "tasks.0.identifier"), ("string", "tasks.0.title"), ("string", "tasks.0.param"), ("replace", "replace"), ("int", "id")],
        [row_for(trophy_path, ctx, proto_ids, {
            "identifier": ctx["trophy"],
            "title": object_value(brief, "trophy", "title", ctx["pet_title"]),
            "conditions": f"asset_or_stuff={ctx['pet']}:1",
            "pass_conditions": f"asset={ctx['trophy']}:1",
            "tasks.0.identifier": task_ids["trophy"],
            "tasks.0.title": object_value(brief, "trophy", "title", ctx["pet_title"]),
            "tasks.0.param": ctx["pet"],
            "replace": (OLD_PREFIX, ctx["new_prefix"]),
        })],
    )

    report = {
        "new_prefix": ctx["new_prefix"],
        "first_proto_id": proto_ids.first_free,
        "last_proto_id": proto_ids.current,
        "next_proto_id": proto_ids.current + 1 if proto_ids.current is not None else None,
        "first_task_id": task_allocator.first_free,
        "last_task_id": task_allocator.last_used(),
        "next_task_id": task_allocator.next_free(),
        "competition_id": competition_id,
    }
    return families, report


def save_outputs(families: dict[str, tuple[list[tuple[str, str]], list[dict[str, Any]]]], report: dict[str, Any], out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    files: list[Path] = []
    combined = Workbook()
    combined.remove(combined.active)

    readme = combined.create_sheet("README")
    for row, line in enumerate(
        [
            "TwoDaysTournament generated tables",
            f"new_prefix: {report['new_prefix']}",
            f"first_proto_id: {report['first_proto_id']}",
            f"last_proto_id: {report['last_proto_id']}",
            f"next_proto_id: {report['next_proto_id']}",
            f"first_task_id: {report['first_task_id']}",
            f"last_task_id: {report['last_task_id']}",
            f"next_task_id: {report['next_task_id']}",
            f"competition_id: {report['competition_id']}",
        ],
        1,
    ):
        readme.cell(row, 1).value = line
    readme.column_dimensions["A"].width = 80

    for family in FAMILY_ORDER:
        fields, rows = families[family]
        wb, ws = new_conf_workbook(f"{family} - {report['new_prefix']}")
        write_table(ws, f"{family} - {report['new_prefix']}", fields, rows)
        path = out_dir / f"{family}.xlsx"
        wb.save(path)
        files.append(path)

        combined_ws = combined.create_sheet(family[:31])
        for source_row in ws.iter_rows():
            for source_cell in source_row:
                target_cell = combined_ws.cell(source_cell.row, source_cell.column)
                target_cell.value = source_cell.value
                copy_cell_style(source_cell, target_cell)
        for merged in ws.merged_cells.ranges:
            combined_ws.merge_cells(str(merged))
        for key, dim in ws.column_dimensions.items():
            combined_ws.column_dimensions[key].width = dim.width

    combined_path = out_dir / f"{report['new_prefix']}_proto_tables.xlsx"
    combined.save(combined_path)
    files.append(combined_path)

    report_path = out_dir / "generation_report.txt"
    report_path.write_text("\n".join(f"{key}: {value}" for key, value in report.items()) + "\n", encoding="utf-8")
    files.append(report_path)
    return files


def main() -> None:
    parser = argparse.ArgumentParser(description="Build TwoDaysTournament final converter tables from master brief.")
    parser.add_argument("--brief", type=Path, default=DEFAULT_BRIEF)
    parser.add_argument("--source-root", type=Path, default=SOURCE_ROOT)
    parser.add_argument("--out-dir", type=Path)
    args = parser.parse_args()

    brief = parse_brief(args.brief)
    ctx = context(brief)
    out_dir = args.out_dir or OUT_ROOT / f"{ctx['new_prefix']}_pack_tables"
    families, report = build_rows(args.source_root, brief)
    files = save_outputs(families, report, out_dir)
    print(out_dir)
    for file in files:
        print(file)


if __name__ == "__main__":
    main()
