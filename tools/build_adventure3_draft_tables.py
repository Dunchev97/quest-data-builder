from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_ROOT = Path("input/Adventure_3")
OUTPUT_XLSX = Path("output/Adventure_3_proto_tables_draft.xlsx")
DIRECTIVES = {"sl", "ml", "temp_01", "temp_02"}
PAIR_TYPES = {"object", "array_of_objects", "replace"}
SKIP_FIELDS = {"class"}

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="E2F0D9")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass
class Field:
    typ: str
    key: str
    values: list[Any]
    width: int = 1


def load_proto(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def json_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def cell_type(value: Any) -> str:
    if isinstance(value, int) and not isinstance(value, bool):
        return "int"
    return "string"


def folder_ref_path(folder: Path) -> str:
    raw = (folder / "Path.txt").read_text(encoding="utf-8").strip()
    normalized = raw.replace("\\", "/").strip("/")
    return "/" + normalized


def proto_ref(folder: Path, proto: Path) -> str:
    return f"{folder_ref_path(folder)}/{proto.name}"


def role_for(folder: str, filename: str) -> str:
    stem = filename.removesuffix(".proto.js")
    if folder == "quest_action":
        if stem.startswith("action_Free_Reset"):
            return "free_reset_action"
        if stem.startswith("action_Reset"):
            return "paid_reset_action"
        if stem.startswith("action_Skip"):
            return "skip_action"
    if stem.endswith("_tech"):
        return "tech"
    if stem.endswith("_rewards"):
        return "rewards"
    if "restarter" in stem:
        return "restarter"
    if "quest_task_" in stem:
        return "task"
    return "main"


def object_field(key: str, value: dict[str, Any]) -> Field:
    return Field("object", key, [(k, json_cell(v)) for k, v in value.items()], width=2)


def array_field(key: str, value: list[Any]) -> Field:
    return Field("array", key, [json_cell(v) for v in value])


def fields_for_proto(data: dict[str, Any]) -> list[Field]:
    fields: list[Field] = []
    for key, value in data.items():
        if key in SKIP_FIELDS:
            continue
        if isinstance(value, dict):
            fields.append(object_field(key, value))
        elif isinstance(value, list):
            fields.append(array_field(key, value))
        else:
            fields.append(Field(cell_type(value), key, [json_cell(value)]))
    return fields


def write_title(ws, row: int, title: str) -> int:
    ws.cell(row, 2, title)
    ws.cell(row, 2).fill = TITLE_FILL
    ws.cell(row, 2).font = Font(bold=True)
    return row + 1


def write_headers(ws, row: int, fields: list[Field]) -> list[int]:
    cols: list[int] = []
    col = 4
    for field in fields:
        cols.append(col)
        ws.cell(row, col, field.typ)
        ws.cell(row + 1, col, field.key)
        if field.width > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + field.width - 1)
            ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + field.width - 1)
        col += field.width
    return cols


def write_values(ws, row: int, col: int, field: Field) -> int:
    if field.typ in PAIR_TYPES:
        rows = field.values or [(None, None)]
        for offset, (left, right) in enumerate(rows):
            ws.cell(row + offset, col, left)
            ws.cell(row + offset, col + 1, right)
        return len(rows)

    rows = field.values or [None]
    for offset, value in enumerate(rows):
        ws.cell(row + offset, col, value)
    return len(rows)


def write_ml_block(ws, row: int, title: str, input_path: str, output_path: str, fields: list[Field]) -> int:
    row = write_title(ws, row, title)
    ws.cell(row, 1, "ml")
    ws.cell(row, 2, "string")
    ws.cell(row, 3, "string")
    cols = write_headers(ws, row, fields)
    row += 1
    ws.cell(row, 2, "input")
    ws.cell(row, 3, "output")
    row += 1
    ws.cell(row, 2, input_path)
    ws.cell(row, 3, output_path)

    max_rows = 1
    for col, field in zip(cols, fields):
        max_rows = max(max_rows, write_values(ws, row, col, field))
    return row + max_rows + 2


def simple_value(field: Field) -> Any:
    if len(field.values) != 1:
        return None
    value = field.values[0]
    if isinstance(value, (dict, list, tuple)):
        return None
    return value


def write_sl_group(ws, row: int, title: str, rows: list[tuple[str, str, list[Field]]]) -> int:
    if not rows:
        return row
    ordered_keys: list[str] = []
    field_by_key: dict[str, Field] = {}
    for _input_path, _output_path, fields in rows:
        for field in fields:
            if field.width == 1 and field.key not in ordered_keys:
                ordered_keys.append(field.key)
                field_by_key[field.key] = field

    row = write_title(ws, row, title)
    ws.cell(row, 1, "sl")
    ws.cell(row, 2, "string")
    ws.cell(row, 3, "string")
    for offset, key in enumerate(ordered_keys, start=4):
        ws.cell(row, offset, field_by_key[key].typ)
    row += 1
    ws.cell(row, 2, "input")
    ws.cell(row, 3, "output")
    for offset, key in enumerate(ordered_keys, start=4):
        ws.cell(row, offset, key)
    row += 1

    for input_path, output_path, fields in rows:
        by_key = {field.key: simple_value(field) for field in fields}
        ws.cell(row, 2, input_path)
        ws.cell(row, 3, output_path)
        for offset, key in enumerate(ordered_keys, start=4):
            ws.cell(row, offset, by_key.get(key))
        row += 1
    return row + 1


def can_be_sl(fields: list[Field]) -> bool:
    return all(field.width == 1 and len(field.values) == 1 for field in fields)


def write_family_sheet(wb: Workbook, folder: Path) -> None:
    ws = wb.create_sheet(folder.name[:31])
    row = 1
    row = write_title(ws, row, f"folder: {folder.name}, input base: {folder_ref_path(folder)}")

    sl_rows_by_role: dict[str, list[tuple[str, str, list[Field]]]] = {}

    for proto in sorted(folder.glob("*.proto.js")):
        data = load_proto(proto)
        fields = fields_for_proto(data)
        input_path = proto_ref(folder, proto)
        output_path = input_path
        role = role_for(folder.name, proto.name)
        title = f"{role} - {proto.name}"

        if can_be_sl(fields):
            sl_rows_by_role.setdefault(role, []).append((input_path, output_path, fields))
        else:
            row = write_ml_block(ws, row, title, input_path, output_path, fields)

    for role, rows in sorted(sl_rows_by_role.items()):
        row = write_sl_group(ws, row, f"{role} - flat rows", rows)

    style_sheet(ws)


def style_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value is not None:
                cell.border = BORDER
            if cell.column == 1 and cell.value in DIRECTIVES:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 8
        for cell in ws[letter]:
            if cell.value is None:
                continue
            parts = str(cell.value).splitlines() or [""]
            width = max(width, min(65, max(len(part) for part in parts) + 2))
        ws.column_dimensions[letter].width = width


def write_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    rows = [
        ["Adventure_3 draft proto tables"],
        ["This workbook is for manual cleanup and learning."],
        ["Each family is on a separate sheet for review."],
        ["Final converter-ready delivery must be copied/merged into one first sheet named conf."],
        ["Column A is kept converter-like: only sl/ml directives in table sheets."],
        ["input paths come from each folder Path.txt plus the proto filename."],
        ["output currently mirrors input because Adventure_3 is the implemented reference package."],
        ["After manual edits, compare this draft with the edited workbook to update $proto rules."],
    ]
    for r, row in enumerate(rows, start=1):
        ws.cell(r, 1, row[0])
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 1).fill = NOTE_FILL if r == 4 else PatternFill(fill_type=None)
    ws.column_dimensions["A"].width = 100


def build() -> None:
    wb = Workbook()
    write_readme(wb)
    for folder in sorted(p for p in INPUT_ROOT.iterdir() if p.is_dir()):
        write_family_sheet(wb, folder)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


def validate() -> list[str]:
    errors: list[str] = []
    wb = load_workbook(OUTPUT_XLSX, data_only=False)
    if "README" not in wb.sheetnames:
        errors.append("README sheet is missing")

    text_values: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            text_values.extend(str(value) for value in row if value is not None)
        if ws.title == "README":
            continue
        for r in range(1, ws.max_row + 1):
            value = ws.cell(r, 1).value
            if value in (None, ""):
                continue
            if value not in DIRECTIVES:
                errors.append(f"{ws.title}: row {r} has non-directive in column A: {value!r}")
                continue
            if ws.cell(r + 1, 2).value != "input" or ws.cell(r + 1, 3).value != "output":
                errors.append(f"{ws.title}: row {r} is not followed by input/output row")
    for marker in ["????", "Рџ", "Рє", "Рё", "Р¦", "Рђ", "РЎ", "Рњ", "СЃ", "С‚", "С‹"]:
        if any(marker in value for value in text_values):
            errors.append(f"encoding marker found: {marker}")
    return errors


def main() -> None:
    build()
    errors = validate()
    if errors:
        raise SystemExit("\n".join(errors))
    wb = load_workbook(OUTPUT_XLSX, data_only=False)
    print(f"OK: {OUTPUT_XLSX.resolve()}")
    print(f"sheets={len(wb.sheetnames)} names={wb.sheetnames}")


if __name__ == "__main__":
    main()
