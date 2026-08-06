from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BRIEF = ROOT / "input" / "Adventure_master_brief.xlsx"

PACK_HEADERS = [
    "pack_no",
    "status",
    "new_prefix",
    "event_folder",
    "main_resource",
    "event_title_ru",
    "quest_group_title",
    "output_root_hint",
    "duration_text",
    "source_reference",
    "notes",
    "Последний id proto",
    "Последний id task",
]

OLD_PREFIX = "Adventure_3_Sport_9"
OLD_EVENT_FOLDER = "Adventure_3_Sport"
OLD_MAIN_RESOURCE = "Adventure_GR_3"
OLD_PET = "Pet17_Sport_9"
OLD_TROPHY = "TrophyPet17_Sport_9"


def pack_rows(ws) -> list[int]:
    rows: list[int] = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, int):
            rows.append(row)
    return rows


def pack_no_from_row(ws, row: int) -> int:
    return int(ws.cell(row, 1).value)


def event_folder_formula(prefix_cell: str) -> str:
    return (
        f'=IF({prefix_cell}="","",'
        f'IFERROR(LEFT({prefix_cell},FIND("~",SUBSTITUTE({prefix_cell},"_","~",'
        f'LEN({prefix_cell})-LEN(SUBSTITUTE({prefix_cell},"_",""))))-1),{prefix_cell}))'
    )


def output_root_formula(prefix_cell: str) -> str:
    folder = event_folder_formula(prefix_cell)[1:]
    return f'=IF({prefix_cell}="","","стандартные пути Adventure/"&({folder})&"/"&{prefix_cell})'


def set_header_style(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def simplify_pack_sheet(wb) -> None:
    ws = wb["01_Паки"]
    current_headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    if "theme_short" in current_headers:
        ws.delete_cols(current_headers.index("theme_short") + 1)

    for col, header in enumerate(PACK_HEADERS, 1):
        ws.cell(1, col).value = header

    for row in pack_rows(ws):
        ws.cell(row, 4).value = event_folder_formula(f"C{row}")
        ws.cell(row, 8).value = output_root_formula(f"C{row}")
        ws.cell(row, 12).number_format = "0"
        ws.cell(row, 13).number_format = "@"

    explanations = {
        16: ("Пояснения к 01_Паки", None),
        17: ("new_prefix", "Главный classname нового набора. Заполняется вручную и протягивается в производные поля."),
        18: ("event_folder", "Формула из new_prefix: Adventure_4_1 -> Adventure_4. Руками обычно не трогать."),
        19: ("main_resource", "Ресурс прогресса, который игрок копит в Adventure. Заполняется вручную."),
        20: ("event_title_ru", "Видимое название события или rewards quest. Заполняется вручную."),
        21: ("quest_group_title", "Title quest_group. Заполняется вручную один раз на этой странице."),
        22: ("output_root_hint", "Формула-подсказка по путям из new_prefix/event_folder."),
        23: ("duration_text", "Текстовая длительность для rule window: 2 дня, 3 дня и т.п."),
        24: ("source_reference", "Донор/шаблон, от которого собран бриф."),
        25: ("Последний id proto", "Последний занятый числовой proto id. В выходных таблицах первый новый будет +1."),
        26: ("Последний id task", "Последний занятый task identifier, например e8697. В tasks первый новый будет e8698."),
    }
    for row, (left, right) in explanations.items():
        ws.cell(row, 1).value = left
        ws.cell(row, 2).value = right

    widths = {
        "A": 10,
        "B": 14,
        "C": 20,
        "D": 20,
        "E": 20,
        "F": 26,
        "G": 26,
        "H": 48,
        "I": 14,
        "J": 34,
        "K": 28,
        "L": 18,
        "M": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws["L1"].comment = Comment("Заполняется вручную перед сборкой, если нужны новые id в proto.", "Codex")
    ws["M1"].comment = Comment("Заполняется вручную перед сборкой, если нужны новые identifier в tasks.", "Codex")
    set_header_style(ws)


def simplify_replace_sheet(wb) -> None:
    ws = wb["02_Замены"]
    for row in pack_rows(wb["01_Паки"]):
        pack_no = pack_no_from_row(wb["01_Паки"], row)
        base = 2 + (pack_no - 1) * 6
        rows = [
            ("да", OLD_PREFIX, f"='01_Паки'!C{row}", "all", "Главный prefix/classname набора"),
            ("да", OLD_EVENT_FOLDER, f"='01_Паки'!D{row}", "paths/event_frame", "Папка события"),
            ("да", OLD_MAIN_RESOURCE, f"='01_Паки'!E{row}", "quest/rewards", "Главный ресурс прогресса"),
            ("да", OLD_PET, f'="Pet17_"&\'01_Паки\'!C{row}', "pet/reward strings", "Pet classname из new_prefix"),
            ("да", OLD_TROPHY, f'="TrophyPet17_"&\'01_Паки\'!C{row}', "collection_item/trophy", "Trophy classname из new_prefix"),
            ("нет", None, None, "extra", "Дополнительная замена, если всплывёт общий classname"),
        ]
        for offset, values in enumerate(rows):
            target_row = base + offset
            ws.cell(target_row, 1).value = pack_no
            for col, value in enumerate(values, 2):
                ws.cell(target_row, col).value = value
    for col, width in {"A": 10, "B": 12, "C": 28, "D": 30, "E": 24, "F": 48}.items():
        ws.column_dimensions[col].width = width
    set_header_style(ws)


def simplify_quest_sheet(wb) -> None:
    ws = wb["04_Квесты"]
    formulas_by_key = {
        "main_quest": "&\"_quest\"",
        "quest_task_1": "&\"_quest_task_1\"",
        "quest_task_2": "&\"_quest_task_2\"",
        "quest_task_3": "&\"_quest_task_3\"",
        "rewards_quest": "&\"_rewards\"",
        "quest_group": "",
    }
    pack_sheet = wb["01_Паки"]
    for row in range(2, ws.max_row + 1):
        pack_no = ws.cell(row, 1).value
        quest_key = ws.cell(row, 2).value
        if not isinstance(pack_no, int) or quest_key not in formulas_by_key:
            continue
        pack_row = pack_no + 1
        suffix = formulas_by_key[quest_key]
        ws.cell(row, 4).value = f"='01_Паки'!C{pack_row}{suffix}"
        if quest_key == "rewards_quest":
            ws.cell(row, 5).value = f"='01_Паки'!F{pack_row}"
        elif quest_key == "quest_group":
            ws.cell(row, 5).value = f"='01_Паки'!G{pack_row}"
    for col, width in {"A": 10, "B": 18, "C": 34, "D": 34, "E": 28, "F": 48, "G": 38, "H": 88, "I": 48}.items():
        ws.column_dimensions[col].width = width
    set_header_style(ws)


def remove_unused_sheets(wb) -> None:
    for title in ["06_Экономика", "07_Выходные файлы", "08_Чеклист"]:
        if title in wb.sheetnames:
            del wb[title]


def refresh_readme(wb) -> None:
    ws = wb["README"] if "README" in wb.sheetnames else wb.create_sheet("README")
    ws.delete_rows(1, ws.max_row)
    lines = [
        "Adventure master brief: упрощённая версия",
        "",
        "Основная ручная вкладка: 01_Паки. Здесь заполняются new_prefix, main_resource, event_title_ru, quest_group_title, duration_text/source_reference при необходимости, Последний id proto и Последний id task.",
        "03_Объекты остаётся ручной рабочей вкладкой для названий, описаний, цен, subgroup и extra_json_or_notes.",
        "02_Замены и 04_Квесты получают технические identifier/replace формулами из 01_Паки.",
        "05_Награды можно заполнять автоматически по объектам и вручную править текст, если нужно.",
        "06_Экономика, 07_Выходные файлы и 08_Чеклист удалены: в текущем workflow они не участвовали в итоговой сборке.",
        "",
        "ID logic:",
        "Последний id proto = последний уже занятый числовой proto id. Достаточно заполнить один раз в любой строке pack-таблицы; генератор протянет цепочку по всем pack.",
        "Последний id task = последний уже занятый task identifier, например e8697. Достаточно заполнить один раз; в каждом quest.xlsx сначала получат id 6 обычных tasks.0.identifier, затем 7 rewards quest tasks.",
    ]
    for row, line in enumerate(lines, 1):
        ws.cell(row, 1).value = line
    ws.column_dimensions["A"].width = 130
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def simplify_brief(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    remove_unused_sheets(wb)
    simplify_pack_sheet(wb)
    simplify_replace_sheet(wb)
    simplify_quest_sheet(wb)
    refresh_readme(wb)
    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Simplify Adventure master brief workbook.")
    parser.add_argument("--brief", type=Path, default=DEFAULT_BRIEF)
    args = parser.parse_args()
    simplify_brief(args.brief)
    print(args.brief)


if __name__ == "__main__":
    main()
