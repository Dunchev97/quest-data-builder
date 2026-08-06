from __future__ import annotations

import argparse
import json
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BRIEF = ROOT / "input" / "Adventure_master_brief.xlsx"
TEMPLATE_DIR = ROOT / "skills" / "proto" / "assets" / "Adventure_template" / "split_xlsx"
OUT_ROOT = ROOT / "output"

OLD_PREFIX = "Adventure_3_Sport_9"
OLD_EVENT_FOLDER = "Adventure_3_Sport"
OLD_MAIN_RESOURCE = "Adventure_GR_3"
DIRECTIVES = {"sl", "ml", "temp_01", "temp_02"}
PROTO_ID_HEADERS = ("Последний id proto", "last_proto_id")
TASK_ID_HEADERS = ("Последний id task", "last_task_id")
REGULAR_QUEST_TASK_IDENTIFIER_COLUMNS = [
    # label, directive row, value column
    ("quest_task_1", 18, 9),
    ("quest_task_1_restarter", 27, 7),
    ("quest_task_2", 36, 9),
    ("quest_task_2_restarter", 45, 7),
    ("quest_task_3", 54, 9),
    ("quest_task_3_restarter", 63, 7),
]
PROTO_IDS_PER_ADVENTURE_PACK = 35
TASK_IDS_PER_ADVENTURE_PACK = 13

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def text(value: Any) -> str:
    return "" if value is None else str(value)


def first_value(record: dict[str, Any], keys: tuple[str, ...]) -> Any:
    for key in keys:
        value = record.get(key)
        if value not in (None, ""):
            return value
    return None


def parse_last_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    match = re.fullmatch(r"\d+", str(value).strip())
    return int(match.group(0)) if match else None


class ProtoIdAllocator:
    def __init__(self, last_id: Any, offset: int = 0):
        parsed = parse_last_int(last_id)
        self.start = parsed + offset if parsed is not None else None
        self.current = self.start

    @property
    def enabled(self) -> bool:
        return self.current is not None

    def next(self) -> int | None:
        if self.current is None:
            return None
        self.current += 1
        return self.current


class TaskIdAllocator:
    def __init__(self, last_identifier: Any, offset: int = 0):
        self.prefix = ""
        self.width = 0
        self.current: int | None = None
        self.start: str | None = None
        if last_identifier in (None, ""):
            return
        match = re.fullmatch(r"([A-Za-z]*)(\d+)", str(last_identifier).strip())
        if not match:
            return
        self.prefix = match.group(1)
        numeric = match.group(2)
        self.width = len(numeric)
        self.current = int(numeric) + offset
        self.start = f"{self.prefix}{self.current:0{self.width}d}"

    @property
    def enabled(self) -> bool:
        return self.current is not None

    def next(self) -> str | None:
        if self.current is None:
            return None
        self.current += 1
        return f"{self.prefix}{self.current:0{self.width}d}"


def compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def headers(ws) -> list[str]:
    return [text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]


def normalize_pack_no(value: Any) -> int | None:
    value = clean(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and value.isdigit():
        return int(value)
    return None


def all_pack_records(ws) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    cols = headers(ws)
    for row in range(2, ws.max_row + 1):
        pack_no = normalize_pack_no(ws.cell(row, 1).value)
        if pack_no is None:
            continue
        record = {cols[c - 1]: clean(ws.cell(row, c).value) for c in range(1, len(cols) + 1)}
        record["pack_no"] = pack_no
        record["_row"] = row
        result.append(record)
    return result


def records(ws, pack_no: int) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    cols = headers(ws)
    for row in range(2, ws.max_row + 1):
        first = clean(ws.cell(row, 1).value)
        if first != pack_no:
            continue
        record = {cols[c - 1]: clean(ws.cell(row, c).value) for c in range(1, len(cols) + 1)}
        record["_row"] = row
        result.append(record)
    return result


def first_record(ws, pack_no: int) -> dict[str, Any]:
    found = records(ws, pack_no)
    if not found:
        raise ValueError(f"Pack {pack_no} not found on sheet {ws.title}")
    return found[0]


def pack_index(pack_rows: list[dict[str, Any]], pack_no: int) -> int:
    pack_numbers = sorted(int(row["pack_no"]) for row in pack_rows)
    if pack_no in pack_numbers:
        return pack_numbers.index(pack_no)
    return max(pack_no - 1, 0)


def resolve_shared_baseline(
    pack_rows: list[dict[str, Any]],
    pack_no: int,
    keys: tuple[str, ...],
    ids_per_pack: int,
) -> tuple[Any, int]:
    filled: list[tuple[int, int, Any]] = []
    sorted_rows = sorted(pack_rows, key=lambda row: int(row["pack_no"]))
    for index, row in enumerate(sorted_rows):
        value = first_value(row, keys)
        if value not in (None, ""):
            filled.append((index, int(row["pack_no"]), value))
    if not filled:
        return None, 0

    current_index = pack_index(sorted_rows, pack_no)
    if len(filled) == 1:
        return filled[0][2], current_index * ids_per_pack

    previous = [item for item in filled if item[0] <= current_index]
    if previous:
        baseline_index, _baseline_pack_no, value = previous[-1]
        return value, (current_index - baseline_index) * ids_per_pack

    baseline_index, _baseline_pack_no, value = filled[0]
    return value, (current_index - baseline_index) * ids_per_pack


def parse_kv(raw: Any) -> dict[str, str]:
    result: dict[str, str] = {}
    if not isinstance(raw, str):
        return result
    for part in raw.split(";"):
        part = part.strip()
        if not part or "=" not in part:
            continue
        key, value = part.split("=", 1)
        result[key.strip()] = value.strip()
    return result


def parse_jsonish(value: Any, default: Any) -> Any:
    if value is None:
        return default
    if not isinstance(value, str):
        return value
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return default


def infer_event_folder(pack: dict[str, Any], objects_by_kind: dict[str, dict[str, Any]]) -> str:
    hint = text(pack.get("output_root_hint"))
    prefix = text(pack.get("new_prefix"))
    match = re.search(r"Adventure[/\\]([^/\\]+)[/\\]" + re.escape(prefix), hint)
    if match:
        return match.group(1)

    pet_extra = parse_kv(objects_by_kind.get("pet", {}).get("extra_json_or_notes"))
    if pet_extra.get("event_frame"):
        return pet_extra["event_frame"]

    event_folder = text(pack.get("event_folder"))
    if event_folder and event_folder != prefix:
        return event_folder

    match = re.match(r"^(Adventure_\d+)_\d+$", prefix)
    if match:
        return match.group(1)
    return event_folder or prefix


def replace_all(value: str, mapping: dict[str, str]) -> str:
    result = value
    for old in sorted(mapping, key=len, reverse=True):
        new = mapping[old]
        if old and new is not None:
            result = result.replace(old, new)
    return result


def block_end(ws, start_row: int) -> int:
    row = start_row + 1
    while row <= ws.max_row:
        if ws.cell(row, 1).value in DIRECTIVES:
            return row - 1
        row += 1
    return ws.max_row


def protected_cells(ws) -> set[tuple[int, int]]:
    protected: set[tuple[int, int]] = set()
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value not in DIRECTIVES:
            continue
        key_row = row + 1
        end = block_end(ws, row)
        for col in range(1, ws.max_column + 1):
            key = ws.cell(key_row, col).value
            if key in {"input", "find"}:
                for data_row in range(row + 2, end + 1):
                    protected.add((data_row, col))
    return protected


def generic_transform(ws, mapping: dict[str, str]) -> None:
    protected = protected_cells(ws)
    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in protected:
                continue
            if isinstance(cell.value, str):
                cell.value = replace_all(cell.value, mapping)


def row_has_proto_path(ws, key_row: int, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        if ws.cell(key_row, col).value not in {"input", "output"}:
            continue
        value = ws.cell(row, col).value
        if isinstance(value, str) and value.startswith("/"):
            return True
    return False


def apply_proto_ids(ws, allocator: ProtoIdAllocator) -> None:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value not in DIRECTIVES:
            continue
        key_row = row + 1
        end = block_end(ws, row)
        for col in range(1, ws.max_column + 1):
            if ws.cell(key_row, col).value == "id":
                for data_row in range(row + 2, end + 1):
                    cell = ws.cell(data_row, col)
                    if row_has_proto_path(ws, key_row, data_row):
                        cell.value = allocator.next() if allocator.enabled else None
                    else:
                        cell.value = None


def set_row(ws, row: int, values: dict[int, Any]) -> None:
    for col, value in values.items():
        ws.cell(row, col).value = value


def set_extra_value(ws, key: str, value: Any) -> bool:
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, 8).value == key:
            ws.cell(row, 9).value = value
            return True
    return False


def family_path(family: str, event_folder: str, classname: str) -> str:
    if family == "collection_item":
        return f"/collection_item/trophy/{classname}.proto.js"
    if family == "furniture":
        return f"/furniture/Adventure/{event_folder}/{classname}.proto.js"
    if family == "pet":
        return f"/pet/{classname}.proto.js"
    if family == "pot":
        return f"/pot/Adventure/{event_folder}/{classname}.proto.js"
    if family == "trophy":
        return f"/quest/trophy/{classname}.proto.js"
    raise ValueError(f"Unsupported family path: {family}")


def quest_family_path(family: str, event_folder: str, prefix: str, filename: str) -> str:
    if family == "quest":
        return f"/quest/marketing_action/Adventure/{event_folder}/{prefix}/{filename}"
    if family == "quest_action":
        return f"/quest_action/Adventure/{event_folder}/{prefix}/{filename}"
    if family == "quest_group":
        return f"/quest_group/marketing_action/Adventure/{event_folder}/{prefix}/{filename}"
    raise ValueError(f"Unsupported quest family path: {family}")


def reward_asset_class(reward: str) -> str | None:
    match = re.fullmatch(r"asset=([^:+]+):\d+", reward or "")
    return match.group(1) if match else None


def looks_technical(value: Any) -> bool:
    return isinstance(value, str) and bool(re.fullmatch(r"[A-Za-z0-9_]+", value))


def normalize_reward_hint(
    reward: dict[str, Any],
    objects_by_class: dict[str, dict[str, Any]],
    pet_title: str,
    warnings: list[str],
) -> str:
    hint = text(reward.get("hint"))
    reward_str = text(reward.get("reward"))
    single_asset = reward_asset_class(reward_str)
    obj = objects_by_class.get(single_asset or "")

    stale_markers = ["волейбол", "спортсмен", "чемпион"]
    if obj and obj.get("description") and any(marker in hint.lower() for marker in stale_markers):
        warnings.append(f"Reward task {reward.get('task_index')}: hint replaced from object description")
        return text(obj["description"])

    if "Волейболистка" in hint:
        warnings.append(f"Reward task {reward.get('task_index')}: pet title in hint normalized")
        hint = hint.replace("Волейболистка", pet_title)
    return hint


REWARD_AMOUNTS = [500, 1100, 1800, 2500, 3300, 4100, 5000]
SPEED_REWARD = {
    "reward": "asset=DomovoySpeedModifier_12h:1",
    "title": 'Чары "Шустрик"',
    "hint": "Увеличивает скорость передвижения домовёнка. Рекомендуется применять, если дел много, а времени мало. Действует 12 часов.",
    "congratulation": 'Чары "Шустрик"',
}


def object_hint(record: dict[str, Any]) -> str:
    return text(record.get("title_attention / hint") or record.get("description") or record.get("title"))


def make_reward(
    index: int,
    classname: str,
    reward: str,
    title: str,
    hint: str,
    congratulation: str | None = None,
) -> dict[str, Any]:
    return {
        "pack_no": None,
        "task_index": index,
        "amount": REWARD_AMOUNTS[index],
        "classname": classname,
        "reward": reward,
        "title": title,
        "hint": hint,
        "congratulation": congratulation or title,
        "identifier": None,
        "notes": "autofilled from 03_Objects by Codex",
    }


def derive_rewards(pack: dict[str, Any], objects_by_kind: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    main_resource = text(pack.get("main_resource") or objects_by_kind.get("main_resource", {}).get("new_classname"))
    furniture = objects_by_kind["furniture"]
    pot_multi = objects_by_kind["pot_multi"]
    pot_multi_magic = objects_by_kind["pot_multi_magic"]
    box_1 = objects_by_kind["box_1"]
    box_2 = objects_by_kind["box_2"]
    box_3 = objects_by_kind["box_3"]
    pet = objects_by_kind["pet"]
    pet_title = text(pet.get("title"))
    final_hint = object_hint(box_3)
    if pet_title:
        final_hint = f'{final_hint}\\nСупер-приз кроха "{pet_title}" за достижение последней отметки'

    rewards = [
        make_reward(
            0,
            main_resource,
            f"asset={furniture['new_classname']}:1",
            text(furniture.get("title")),
            object_hint(furniture),
        ),
        make_reward(1, main_resource, SPEED_REWARD["reward"], SPEED_REWARD["title"], SPEED_REWARD["hint"], SPEED_REWARD["congratulation"]),
        make_reward(2, main_resource, f"asset={box_1['new_classname']}:1", text(box_1.get("title")), object_hint(box_1)),
        make_reward(3, main_resource, f"asset={pot_multi['new_classname']}:1", text(pot_multi.get("title")), object_hint(pot_multi)),
        make_reward(4, main_resource, f"asset={box_2['new_classname']}:1", text(box_2.get("title")), object_hint(box_2)),
        make_reward(
            5,
            main_resource,
            f"asset={pot_multi_magic['new_classname']}:1",
            text(pot_multi_magic.get("title")),
            object_hint(pot_multi_magic),
        ),
        make_reward(
            6,
            main_resource,
            f"asset={box_3['new_classname']}:1+asset={pet['new_classname']}:1",
            text(box_3.get("title")),
            final_hint,
            f"{box_3.get('title')} + Кроха {pet_title}" if pet_title else text(box_3.get("title")),
        ),
    ]
    for reward in rewards:
        reward["pack_no"] = pack.get("pack_no")
    return rewards


def merge_rewards(
    pack: dict[str, Any],
    objects_by_kind: dict[str, dict[str, Any]],
    manual_rewards: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged = {reward["task_index"]: reward for reward in derive_rewards(pack, objects_by_kind)}
    for reward in manual_rewards:
        try:
            index = int(reward.get("task_index"))
        except (TypeError, ValueError):
            continue
        if index not in merged:
            continue
        for key, value in reward.items():
            if key == "_row":
                continue
            if value not in (None, ""):
                merged[index][key] = value
        merged[index]["_row"] = reward.get("_row")
    return [merged[index] for index in range(7)]


def extract_pack_data(brief: Path, pack_no: int) -> dict[str, Any]:
    wb = load_workbook(brief, data_only=True)
    pack_rows = all_pack_records(wb["01_Паки"])
    pack = first_record(wb["01_Паки"], pack_no)
    objects = records(wb["03_Объекты"], pack_no)
    quests = records(wb["04_Квесты"], pack_no)
    rewards = records(wb["05_Награды"], pack_no)
    economy = records(wb["06_Экономика"], pack_no) if "06_Экономика" in wb.sheetnames else []

    objects_by_kind = {text(r.get("entity_kind")): r for r in objects}
    quests_by_key = {text(r.get("quest_key")): r for r in quests}
    economy_by_target_field = {
        (text(r.get("target")), text(r.get("field"))): r for r in economy
    }
    event_folder = infer_event_folder(pack, objects_by_kind)

    missing: list[str] = []
    for field in ["new_prefix", "main_resource"]:
        if not pack.get(field):
            missing.append(f"01_Паки.{field}")
    for kind in ["main_resource", "pet", "collection_item", "furniture", "pot_multi", "pot_multi_magic", "trophy", "box_1", "box_2", "box_3"]:
        rec = objects_by_kind.get(kind)
        if not rec:
            missing.append(f"03_Объекты.{kind}")
            continue
        for field in ["new_classname", "title"]:
            if not rec.get(field):
                missing.append(f"03_Объекты.{kind}.{field}")
    if missing:
        raise ValueError("Missing required fields:\n" + "\n".join(missing))

    rewards = merge_rewards(pack, objects_by_kind, rewards)
    for reward in rewards:
        for field in ["amount", "classname", "reward", "title"]:
            if reward.get(field) in (None, ""):
                missing.append(f"05_Награды.row{reward.get('_row')}.{field}")
    if missing:
        raise ValueError("Missing required fields:\n" + "\n".join(missing))

    proto_baseline, proto_offset = resolve_shared_baseline(
        pack_rows,
        pack_no,
        PROTO_ID_HEADERS,
        PROTO_IDS_PER_ADVENTURE_PACK,
    )
    task_baseline, task_offset = resolve_shared_baseline(
        pack_rows,
        pack_no,
        TASK_ID_HEADERS,
        TASK_IDS_PER_ADVENTURE_PACK,
    )

    return {
        "pack": pack,
        "objects": objects,
        "objects_by_kind": objects_by_kind,
        "quests_by_key": quests_by_key,
        "rewards": rewards,
        "economy": economy_by_target_field,
        "event_folder": event_folder,
        "proto_ids": ProtoIdAllocator(proto_baseline, proto_offset),
        "task_ids": TaskIdAllocator(task_baseline, task_offset),
    }


def object_extra(record: dict[str, Any]) -> dict[str, str]:
    result = parse_kv(record.get("extra_json_or_notes"))
    price = record.get("price")
    if isinstance(price, str) and "=" in price:
        result.update(parse_kv(price))
    return result


def economy_value(data: dict[str, Any], target: str, field: str, fallback: Any) -> Any:
    rec = data["economy"].get((target, field))
    value = rec.get("new_value") if rec else None
    return fallback if value in (None, "") else value


def source_workbook(family: str):
    return load_workbook(TEMPLATE_DIR / f"{family}.xlsx", data_only=False)


def save_family(wb, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def apply_common_family(wb, mapping: dict[str, str]) -> None:
    ws = wb["conf"]
    generic_transform(ws, mapping)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_collection_item(data: dict[str, Any], mapping: dict[str, str], out_dir: Path) -> Path:
    rec = data["objects_by_kind"]["collection_item"]
    wb = source_workbook("collection_item")
    apply_common_family(wb, mapping)
    ws = wb["conf"]
    extra = parse_kv(rec.get("extra_json_or_notes"))
    one_of = extra.get("rand_reward.one_of") or ws.cell(5, 8).value
    set_row(
        ws,
        5,
        {
            3: family_path("collection_item", data["event_folder"], rec["new_classname"]),
            4: rec["new_classname"],
            5: rec["title"],
            6: rec.get("description"),
            7: "one_of",
            8: one_of,
            9: None,
        },
    )
    ws.cell(2, 2).value = f"main - {rec['new_classname']}.proto.js"
    path = out_dir / "collection_item.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def build_furniture(data: dict[str, Any], mapping: dict[str, str], out_dir: Path) -> Path:
    rec = data["objects_by_kind"]["furniture"]
    wb = source_workbook("furniture")
    apply_common_family(wb, mapping)
    ws = wb["conf"]
    extra = parse_kv(rec.get("extra_json_or_notes"))
    set_row(
        ws,
        5,
        {
            3: family_path("furniture", data["event_folder"], rec["new_classname"]),
            4: rec["new_classname"],
            5: rec["title"],
            6: rec.get("description"),
            7: rec.get("subgroup"),
            8: extra.get("inventory_section") or "cabinet",
            9: rec.get("price"),
            10: rec.get("currency"),
            11: extra.get("meta_info"),
            12: None,
        },
    )
    path = out_dir / "furniture.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def build_pet(data: dict[str, Any], mapping: dict[str, str], out_dir: Path) -> Path:
    rec = data["objects_by_kind"]["pet"]
    wb = source_workbook("pet")
    apply_common_family(wb, mapping)
    ws = wb["conf"]
    extra = object_extra(rec)

    set_row(
        ws,
        5,
        {
            3: family_path("pet", data["event_folder"], rec["new_classname"]),
            4: rec["new_classname"],
            5: rec["title"],
            6: rec.get("description"),
            7: rec.get("subgroup"),
            10: None,
        },
    )
    set_extra_value(ws, "default_name", rec.get("default_name"))
    set_extra_value(ws, "event_frame", extra.get("event_frame") or data["event_folder"])
    set_extra_value(ws, "open_price", extra.get("open_price") or economy_value(data, "pet", "open_price", ws.cell(11, 9).value))
    set_extra_value(ws, "feed_price", extra.get("feed_price") or economy_value(data, "pet", "feed_price", ws.cell(12, 9).value))
    set_extra_value(ws, "spoil_price", extra.get("spoil_price") or economy_value(data, "pet", "spoil_price", ws.cell(13, 9).value))
    if extra.get("title_feed"):
        set_extra_value(ws, "title_feed", extra["title_feed"])
    if rec.get("title_attention / hint"):
        set_extra_value(ws, "title_attention", rec.get("title_attention / hint"))
    ws.cell(2, 2).value = f"main - {rec['new_classname']}.proto.js"
    path = out_dir / "pet.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def build_pot(data: dict[str, Any], mapping: dict[str, str], out_dir: Path) -> Path:
    wb = source_workbook("pot")
    apply_common_family(wb, mapping)
    ws = wb["conf"]
    for row, kind in [(5, "pot_multi"), (6, "pot_multi_magic")]:
        rec = data["objects_by_kind"][kind]
        set_row(
            ws,
            row,
            {
                3: family_path("pot", data["event_folder"], rec["new_classname"]),
                4: rec["new_classname"],
                5: rec["title"],
                6: rec.get("description"),
                7: rec.get("subgroup"),
                8: rec.get("price"),
                9: rec.get("currency"),
                10: None,
            },
        )
    path = out_dir / "pot.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def build_trophy(data: dict[str, Any], mapping: dict[str, str], out_dir: Path) -> Path:
    rec = data["objects_by_kind"]["trophy"]
    wb = source_workbook("trophy")
    apply_common_family(wb, mapping)
    ws = wb["conf"]
    set_row(
        ws,
        5,
        {
            3: family_path("trophy", data["event_folder"], rec["new_classname"]),
            4: rec["new_classname"],
            5: rec["title"],
            6: "TrophyPet17_Sport_9",
            7: rec["new_classname"],
            8: None,
        },
    )
    ws.cell(2, 2).value = f"main - {rec['new_classname']}.proto.js"
    path = out_dir / "trophy.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def build_quest_group(data: dict[str, Any], mapping: dict[str, str], out_dir: Path) -> Path:
    wb = source_workbook("quest_group")
    apply_common_family(wb, mapping)
    ws = wb["conf"]
    prefix = data["pack"]["new_prefix"]
    quest_group = data["quests_by_key"].get("quest_group", {})
    rule_description = quest_group.get("rule_window_description")

    set_row(
        ws,
        5,
        {
            3: quest_family_path("quest_group", data["event_folder"], prefix, f"{prefix}.proto.js"),
            4: prefix,
            5: quest_group.get("title") or data["pack"].get("quest_group_title"),
            6: f"{prefix}_quest",
            7: f"{prefix}_rewards",
            10: None,
        },
    )
    if rule_description:
        current = parse_jsonish(ws.cell(8, 9).value, {})
        current["description"] = rule_description
        ws.cell(8, 9).value = compact_json(current)
    path = out_dir / "quest_group.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def build_quest_action(data: dict[str, Any], mapping: dict[str, str], out_dir: Path) -> Path:
    wb = source_workbook("quest_action")
    apply_common_family(wb, mapping)
    ws = wb["conf"]
    for row, target in [(47, "paid_reset_task_1"), (56, "paid_reset_task_2"), (65, "paid_reset_task_3")]:
        ws.cell(row, 6).value = economy_value(data, target, "open_price", ws.cell(row, 6).value)
    path = out_dir / "quest_action.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def set_regular_quest_task_identifiers(ws, data: dict[str, Any]) -> None:
    for _label, directive_row, col in REGULAR_QUEST_TASK_IDENTIFIER_COLUMNS:
        ws.cell(directive_row, col).value = "string"
        ws.cell(directive_row + 1, col).value = "tasks.0.identifier"
        ws.cell(directive_row + 2, col).value = data["task_ids"].next() if data["task_ids"].enabled else None
        ws.cell(directive_row, col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(directive_row + 1, col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(directive_row + 2, col).alignment = Alignment(vertical="top", wrap_text=True)


def build_quest(data: dict[str, Any], mapping: dict[str, str], out_dir: Path, warnings: list[str]) -> Path:
    wb = source_workbook("quest")
    apply_common_family(wb, mapping)
    ws = wb["conf"]
    prefix = data["pack"]["new_prefix"]
    rewards_quest = data["quests_by_key"].get("rewards_quest", {})

    header_title = rewards_quest.get("title")
    header_congrats = rewards_quest.get("congratulation")
    if not header_title:
        header_title = data["pack"].get("event_title_ru")
    if looks_technical(header_title):
        if header_congrats and not looks_technical(header_congrats):
            warnings.append("Rewards quest title looked technical; used rewards quest congratulation as title")
            header_title = header_congrats
            header_congrats = data["pack"].get("event_title_ru") or header_congrats
        else:
            header_title = data["pack"].get("event_title_ru") or header_title

    set_regular_quest_task_identifiers(ws, data)

    set_row(
        ws,
        91,
        {
            3: quest_family_path("quest", data["event_folder"], prefix, f"{prefix}_rewards.proto.js"),
            5: prefix,
            6: header_title,
            7: header_congrats,
            8: None,
        },
    )

    objects_by_class = {
        text(rec.get("new_classname")): rec for rec in data["objects"] if rec.get("new_classname")
    }
    pet_title = text(data["objects_by_kind"]["pet"].get("title"))
    starts = [102, 115, 127, 139, 151, 163, 175]
    for start, reward in zip(starts, data["rewards"]):
        values = {
            "type": "have_asset",
            "classname": reward.get("classname"),
            "amount": reward.get("amount"),
            "reward": reward.get("reward"),
            "title": reward.get("title"),
            "hint": normalize_reward_hint(reward, objects_by_class, pet_title, warnings),
            "congratulation": reward.get("congratulation"),
            "identifier": data["task_ids"].next() if data["task_ids"].enabled else reward.get("identifier"),
        }
        for row, key in zip(range(start, start + len(values)), values):
            ws.cell(row, 4).value = key
            ws.cell(row, 5).value = values[key]
        ws.cell(start, 3).value = quest_family_path("quest", data["event_folder"], prefix, f"{prefix}_rewards.proto.js")

    path = out_dir / "quest.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def style_generated(ws) -> None:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[3]:
        cell.fill = HEADER_FILL
    for cell in ws[4]:
        cell.fill = HEADER_FILL
    for row in ws.iter_rows():
        if row[1].value and row[0].value in (None, ""):
            row[1].fill = TITLE_FILL
            row[1].font = Font(bold=True)


def new_workbook() -> tuple[Workbook, Any]:
    wb = Workbook()
    ws = wb.active
    ws.title = "conf"
    return wb, ws


def build_quest_item(data: dict[str, Any], out_dir: Path) -> Path:
    rec = data["objects_by_kind"]["main_resource"]
    wb, ws = new_workbook()
    ws.append([None, "folder: quest_item, input base: /quest_item"])
    ws.append([None, f"main - {rec['new_classname']}.proto.js"])
    ws.append(["sl", "string", "string", "string", "string", "string", "string", "int"])
    ws.append([None, "input", "output", "classname", "title", "subgroup", "inventory_section", "id"])
    ws.append(
        [
            None,
            "/quest_item/Adventure_GR_3.proto.js",
            f"/quest_item/{rec['new_classname']}.proto.js",
            rec["new_classname"],
            rec["title"],
            rec.get("subgroup") or "quest_resource",
            parse_kv(rec.get("extra_json_or_notes")).get("inventory_section") or "constructor",
            None,
        ]
    )
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    style_generated(ws)
    path = out_dir / "quest_item.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def build_mystery_box(data: dict[str, Any], out_dir: Path) -> Path:
    wb, ws = new_workbook()
    ws.append([None, "folder: mystery_box, input base: /mystery_box/Adventure"])
    ws.append([None, "main - flat rows"])
    ws.append(["ml", "string", "string", "string", "string", "string", "int", "string", "string", "object", None, "int"])
    ws.append([None, "input", "output", "classname", "title", "description", "price", "currency", "meta_info", "rand_reward", None, "id"])
    ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=11)
    ws.merge_cells(start_row=4, start_column=10, end_row=4, end_column=11)
    for kind in ["box_1", "box_2", "box_3"]:
        rec = data["objects_by_kind"][kind]
        extra = parse_kv(rec.get("extra_json_or_notes"))
        rand_reward = parse_jsonish(extra.get("rand_reward"), {})
        ws.append(
            [
                None,
                f"/mystery_box/Adventure/{rec['reference_classname']}.proto.js",
                f"/mystery_box/Adventure/{rec['new_classname']}.proto.js",
                rec["new_classname"],
                rec["title"],
                rec.get("description"),
                rec.get("price"),
                rec.get("currency"),
                extra.get("meta_info"),
                "all",
                compact_json(rand_reward.get("all", [])),
                None,
            ]
        )
    for col, width in {"B": 48, "C": 48, "D": 24, "E": 28, "F": 68, "K": 84}.items():
        ws.column_dimensions[col].width = width
    style_generated(ws)
    path = out_dir / "mystery_box.xlsx"
    apply_proto_ids(ws, data["proto_ids"])
    save_family(wb, path)
    return path


def build_mapping(data: dict[str, Any]) -> dict[str, str]:
    mapping = {
        OLD_PREFIX: data["pack"]["new_prefix"],
        OLD_EVENT_FOLDER: data["event_folder"],
        OLD_MAIN_RESOURCE: data["pack"]["main_resource"],
    }
    for rec in data["objects"]:
        old = rec.get("reference_classname")
        new = rec.get("new_classname")
        if old and new:
            mapping[text(old)] = text(new)
    return mapping


def stale_cells(ws, stale_values: list[str]) -> list[str]:
    protected = protected_cells(ws)
    found: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in protected:
                continue
            value = cell.value
            if not isinstance(value, str):
                continue
            for stale in stale_values:
                if stale and stale in value:
                    found.append(f"{ws.title}!{cell.coordinate}: {stale}")
    return found


def validate_stale(paths: list[Path], mapping: dict[str, str]) -> list[str]:
    stale = [old for old in mapping if old not in {OLD_EVENT_FOLDER} or mapping[old] != old]
    problems: list[str] = []
    for path in paths:
        wb = load_workbook(path, data_only=False)
        ws = wb.worksheets[0]
        for problem in stale_cells(ws, stale):
            problems.append(f"{path.name}: {problem}")
    return problems


def write_report(out_dir: Path, data: dict[str, Any], files: list[Path], warnings: list[str], stale: list[str]) -> Path:
    report = out_dir / "generation_report.txt"
    lines = [
        f"pack_no: {data['pack']['pack_no']}",
        f"new_prefix: {data['pack']['new_prefix']}",
        f"event_folder_used: {data['event_folder']}",
        f"main_resource: {data['pack']['main_resource']}",
        f"proto_id_start_after: {data['proto_ids'].start if data['proto_ids'].enabled else 'blank'}",
        f"proto_id_last_used: {data['proto_ids'].current if data['proto_ids'].enabled else 'blank'}",
        f"task_id_start_after: {data['task_ids'].start if data['task_ids'].enabled else 'blank'}",
        f"task_id_last_used: {data['task_ids'].prefix + str(data['task_ids'].current).zfill(data['task_ids'].width) if data['task_ids'].enabled else 'blank'}",
        "",
        "files:",
    ]
    lines.extend(f"- {p.name}" for p in files)
    lines.append("")
    lines.append("warnings:")
    if warnings:
        lines.extend(f"- {item}" for item in warnings)
    else:
        lines.append("- none")
    lines.append("")
    lines.append("stale_check:")
    if stale:
        lines.extend(f"- {item}" for item in stale[:100])
    else:
        lines.append("- no stale target values found")
    report.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return report


def build_pack(brief: Path, pack_no: int, out_dir: Path | None = None) -> tuple[Path, list[Path], Path, list[str], list[str]]:
    data = extract_pack_data(brief, pack_no)
    prefix = data["pack"]["new_prefix"]
    out_dir = out_dir or OUT_ROOT / f"{prefix}_pack_tables"
    out_dir.mkdir(parents=True, exist_ok=True)
    mapping = build_mapping(data)
    warnings: list[str] = []

    files = [
        build_collection_item(data, mapping, out_dir),
        build_furniture(data, mapping, out_dir),
        build_pet(data, mapping, out_dir),
        build_pot(data, mapping, out_dir),
        build_quest(data, mapping, out_dir, warnings),
        build_quest_action(data, mapping, out_dir),
        build_quest_group(data, mapping, out_dir),
        build_trophy(data, mapping, out_dir),
        build_quest_item(data, out_dir),
        build_mystery_box(data, out_dir),
    ]
    stale = validate_stale(files, mapping)
    report = write_report(out_dir, data, files, warnings, stale)
    return out_dir, files, report, warnings, stale


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Adventure converter tables from Adventure_master_brief.xlsx")
    parser.add_argument("--brief", type=Path, default=DEFAULT_BRIEF)
    parser.add_argument("--pack-no", type=int, default=1)
    parser.add_argument("--out-dir", type=Path)
    args = parser.parse_args()

    out_dir, files, report, warnings, stale = build_pack(args.brief, args.pack_no, args.out_dir)
    print(out_dir)
    for path in files:
        print(path)
    print(report)
    print("warnings=" + str(len(warnings)))
    print("stale=" + str(len(stale)))


if __name__ == "__main__":
    main()
