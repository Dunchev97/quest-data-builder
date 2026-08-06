from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BRIEF = ROOT / "input" / "Adventure_master_brief.xlsx"
DEFAULT_OUTPUT = ROOT / "output" / "Adventure_4_all_packs_report.xlsx"

PACKS = range(1, 13)
FAMILIES = [
    ("quest_item", "Ресурсы"),
    ("collection_item", "Коллекции"),
    ("furniture", "Мебель"),
    ("pet", "Петы"),
    ("pot", "Горшки"),
    ("mystery_box", "Коробки"),
    ("quest", "Квесты"),
    ("quest_action", "Квест экшены"),
    ("quest_group", "Квест группы"),
    ("trophy", "Трофеи"),
]

PACK_FILL = PatternFill("solid", fgColor="1F4E78")
PACK_FONT = Font(color="FFFFFF", bold=True, size=12)
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def safe_value(value: Any) -> Any:
    return "" if value is None else value


def brief_records(brief: Path) -> dict[int, dict[str, Any]]:
    wb = load_workbook(brief, data_only=True)
    ws = wb["01_Паки"]
    headers = [str(ws.cell(1, col).value or "") for col in range(1, ws.max_column + 1)]
    records: dict[int, dict[str, Any]] = {}
    for row in range(2, ws.max_row + 1):
        pack_no = ws.cell(row, 1).value
        if isinstance(pack_no, int):
            records[pack_no] = {
                headers[col - 1]: ws.cell(row, col).value for col in range(1, len(headers) + 1)
            }
    return records


def copy_cell(src, dst) -> None:
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    if src.hyperlink:
        dst._hyperlink = copy(src.hyperlink)
    if src.comment:
        dst.comment = copy(src.comment)


def copy_conf_into_report(src_ws, dst_ws, start_row: int) -> int:
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(
            start_row=start_row + merged.min_row - 1,
            start_column=merged.min_col,
            end_row=start_row + merged.max_row - 1,
            end_column=merged.max_col,
        )

    for src_row in src_ws.iter_rows():
        for src in src_row:
            if isinstance(src, MergedCell):
                continue
            dst = dst_ws.cell(start_row + src.row - 1, src.column)
            copy_cell(src, dst)

    for idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[start_row + idx - 1].height = dim.height

    return start_row + src_ws.max_row


def apply_source_column_widths(src_ws, dst_ws) -> None:
    for col_idx in range(1, src_ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = src_ws.column_dimensions[letter].width
        if width:
            current = dst_ws.column_dimensions[letter].width or 0
            dst_ws.column_dimensions[letter].width = max(current, width)


def add_pack_separator(ws, row: int, max_col: int, text: str) -> None:
    if max_col > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.fill = PACK_FILL
    cell.font = PACK_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER
    ws.row_dimensions[row].height = 24


def build_summary(wb: Workbook, packs: dict[int, dict[str, Any]], out_root: Path) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Отчёт по Adventure_4 packs"])
    ws.append(["Источник", str(out_root)])
    ws.append([])
    headers = [
        "pack_no",
        "new_prefix",
        "theme_short",
        "event_title_ru",
        "quest_group_title",
        "folder",
        "xlsx_count",
        "notes",
    ]
    ws.append(headers)
    for cell in ws[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for pack_no in PACKS:
        record = packs.get(pack_no, {})
        folder = out_root / f"Adventure_4_{pack_no}_pack_tables"
        xlsx_count = len(list(folder.glob("*.xlsx"))) if folder.exists() else 0
        ws.append(
            [
                pack_no,
                record.get("new_prefix"),
                record.get("theme_short"),
                record.get("event_title_ru"),
                record.get("quest_group_title"),
                str(folder),
                xlsx_count,
                "в отчёт включены все family-вкладки подряд",
            ]
        )
    ws.append([])
    ws.append(["Вкладка", "Источник"])
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for family, sheet_name in FAMILIES:
        ws.append([sheet_name, f"{family}.xlsx из каждого pack"])

    widths = {
        "A": 12,
        "B": 24,
        "C": 34,
        "D": 28,
        "E": 28,
        "F": 62,
        "G": 12,
        "H": 44,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A2"].fill = NOTE_FILL


def build_family_sheet(
    wb: Workbook,
    out_root: Path,
    packs: dict[int, dict[str, Any]],
    family: str,
    sheet_name: str,
) -> None:
    ws = wb.create_sheet(sheet_name)
    current_row = 1
    max_col_seen = 1

    for pack_no in PACKS:
        folder = out_root / f"Adventure_4_{pack_no}_pack_tables"
        source = folder / f"{family}.xlsx"
        if not source.exists():
            raise FileNotFoundError(source)
        src_wb = load_workbook(source, data_only=False)
        src_ws = src_wb.worksheets[0]
        if src_ws.title != "conf":
            raise ValueError(f"{source}: first sheet should be conf, got {src_ws.title}")

        max_col_seen = max(max_col_seen, src_ws.max_column)
        record = packs.get(pack_no, {})
        prefix = record.get("new_prefix") or f"Adventure_4_{pack_no}"
        title = record.get("event_title_ru") or ""
        add_pack_separator(
            ws,
            current_row,
            src_ws.max_column,
            f"pack {pack_no} - {prefix} - {title} - source: {source.name}",
        )
        current_row += 1
        copy_conf_into_report(src_ws, ws, current_row)
        apply_source_column_widths(src_ws, ws)
        current_row += src_ws.max_row + 2

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for col_idx in range(1, max_col_seen + 1):
        letter = get_column_letter(col_idx)
        if not ws.column_dimensions[letter].width:
            ws.column_dimensions[letter].width = 16
    if ws.column_dimensions["B"].width < 42:
        ws.column_dimensions["B"].width = 42
    if ws.column_dimensions["C"].width < 42:
        ws.column_dimensions["C"].width = 42


def build_report(brief: Path, output: Path, out_root: Path) -> Path:
    packs = brief_records(brief)
    missing_packs = [pack_no for pack_no in PACKS if pack_no not in packs]
    if missing_packs:
        raise ValueError(f"Missing packs in brief: {missing_packs}")

    wb = Workbook()
    build_summary(wb, packs, out_root)
    for family, sheet_name in FAMILIES:
        build_family_sheet(wb, out_root, packs, family, sheet_name)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    reopened = load_workbook(output, data_only=False)
    expected = ["Сводка"] + [sheet_name for _, sheet_name in FAMILIES]
    assert reopened.sheetnames == expected, reopened.sheetnames
    for _, sheet_name in FAMILIES:
        ws = reopened[sheet_name]
        pack_rows = sum(
            1
            for row in range(1, ws.max_row + 1)
            if isinstance(ws.cell(row, 1).value, str) and ws.cell(row, 1).value.startswith("pack ")
        )
        assert pack_rows == 12, (sheet_name, pack_rows)
    all_text = " ".join(
        str(cell.value)
        for ws in reopened.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "????" not in all_text
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Build one Adventure report workbook from generated pack tables.")
    parser.add_argument("--brief", type=Path, default=DEFAULT_BRIEF)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--out-root", type=Path, default=ROOT / "output")
    args = parser.parse_args()

    path = build_report(args.brief, args.output, args.out_root)
    print(path)


if __name__ == "__main__":
    main()
