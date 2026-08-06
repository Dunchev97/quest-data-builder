from __future__ import annotations

import argparse
import json
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = ROOT / "input" / "TwoDaysTournament_2026_05"
OUT_ROOT = ROOT / "output" / "TwoDaysTournament_2026_05_reference_tables"
SKILL_ASSET_ROOT = ROOT / "skills" / "proto" / "assets" / "TwoDaysTournament_template"
GLOBAL_SKILL_ASSET_ROOT = Path.home() / ".codex" / "skills" / "proto" / "assets" / "TwoDaysTournament_template"
BRIEF_PATH = ROOT / "input" / "TwoDaysTournament_master_brief.xlsx"

OLD_PREFIX = "TwoDaysTournament_2026_05_01"
OLD_QUEST_GROUP_FOLDER = "2026_05_01"
OLD_MONTH_PREFIX = "TwoDaysTournament_2026_05"
NEXT_PREFIX = "TwoDaysTournament_2026_08_01"
NEXT_QUEST_GROUP_FOLDER = "2026_08_01"
NEXT_MONTH_PREFIX = "TwoDaysTournament_2026_08"

DIRECTIVES = {"sl", "ml", "temp_01", "temp_02"}

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="E2F0D9")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_proto(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def source_path(proto_path: Path) -> str:
    path_file = proto_path.parent / "Path.txt"
    if not path_file.exists():
        raise FileNotFoundError(path_file)
    base = path_file.read_text(encoding="utf-8-sig").strip().replace("\\", "/").strip("/")
    return "/" + "/".join(part for part in [base, proto_path.name] if part)


def deep_get(data: dict[str, Any], path: str, default: Any = None) -> Any:
    value: Any = data
    for part in path.split("."):
        if isinstance(value, list):
            try:
                value = value[int(part)]
            except (ValueError, IndexError):
                return default
        elif isinstance(value, dict):
            value = value.get(part, default)
        else:
            return default
    return value


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def new_conf_workbook(title: str) -> tuple[Workbook, Any]:
    wb = Workbook()
    ws = wb.active
    ws.title = "conf"
    ws.sheet_view.showGridLines = False
    ws.cell(1, 2).value = title
    ws.cell(1, 2).font = Font(bold=True)
    ws.cell(1, 2).fill = TITLE_FILL
    return wb, ws


def style_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_no in range(1, min(ws.max_row, 4) + 1):
        for cell in ws[row_no]:
            if cell.value not in (None, ""):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows():
        if row[1].value and row[0].value in (None, ""):
            row[1].fill = TITLE_FILL
            row[1].font = Font(bold=True)
    for cell in ws[2]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
    for cell in ws[3]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
    ws.freeze_panes = "A4"


def write_table(
    ws,
    title: str,
    fields: list[tuple[str, str]],
    rows: list[dict[str, Any]],
) -> None:
    ws.cell(1, 2).value = title
    directive = "ml" if any(field_type in {"object", "replace"} for field_type, _key in fields) else "sl"
    ws.cell(2, 1).value = directive
    ws.cell(2, 2).value = "string"
    ws.cell(2, 3).value = "string"
    ws.cell(3, 2).value = "input"
    ws.cell(3, 3).value = "output"

    col = 4
    field_columns: list[tuple[str, str, int]] = []
    for field_type, key in fields:
        ws.cell(2, col).value = field_type
        if field_type == "replace":
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
            ws.cell(3, col).value = "find"
            ws.cell(3, col + 1).value = "replace"
            field_columns.append((field_type, key, col))
            col += 2
        elif field_type == "object":
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
            ws.cell(3, col).value = key
            field_columns.append((field_type, key, col))
            col += 2
        else:
            ws.cell(3, col).value = key
            field_columns.append((field_type, key, col))
            col += 1

    row_no = 4
    for row_data in rows:
        object_lengths = [
            len(row_data.get(key) or {})
            for field_type, key, _col in field_columns
            if field_type == "object"
        ]
        repeat = max([1] + object_lengths)
        ws.cell(row_no, 2).value = row_data["input"]
        ws.cell(row_no, 3).value = row_data.get("output", row_data["input"])
        for field_type, key, field_col in field_columns:
            value = row_data.get(key)
            if field_type == "replace":
                find_value, replace_value = value or ("", "")
                ws.cell(row_no, field_col).value = find_value
                ws.cell(row_no, field_col + 1).value = replace_value
            elif field_type == "object":
                items = list((value or {}).items())
                for offset, (object_key, object_value) in enumerate(items):
                    ws.cell(row_no + offset, field_col).value = object_key
                    ws.cell(row_no + offset, field_col + 1).value = (
                        compact_json(object_value) if isinstance(object_value, (dict, list)) else object_value
                    )
            else:
                ws.cell(row_no, field_col).value = value
        row_no += repeat

    for column in range(2, ws.max_column + 1):
        max_len = 12
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, column).value
            if value is None:
                continue
            max_len = max(max_len, min(80, len(str(value)) + 2))
        ws.column_dimensions[ws.cell(1, column).column_letter].width = max_len
    style_sheet(ws)


def one_row(proto_path: Path, data: dict[str, Any], extra: dict[str, Any] | None = None) -> dict[str, Any]:
    row = {
        "input": source_path(proto_path),
        "output": source_path(proto_path),
        "replace": (OLD_PREFIX, OLD_PREFIX),
        "id": None,
    }
    if extra:
        row.update(extra)
    return row


def asset_rows(root: Path, folder: str, field_paths: list[str]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for proto_path in sorted((root / folder).glob("*.proto.js")):
        data = load_proto(proto_path)
        extra = {"classname": data.get("classname"), "replace": (OLD_PREFIX, OLD_PREFIX), "id": None}
        for key in field_paths:
            extra[key] = deep_get(data, key)
        result.append(one_row(proto_path, data, extra))
    return result


def build_family_workbooks(root: Path, out_dir: Path, skill_asset_root: Path) -> list[Path]:
    split_dir = skill_asset_root / "split_xlsx"
    split_dir.mkdir(parents=True, exist_ok=True)
    out_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[Path] = []

    families: list[tuple[str, list[tuple[str, str]], list[dict[str, Any]]]] = []

    qi_path = next((root / "quest_item").glob("*.proto.js"))
    qi = load_proto(qi_path)
    families.append((
        "quest_item",
        [("string", "classname"), ("string", "title"), ("string", "subgroup"), ("string", "inventory_section"), ("int", "id")],
        [one_row(qi_path, qi, {"classname": qi["classname"], "title": qi["title"], "subgroup": qi.get("subgroup"), "inventory_section": qi.get("inventory_section"), "id": None})],
    ))

    ci_path = next((root / "collection_item").glob("*.proto.js"))
    ci = load_proto(ci_path)
    families.append((
        "collection_item",
        [("string", "classname"), ("string", "title"), ("string", "description"), ("object", "rand_reward"), ("int", "id")],
        [one_row(ci_path, ci, {"classname": ci["classname"], "title": ci.get("title"), "description": ci.get("description"), "rand_reward": ci.get("rand_reward"), "id": None})],
    ))

    families.append((
        "furniture",
        [
            ("string", "classname"),
            ("string", "title"),
            ("string", "subgroup"),
            ("int", "price"),
            ("string", "currency"),
            ("string", "meta_info"),
            ("string", "extra.wdr_type"),
            ("int", "id"),
        ],
        asset_rows(root, "furniture", ["title", "subgroup", "price", "currency", "meta_info", "extra.wdr_type"]),
    ))

    pet_path = next((root / "pet").glob("*.proto.js"))
    pet = load_proto(pet_path)
    families.append((
        "pet",
        [
            ("string", "classname"),
            ("string", "title"),
            ("string", "description"),
            ("string", "subgroup"),
            ("string", "extra.default_name"),
            ("string", "extra.event_frame"),
            ("string", "extra.feed_price"),
            ("string", "extra.title_attention"),
            ("string", "extra.encyclopedia_conditions"),
            ("int", "id"),
        ],
        [one_row(pet_path, pet, {
            "classname": pet["classname"],
            "title": pet.get("title"),
            "description": pet.get("description"),
            "subgroup": pet.get("subgroup"),
            "extra.default_name": deep_get(pet, "extra.default_name"),
            "extra.event_frame": deep_get(pet, "extra.event_frame"),
            "extra.feed_price": deep_get(pet, "extra.feed_price"),
            "extra.title_attention": deep_get(pet, "extra.title_attention"),
            "extra.encyclopedia_conditions": deep_get(pet, "extra.encyclopedia_conditions"),
            "id": None,
        })],
    ))

    mb_rows: list[dict[str, Any]] = []
    for proto_path in sorted((root / "mystery_box").glob("*.proto.js")):
        data = load_proto(proto_path)
        mb_rows.append(one_row(proto_path, data, {
            "classname": data.get("classname"),
            "title": data.get("title"),
            "description": data.get("description"),
            "price": data.get("price"),
            "currency": data.get("currency"),
            "meta_info": data.get("meta_info"),
            "rand_reward": data.get("rand_reward"),
            "id": None,
        }))
    families.append((
        "mystery_box",
        [
            ("string", "classname"),
            ("string", "title"),
            ("string", "description"),
            ("int", "price"),
            ("string", "currency"),
            ("string", "meta_info"),
            ("object", "rand_reward"),
            ("int", "id"),
        ],
        mb_rows,
    ))

    comp_path = next((root / "competition").glob("*.proto.js"))
    comp = load_proto(comp_path)
    families.append((
        "competition",
        [
            ("string", "identifier"),
            ("string", "quest_task_identifier"),
            ("string", "quest_group_identifier"),
            ("string", "update_top_classname"),
            ("string", "finish_conditions"),
            ("string", "active_conditions"),
            ("int", "leaders_top_amount"),
            ("int", "leaders_top_amount_my_mail"),
            ("int", "leaders_top_amount_odnoklassniki"),
            ("string", "icon"),
            ("object", "extra"),
            ("object", "rewards"),
            ("object", "rewards_my_mail"),
            ("object", "rewards_odnoklassniki"),
            ("int", "id"),
        ],
        [one_row(comp_path, comp, {key: comp.get(key) for key in [
            "identifier", "quest_task_identifier", "quest_group_identifier", "update_top_classname",
            "finish_conditions", "active_conditions", "leaders_top_amount", "leaders_top_amount_my_mail",
            "leaders_top_amount_odnoklassniki", "icon", "extra", "rewards", "rewards_my_mail",
            "rewards_odnoklassniki"
        ]} | {"id": None})],
    ))

    magazine_path = next((root / "magazine").glob("*.proto.js"))
    magazine = load_proto(magazine_path)
    magazine_fields = [
        "identifier",
        "title",
        "view",
        "pages.0.label_title",
        "pages.0.text_fields.0.text",
        "pages.0.text_fields.1.text",
        "pages.1.label_title",
        "pages.1.text_fields.1.text",
        "pages.2.label_title",
        "pages.2.text_fields.0.text",
        "pages.2.text_fields.1.text",
        "pages.2.text_fields.2.text",
        "pages.2.text_fields.3.text",
        "pages.3.label_title",
        "pages.3.text_fields.0.text",
        "pages.3.text_fields.1.text",
        "pages.3.main_resource",
    ]
    families.append((
        "magazine",
        [("string", key) for key in magazine_fields] + [("replace", "replace"), ("int", "id")],
        [one_row(magazine_path, magazine, {key: deep_get(magazine, key) for key in magazine_fields} | {"replace": (OLD_PREFIX, OLD_PREFIX), "id": None})],
    ))

    recipe_rows: list[dict[str, Any]] = []
    for proto_path in sorted((root / "recipe").glob("*.proto.js")):
        data = load_proto(proto_path)
        recipe_rows.append(one_row(proto_path, data, {
            "identifier": data.get("identifier"),
            "reward": data.get("reward"),
            "ingredients": data.get("ingredients"),
            "id": None,
        }))
    families.append(("recipe", [("string", "identifier"), ("string", "reward"), ("string", "ingredients"), ("int", "id")], recipe_rows))

    quest_rows: list[dict[str, Any]] = []
    for proto_path in sorted((root / "quest").glob("*.proto.js")) + sorted((root / "quest" / "competition").glob("*.proto.js")):
        data = load_proto(proto_path)
        quest_rows.append(one_row(proto_path, data, {
            "identifier": data.get("identifier"),
            "conditions": data.get("conditions"),
            "group_identifier": data.get("group_identifier"),
            "progress_arrow_id": data.get("progress_arrow_id"),
            "title": data.get("title"),
            "congratulation": data.get("congratulation"),
            "tasks.0.identifier": deep_get(data, "tasks.0.identifier"),
            "tasks.0.type": deep_get(data, "tasks.0.type"),
            "tasks.0.generator": deep_get(data, "tasks.0.generator"),
            "tasks.0.reward_classname": deep_get(data, "tasks.0.reward_classname"),
            "tasks.0.classname": deep_get(data, "tasks.0.classname"),
            "tasks.0.title": deep_get(data, "tasks.0.title"),
            "tasks.0.amount": deep_get(data, "tasks.0.amount"),
            "replace": (OLD_PREFIX, OLD_PREFIX),
            "id": None,
        }))
    families.append((
        "quest",
        [
            ("string", "identifier"),
            ("string", "conditions"),
            ("string", "group_identifier"),
            ("string", "progress_arrow_id"),
            ("string", "title"),
            ("string", "congratulation"),
            ("string", "tasks.0.identifier"),
            ("string", "tasks.0.type"),
            ("string", "tasks.0.generator"),
            ("string", "tasks.0.reward_classname"),
            ("string", "tasks.0.classname"),
            ("string", "tasks.0.title"),
            ("int", "tasks.0.amount"),
            ("replace", "replace"),
            ("int", "id"),
        ],
        quest_rows,
    ))

    action_rows: list[dict[str, Any]] = []
    for proto_path in sorted((root / "quest_action").glob("*.proto.js")):
        data = load_proto(proto_path)
        action_rows.append(one_row(proto_path, data, {
            "identifier": data.get("identifier"),
            "conditions": data.get("conditions"),
            "open_price": data.get("open_price"),
            "replace": (OLD_PREFIX, OLD_PREFIX),
            "id": None,
        }))
    families.append(("quest_action", [("string", "identifier"), ("string", "conditions"), ("string", "open_price"), ("replace", "replace"), ("int", "id")], action_rows))

    group_rows: list[dict[str, Any]] = []
    for proto_path in sorted((root / "quest_group").glob("*.proto.js")):
        data = load_proto(proto_path)
        group_rows.append(one_row(proto_path, data, {
            "identifier": data.get("identifier"),
            "title": data.get("title"),
            "description": data.get("description"),
            "description_spoil": data.get("description_spoil"),
            "spoil_conditions": data.get("spoil_conditions"),
            "first_quest": data.get("first_quest"),
            "last_quest": data.get("last_quest"),
            "extra": data.get("extra"),
            "replace": (OLD_PREFIX, OLD_PREFIX),
            "id": None,
        }))
    families.append((
        "quest_group",
        [
            ("string", "identifier"),
            ("string", "title"),
            ("string", "description"),
            ("string", "description_spoil"),
            ("string", "spoil_conditions"),
            ("string", "first_quest"),
            ("string", "last_quest"),
            ("object", "extra"),
            ("replace", "replace"),
            ("int", "id"),
        ],
        group_rows,
    ))

    trophy_path = next((root / "trophy").glob("*.proto.js"))
    trophy = load_proto(trophy_path)
    families.append((
        "trophy",
        [
            ("string", "identifier"),
            ("string", "title"),
            ("string", "conditions"),
            ("string", "pass_conditions"),
            ("string", "tasks.0.identifier"),
            ("string", "tasks.0.title"),
            ("string", "tasks.0.param"),
            ("replace", "replace"),
            ("int", "id"),
        ],
        [one_row(trophy_path, trophy, {
            "identifier": trophy.get("identifier"),
            "title": trophy.get("title"),
            "conditions": trophy.get("conditions"),
            "pass_conditions": trophy.get("pass_conditions"),
            "tasks.0.identifier": deep_get(trophy, "tasks.0.identifier"),
            "tasks.0.title": deep_get(trophy, "tasks.0.title"),
            "tasks.0.param": deep_get(trophy, "tasks.0.param"),
            "replace": (OLD_PREFIX, OLD_PREFIX),
            "id": None,
        })],
    ))

    combined = Workbook()
    combined.remove(combined.active)
    for family, fields, rows in families:
        wb, ws = new_conf_workbook(f"{family} - TwoDaysTournament template")
        write_table(ws, f"{family} - TwoDaysTournament template", fields, rows)
        split_path = split_dir / f"{family}.xlsx"
        wb.save(split_path)
        outputs.append(split_path)

        combined_ws = combined.create_sheet(family[:31])
        for source_row in ws.iter_rows():
            for source_cell in source_row:
                target_cell = combined_ws.cell(source_cell.row, source_cell.column)
                target_cell.value = source_cell.value
                copy_cell_style(source_cell, target_cell)
        for merged in ws.merged_cells.ranges:
            combined_ws.merge_cells(str(merged))
        for key, dim in ws.column_dimensions.items():
            combined_ws.column_dimensions[key].width = dim.width

    readme = combined.create_sheet("README", 0)
    readme["A1"] = "TwoDaysTournament template workbook"
    readme["A2"] = f"Source: {root}"
    readme["A3"] = f"Old prefix: {OLD_PREFIX}"
    readme["A4"] = f"Suggested next prefix: {NEXT_PREFIX}"
    readme["A5"] = "Converter-ready files are in skills/proto/assets/TwoDaysTournament_template/split_xlsx/"
    readme.column_dimensions["A"].width = 120
    combined_path = out_dir / "TwoDaysTournament_2026_05_proto_tables.xlsx"
    combined.save(combined_path)
    outputs.append(combined_path)

    skill_asset_root.mkdir(parents=True, exist_ok=True)
    canonical_path = skill_asset_root / "TwoDaysTournament_template.xlsx"
    combined.save(canonical_path)
    outputs.append(canonical_path)
    return outputs


def build_brief(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "01_Event"
    headers = ["field", "value", "source_value", "notes"]
    ws.append(headers)
    rows = [
        ("source_prefix", OLD_PREFIX, OLD_PREFIX, "Do not edit unless source reference changes."),
        ("new_prefix", NEXT_PREFIX, OLD_PREFIX, "Main value to edit, e.g. TwoDaysTournament_2026_08_01."),
        ("source_month_prefix", OLD_MONTH_PREFIX, OLD_MONTH_PREFIX, "Prefix without final _01."),
        ("new_month_prefix", NEXT_MONTH_PREFIX, OLD_MONTH_PREFIX, "Usually derived from new_prefix without final _01."),
        ("source_quest_group_folder", OLD_QUEST_GROUP_FOLDER, OLD_QUEST_GROUP_FOLDER, "Quest group Path.txt folder."),
        ("new_quest_group_folder", NEXT_QUEST_GROUP_FOLDER, OLD_QUEST_GROUP_FOLDER, "Usually 2026_08_01 for TwoDaysTournament_2026_08_01."),
        ("event_title_ru", "Августовский турнир", "Майский турнир", "Magazine title and quest group visible title."),
        ("event_title_genitive_ru", "Августовского турнира 2026", "Майского турнира 2026", "Cup titles/meta_info."),
        ("resource_title", "Конфета", "Конфета", "Competition resource visible title."),
        ("resource_classname", f"{NEXT_PREFIX}_Resource_Competition", f"{OLD_PREFIX}_Resource_Competition", "Can be formula-derived from new_prefix."),
        ("pet_classname", f"Pet19_{NEXT_PREFIX}_1", f"Pet19_{OLD_PREFIX}_1", "Can be formula-derived from new_prefix."),
        ("trophy_classname", f"TrophyPet19_{NEXT_PREFIX}_1", f"TrophyPet19_{OLD_PREFIX}_1", "Can be formula-derived from new_prefix."),
        ("pet_title", "", "Хомяк", "Manual."),
        ("pet_description", "", "Этот малыш очень активный и весёлый питомец. Он неутомимо может играть или бегать в круге на марафонские дистанции!", "Manual."),
        ("pet_default_name", "", "Мяфлик", "Manual."),
        ("pet_title_attention", "", "Можно получить в майском турнире 2026", "Manual."),
        ("competition_open_time", "", "2026-05-25 09:00", "For activate conditions and magazine text."),
        ("competition_close_time", "", "2026-05-26 22:00", "For activate conditions and magazine text."),
        ("competition_finish_condition", "", "time<2026-05-26 22:05", "Competition finish_conditions."),
        ("event_active_until", "", "time<2026-05-31 23:59", "Competition active_conditions / tech group spoil."),
        ("tech_start_condition", "", "time>2026-05-24 09:00", "Tech1 start."),
        ("pet_encyclopedia_condition", "", "time>2026-05-22 11:00", "Pet extra.encyclopedia_conditions."),
        ("title_date_short", "", "Турнир с 25.05 по 26.05", "Competition extra."),
        ("title_date_full", "", "Турнир с 25 по 26 Мая", "Competition extra."),
        ("last_proto_id", "", "", "Fill before generating final August tables."),
        ("last_task_id", "e100217", "e9815", "Given by user. Next generated task id will be e100218."),
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    for col, width in {"A": 34, "B": 56, "C": 56, "D": 86}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    object_ws = wb.create_sheet("02_Objects_Text")
    object_ws.append(["object_key", "field", "new_value", "source_value", "notes"])
    object_rows = [
        ("resource", "title", "Конфета", "Конфета", "Competition resource title."),
        ("pet", "title", "", "Хомяк", ""),
        ("pet", "description", "", "Этот малыш очень активный и весёлый питомец. Он неутомимо может играть или бегать в круге на марафонские дистанции!", ""),
        ("pet", "default_name", "", "Мяфлик", ""),
        ("pet", "title_attention", "", "Можно получить в майском турнире 2026", ""),
        ("trophy", "title", "", "Хомяк", "Usually same as pet title."),
        ("collection_item", "title", "", "Медалька \"Хомяк\"", ""),
        ("cup_gold", "title", "", "Золотой кубок \"Майского турнира 2026\"", ""),
        ("cup_silver", "title", "", "Серебряный кубок \"Майского турнира 2026\"", ""),
        ("cup_bronze", "title", "", "Бронзовый кубок \"Майского турнира 2026\"", ""),
        ("cup_iron", "title", "", "Кубок участника \"Майского турнира 2026\"", ""),
    ]
    for row in object_rows:
        object_ws.append(row)
    for cell in object_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    for col, width in {"A": 24, "B": 24, "C": 64, "D": 64, "E": 54}.items():
        object_ws.column_dimensions[col].width = width

    text_ws = wb.create_sheet("03_Magazine_Text")
    text_ws.append(["key", "new_text", "source_text", "notes"])
    source_texts = {
        "magazine.title": "Майский турнир",
        "pages.0.label_title": "Турнир!",
        "pages.0.text_fields.0.text": "Турнир!",
        "pages.0.text_fields.1.text": "Участвуй в турнире, и выигрывай уникального кроху!\nА также: кубок, подарки, щедрый ларчик, хрюшку, жаб и другие полезности!",
        "pages.1.text_fields.1.text": "Участвуй турнире и получай достойные призы! Больше конфет — выше\nтвое место в топе, а значит и больше наград! Ты можешь получить кубки, кроху,\nхрюшки с червончиками и подарочные коробки!\n\nТурнир начнётся 25.05.2026 в 09:00 и завершится 26.05.2026 22:00",
        "pages.2.text_fields.1.text": "Выполняй задания и получай за них конфеты. После выполнения задания у тебя будет время до получения нового, чтобы сократить это время, выполни все три задания или можешь ускорить за 2 червончика. Чем больше у тебя конфет, тем выше твоё место в топе.",
        "pages.3.label_title": "Конфетный магазин",
        "pages.3.text_fields.0.text": "Конфетный магазин",
        "pages.3.text_fields.1.text": "Обменивай полученные конфеты на декор с этой страницы! Обменивай в любое время - место в топе при этом не теряется.",
    }
    for key, value in source_texts.items():
        text_ws.append([key, "", value, "Manual text if theme/resource changes."])
    for cell in text_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    for col, width in {"A": 34, "B": 88, "C": 88, "D": 54}.items():
        text_ws.column_dimensions[col].width = width

    ids_ws = wb.create_sheet("04_Task_Id_Order")
    ids_ws.append(["order", "target", "source_identifier", "new_identifier"])
    task_rows = [
        ("1", "trophy.tasks.0.identifier", "e9803", ""),
        ("2", "competition Activate tasks.0.identifier", "e9804", ""),
        ("3", "competition.quest_task_identifier + Counter tasks.0.identifier", "e9805", ""),
        ("4", "Competition Task1 tasks.0.identifier", "e9806", ""),
        ("5", "Competition Task1 restarter tasks.0.identifier", "e9807", ""),
        ("6", "Competition Task2 tasks.0.identifier", "e9808", ""),
        ("7", "Competition Task2 restarter tasks.0.identifier", "e9809", ""),
        ("8", "Competition Task3 tasks.0.identifier", "e9810", ""),
        ("9", "Competition Task3 restarter tasks.0.identifier", "e9811", ""),
        ("10", "Tech1 tasks.0.identifier", "e9812", ""),
        ("11", "Tech2 tasks.0.identifier", "e9813", ""),
        ("12", "Tech_Magazine tasks.0.identifier", "e9814", ""),
        ("13", "Tech_MagazineShopAvailable tasks.0.identifier", "e9815", ""),
    ]
    for row in task_rows:
        ids_ws.append(row)
    for cell in ids_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    for col, width in {"A": 12, "B": 70, "C": 22, "D": 22}.items():
        ids_ws.column_dimensions[col].width = width

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build TwoDaysTournament template tables and master brief.")
    parser.add_argument("--source-root", type=Path, default=SOURCE_ROOT)
    parser.add_argument("--out-root", type=Path, default=OUT_ROOT)
    parser.add_argument("--brief", type=Path, default=BRIEF_PATH)
    args = parser.parse_args()

    outputs = build_family_workbooks(args.source_root, args.out_root, SKILL_ASSET_ROOT)
    build_family_workbooks(args.source_root, args.out_root, GLOBAL_SKILL_ASSET_ROOT)
    build_brief(args.brief)
    print(args.out_root / "TwoDaysTournament_2026_05_proto_tables.xlsx")
    print(args.brief)
    for output in outputs:
        print(output)


if __name__ == "__main__":
    main()
