from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from . import build_actions_table, build_resource_table, export_csv, interactive_objects
    from .campaigns import DEFAULT_CAMPAIGNS_DIR, pack_dir
except ImportError:
    import build_actions_table
    import build_resource_table
    import export_csv
    import interactive_objects
    from campaigns import DEFAULT_CAMPAIGNS_DIR, pack_dir


MAX_SHEET_NAME_LENGTH = 31
STAGE6_WORKBOOK_NAME = "stage6_review.xlsx"

INTERACTIVE_SHEET_NAMES = {
    "chest_1": "ИНТЕРАКТИВ Chest",
    "help_1": "ИНТЕРАКТИВ HELP",
    "friend_action_1": "ИНТЕРАКТИВ Story_FriendAction",
    "story_random_recipe": "ИНТЕРАКТИВ Story_RandomRecipe",
    "exchanger": "ИНТЕРАКТИВ Exchanger",
    "mixer_1": "ИНТЕРАКТИВ Story_Mixer",
}
INTERACTIVE_ORDER = {
    "chest_1": 10,
    "help_1": 20,
    "friend_action_1": 30,
    "exchanger": 40,
    "story_random_recipe": 50,
    "mixer_1": 60,
}


@dataclass
class WorkbookSheet:
    name: str
    rows: list[list[Any]]
    source: str


def cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def safe_sheet_name(value: str) -> str:
    name = re.sub(r"[:\\/?*\[\]]+", " ", str(value or "").strip())
    name = re.sub(r"\s+", " ", name).strip("' ")
    return name or "Лист"


def unique_sheet_name(value: str, used: set[str]) -> str:
    base = safe_sheet_name(value)[:MAX_SHEET_NAME_LENGTH]
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f" {counter}"
        candidate = f"{base[: MAX_SHEET_NAME_LENGTH - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="EAF3E8")
    type_fill = PatternFill("solid", fgColor="F5F7FA")
    header_font = Font(bold=True, color="1F2933")
    section_font = Font(bold=True, color="234E52")
    thin = Side(style="thin", color="D9E2EC")
    bottom_border = Border(bottom=thin)
    widths: dict[int, int] = {}

    for row in ws.iter_rows():
        values = [cell.value for cell in row]
        non_empty = [str(value).strip() for value in values if str(value or "").strip()]
        first = non_empty[0] if non_empty else ""
        is_type_row = first in {"ml", "sl", "temp_01"}
        is_header_row = "input" in non_empty and "output" in non_empty
        is_section_row = bool(non_empty) and len(non_empty) <= 2

        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            if value:
                widths[cell.column] = min(max(widths.get(cell.column, 8), len(value) + 2), 60)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = bottom_border
            elif is_section_row:
                cell.font = section_font
                cell.fill = section_fill
                cell.border = bottom_border
            elif is_type_row:
                cell.fill = type_fill

    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = max(10, width)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def write_workbook(output_path: Path, sheets: list[WorkbookSheet]) -> dict[str, Any]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    used_names: set[str] = set()
    written_sheets: list[dict[str, Any]] = []
    for sheet in sheets:
        sheet_name = unique_sheet_name(sheet.name, used_names)
        ws = wb.create_sheet(title=sheet_name)
        max_columns = 0
        for row in sheet.rows:
            values = [cell_value(value) for value in row]
            max_columns = max(max_columns, len(values))
            ws.append(values)
        style_sheet(ws)
        written_sheets.append(
            {
                "name": sheet_name,
                "source": sheet.source,
                "rows": len(sheet.rows),
                "columns": max_columns,
            }
        )

    wb.save(output_path)
    reopened = load_workbook(output_path, data_only=False)
    expected_names = [item["name"] for item in written_sheets]
    if reopened.sheetnames != expected_names:
        raise ValueError(f"xlsx sheet order mismatch: {reopened.sheetnames} != {expected_names}")
    for ws in reopened.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str) and "????" in value:
                    raise ValueError(f"xlsx contains broken Cyrillic marker on sheet {ws.title}")

    return {"xlsx": str(output_path), "sheets": written_sheets}


def quest_rows_and_summary(filled_tasks: dict[str, Any], quest_group: dict[str, Any]) -> tuple[list[list[Any]], dict[str, int]]:
    quests = filled_tasks.get("quests", [])
    rows = export_csv.iter_csv_rows(quests, quest_group=quest_group)
    return rows, {
        "quests_found": len(quests),
        "quest_group_exported": 1,
        "quest_blocks_exported": len(quests),
        "tasks_exported": sum(len(quest.get("tasks", [])) for quest in quests),
        "rows_written": len(rows),
    }


def interactive_sheet_title(selection: dict[str, Any], template: dict[str, Any]) -> str:
    template_id = str(selection.get("template_id") or "")
    if template_id in INTERACTIVE_SHEET_NAMES:
        return INTERACTIVE_SHEET_NAMES[template_id]
    display = template.get("display_name_ru") or template_id
    return f"ИНТЕРАКТИВ {display}"


def build_interactive_sheets(
    campaign_id: str,
    pack_id: str,
    manifest_path: Path,
) -> tuple[list[WorkbookSheet], dict[str, Any]]:
    templates_data = interactive_objects.load_templates()
    manifest = interactive_objects.read_json(manifest_path)
    validation = interactive_objects.validate_manifest(manifest, templates_data)
    if validation["summary"]["errors"]:
        first = validation["errors"][0]
        raise ValueError(f"interactive_objects.json is invalid: {first['code']}: {first['message']}")

    templates = templates_data["templates"]
    selections = sorted(
        validation["objects"],
        key=lambda item: (
            INTERACTIVE_ORDER.get(str(item.get("template_id") or ""), 1000),
            str(item.get("object_id") or item.get("template_id") or ""),
        ),
    )

    sheets: list[WorkbookSheet] = []
    sheet_items: list[dict[str, Any]] = []
    for selection in selections:
        template_id = str(selection["template_id"])
        template = templates[template_id]
        rows = interactive_objects.build_rows_for_selection(campaign_id, selection)
        sheet_name = interactive_sheet_title(selection, template)
        object_id = str(selection.get("object_id") or template_id)
        sheets.append(WorkbookSheet(sheet_name, rows, f"interactive:{object_id}"))
        sheet_items.append(
            {
                "template_id": template_id,
                "object_id": object_id,
                "result_resource": interactive_objects.result_resource_classname(campaign_id, selection, template),
                "sheet": sheet_name,
                "rows": len(rows),
            }
        )

    summary = {
        "campaign_id": campaign_id,
        "pack_id": pack_id,
        "manifest": str(manifest_path),
        "selected_count": validation["summary"]["selected_count"],
        "sheets_written": sheet_items,
        "files_written": sheet_items,
    }
    return sheets, summary


def build_stage6_review_workbook(
    campaign_id: str,
    pack_id: str,
    campaigns_dir: Path = DEFAULT_CAMPAIGNS_DIR,
    output_xlsx: Path | None = None,
    interactive_manifest_path: Path | None = None,
) -> dict[str, Any]:
    target = pack_dir(campaign_id, pack_id, campaigns_dir)
    campaign_target = campaigns_dir / campaign_id
    output_path = output_xlsx or target / "review" / STAGE6_WORKBOOK_NAME

    filled_tasks = export_csv.read_json(target / "filled_tasks.json")
    quest_group = export_csv.read_json(target / "quest_group.json")
    quest_rows, quest_summary = quest_rows_and_summary(filled_tasks, quest_group)

    actions_rows, actions_summary = build_actions_table.build_actions(
        campaign_id,
        campaigns_dir=campaigns_dir,
        current_pack_id=pack_id,
    )
    resource_rows, resource_summary = build_resource_table.build_resource_table(
        campaign_id,
        campaigns_dir=campaigns_dir,
    )

    sheets = [
        WorkbookSheet("КВЕСТЫ", quest_rows, "generated_quests"),
        WorkbookSheet("ЭКШЕНЫ", actions_rows, "generated_actions"),
        WorkbookSheet("РЕСУРСЫ", resource_rows, "resource_table"),
    ]

    interactive_summary = {
        "campaign_id": campaign_id,
        "pack_id": pack_id,
        "manifest": "",
        "selected_count": 0,
        "sheets_written": [],
        "files_written": [],
    }
    if interactive_manifest_path is not None and interactive_manifest_path.exists():
        interactive_sheets, interactive_summary = build_interactive_sheets(
            campaign_id,
            pack_id,
            interactive_manifest_path,
        )
        sheets.extend(interactive_sheets)

    workbook_summary = write_workbook(output_path, sheets)
    workbook_summary["campaign_id"] = campaign_id
    workbook_summary["pack_id"] = pack_id
    workbook_summary["campaign_dir"] = str(campaign_target)
    return {
        "workbook": workbook_summary,
        "quest_summary": quest_summary,
        "actions_summary": actions_summary,
        "resource_summary": resource_summary,
        "interactive_summary": interactive_summary,
    }
