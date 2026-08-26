import json
import os
import sys
import tempfile
import time
import unittest
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src import workflow_fast
from src.workflow_context import file_sha256


STAGE3_ONE_DIALOG = """
Classname quests: Event_2026_Story_1
title_quest: Test Quest
# quest: 1
# task: 1
Task template ID: TT-001
Task template name: Диалог
Task type: action dialog
description: "Talk to the helper."
Tasks:
[empty]
congratulation: "Done."
Character: Helper
"""


def approved_stage_record(stage: str, campaign_id: str, pack_id: str, review_path: Path) -> dict[str, object]:
    review_path.parent.mkdir(parents=True, exist_ok=True)
    if not review_path.exists():
        review_path.write_text(f"stage {stage} review", encoding="utf-8")
    return {
        "approved": True,
        "campaign_id": campaign_id,
        "pack_id": pack_id,
        "review_path": str(review_path.resolve()),
        "review_sha256": file_sha256(review_path),
    }


def write_usable_review(review_path: Path, stage: int) -> None:
    review_path.parent.mkdir(parents=True, exist_ok=True)
    if stage == 5:
        content = """# Контрольный документ. Этап 5

## Quest group

Название: Проверка
Описание: Описание.
Успех: Успех.
Провал: Провал.
"""
    else:
        content = f"""# Контрольный документ. Этап {stage}

## Квесты

### 1. Проверка
"""
    review_path.write_text(content, encoding="utf-8")


class WorkflowFastTests(unittest.TestCase):
    def test_outputs_are_fresh_checks_all_outputs_against_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_path = root / "input.json"
            output_path = root / "output.csv"
            summary_path = root / "summary.json"
            input_path.write_text("input", encoding="utf-8")
            output_path.write_text("output", encoding="utf-8")
            summary_path.write_text("summary", encoding="utf-8")

            now = time.time()
            os.utime(input_path, (now - 20, now - 20))
            os.utime(output_path, (now - 10, now - 10))
            os.utime(summary_path, (now - 10, now - 10))
            self.assertTrue(workflow_fast.outputs_are_fresh([output_path, summary_path], [input_path]))

            os.utime(input_path, (now, now))
            self.assertFalse(workflow_fast.outputs_are_fresh([output_path, summary_path], [input_path]))

    def test_stage3_fast_command_writes_pack_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            campaigns_dir = Path(temp_dir) / "campaigns"
            pack_dir = campaigns_dir / "Event_2026" / "pack_001"
            pack_dir.mkdir(parents=True)
            (pack_dir / "stage3_quests.txt").write_text(STAGE3_ONE_DIALOG, encoding="utf-8")

            exit_code = workflow_fast.main(
                [
                    "stage3",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(campaigns_dir),
                ]
            )

            self.assertEqual(exit_code, 0)
            self.assertTrue((pack_dir / "quest_plan.json").exists())
            self.assertTrue((pack_dir / "quest_plan.resolved.json").exists())
            self.assertTrue((pack_dir / "review" / "stage3_review.md").exists())
            resolved = json.loads((pack_dir / "quest_plan.resolved.json").read_text(encoding="utf-8"))
            self.assertEqual(resolved["summary"]["issues"], 0)
            self.assertEqual(resolved["quests"][0]["tasks"][0]["task_template_id"], "TT-001")

    def test_status_uses_active_context_when_ids_are_omitted(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            campaigns_dir = root / "campaigns"
            context_path = root / "active_context.json"
            pack_dir = campaigns_dir / "Event_2026" / "pack_001"
            pack_dir.mkdir(parents=True)
            context_path.write_text(
                json.dumps(
                    {
                        "campaign_id": "Event_2026",
                        "pack_id": "pack_001",
                        "stage_approvals": {},
                    }
                ),
                encoding="utf-8",
            )

            exit_code = workflow_fast.main(
                [
                    "status",
                    "--campaigns-dir",
                    str(campaigns_dir),
                    "--context",
                    str(context_path),
                ]
            )

            self.assertEqual(exit_code, 1)

    def test_quest_group_fast_command_reads_choices_file(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            campaigns_dir = root / "campaigns"
            context_path = root / "active_context.json"
            pack_dir = campaigns_dir / "Event_2026" / "pack_001"
            pack_dir.mkdir(parents=True)
            (pack_dir / "filled_tasks.json").write_text(
                json.dumps(
                    {
                        "quests": [
                            {
                                "classname_quests": "Event_2026_Story_1",
                                "title_quest": "Проверка",
                                "tasks": [],
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (pack_dir / "quest_group_choices.json").write_text(
                json.dumps(
                    {
                        "title": "Проверочная группа",
                        "description": "Начало проверки.",
                        "description_complete": "Проверка завершена.",
                        "description_spoil": "Проверка не завершена.",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            context_path.write_text(
                json.dumps(
                    {
                        "campaign_id": "Event_2026",
                        "pack_id": "pack_001",
                        "stage_approvals": {
                            "4": approved_stage_record(
                                "4",
                                "Event_2026",
                                "pack_001",
                                pack_dir / "review" / "stage4_review.md",
                            )
                        },
                    }
                ),
                encoding="utf-8",
            )

            exit_code = workflow_fast.main(
                [
                    "quest-group",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(campaigns_dir),
                    "--context",
                    str(context_path),
                ]
            )

            self.assertEqual(exit_code, 0)
            self.assertTrue((pack_dir / "quest_group.json").exists())
            self.assertTrue((pack_dir / "review" / "stage5_review.md").exists())

    def test_interactive_objects_fast_command_writes_selection_preview(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            campaigns_dir = root / "campaigns"
            pack_dir = campaigns_dir / "Event_2026" / "pack_001"
            pack_dir.mkdir(parents=True)

            exit_code = workflow_fast.main(
                [
                    "interactive-objects",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(campaigns_dir),
                    "--select",
                    "chest_1",
                    "--select",
                    "help_1",
                ]
            )

            self.assertEqual(exit_code, 0)
            self.assertTrue((campaigns_dir / "Event_2026" / "interactive_objects.json").exists())
            self.assertTrue((campaigns_dir / "Event_2026" / "interactive_objects.preview.md").exists())

    def test_review_and_apply_review_commands_roundtrip_stage5(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            campaigns_dir = root / "campaigns"
            pack_dir = campaigns_dir / "Event_2026" / "pack_001"
            pack_dir.mkdir(parents=True)
            (pack_dir / "quest_group_choices.json").write_text(
                json.dumps(
                    {
                        "title": "Старый заголовок",
                        "description": "Описание.",
                        "description_complete": "Успех.",
                        "description_spoil": "Провал.",
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )

            review_code = workflow_fast.main(
                [
                    "review",
                    "--stage",
                    "5",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(campaigns_dir),
                ]
            )
            self.assertEqual(review_code, 0)
            review_path = pack_dir / "review" / "stage5_review.md"
            self.assertTrue(review_path.exists())
            review_path.write_text(
                review_path.read_text(encoding="utf-8").replace("Старый заголовок", "Новый заголовок"),
                encoding="utf-8",
            )

            apply_code = workflow_fast.main(
                [
                    "apply-review",
                    "--stage",
                    "5",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(campaigns_dir),
                ]
            )

            self.assertEqual(apply_code, 0)
            data = json.loads((pack_dir / "quest_group_choices.json").read_text(encoding="utf-8"))
            self.assertEqual(data["title"], "Новый заголовок")

    def test_review_command_preserves_manual_edits_without_force(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            campaigns_dir = Path(temp_dir) / "campaigns"
            pack_dir = campaigns_dir / "Event_2026" / "pack_001"
            pack_dir.mkdir(parents=True)
            choices_path = pack_dir / "quest_group_choices.json"
            choices_path.write_text(
                json.dumps(
                    {
                        "title": "Первый заголовок",
                        "description": "Описание.",
                        "description_complete": "Успех.",
                        "description_spoil": "Провал.",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            base_args = [
                "review",
                "--stage",
                "5",
                "--campaign",
                "Event_2026",
                "--pack",
                "pack_001",
                "--campaigns-dir",
                str(campaigns_dir),
            ]
            self.assertEqual(workflow_fast.main(base_args), 0)
            review_path = pack_dir / "review" / "stage5_review.md"
            review_path.write_text(
                review_path.read_text(encoding="utf-8").replace("Первый заголовок", "Ручная правка"),
                encoding="utf-8",
            )
            choices = json.loads(choices_path.read_text(encoding="utf-8"))
            choices["title"] = "Новый machine source"
            choices_path.write_text(json.dumps(choices, ensure_ascii=False), encoding="utf-8")

            self.assertEqual(workflow_fast.main(base_args), 0)
            self.assertIn("Ручная правка", review_path.read_text(encoding="utf-8"))
            self.assertEqual(workflow_fast.main([*base_args, "--force"]), 0)
            self.assertIn("Новый machine source", review_path.read_text(encoding="utf-8"))

    def test_approve_refuses_missing_review(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            context_path = root / "active_context.json"
            exit_code = workflow_fast.main(
                [
                    "approve",
                    "--stage",
                    "1",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(root / "campaigns"),
                    "--context",
                    str(context_path),
                ]
            )

            self.assertEqual(exit_code, 1)
            self.assertFalse(context_path.exists())

    def test_approve_refuses_out_of_order_stage(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            campaigns_dir = root / "campaigns"
            context_path = root / "active_context.json"
            review_path = campaigns_dir / "Event_2026" / "pack_001" / "review" / "stage2_review.md"
            write_usable_review(review_path, 2)

            exit_code = workflow_fast.main(
                [
                    "approve",
                    "--stage",
                    "2",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(campaigns_dir),
                    "--context",
                    str(context_path),
                ]
            )

            self.assertEqual(exit_code, 1)
            self.assertFalse(context_path.exists())

    def test_approve_applies_existing_review_before_marking_stage(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            campaigns_dir = root / "campaigns"
            context_path = root / "active_context.json"
            pack_dir = campaigns_dir / "Event_2026" / "pack_001"
            pack_dir.mkdir(parents=True)
            (pack_dir / "quest_group_choices.json").write_text(
                json.dumps(
                    {
                        "title": "Старый заголовок",
                        "description": "Описание.",
                        "description_complete": "Успех.",
                        "description_spoil": "Провал.",
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            (pack_dir / "filled_tasks.json").write_text(
                json.dumps(
                    {
                        "quests": [
                            {
                                "classname_quests": "Event_2026_Story_1",
                                "title_quest": "Проверка",
                                "tasks": [],
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            prior_approvals: dict[str, object] = {}
            for prior_stage in range(1, 5):
                prior_review_path = pack_dir / "review" / f"stage{prior_stage}_review.md"
                write_usable_review(prior_review_path, prior_stage)
                prior_approvals[str(prior_stage)] = approved_stage_record(
                    str(prior_stage),
                    "Event_2026",
                    "pack_001",
                    prior_review_path,
                )
            context_path.write_text(
                json.dumps(
                    {
                        "campaign_id": "Event_2026",
                        "pack_id": "pack_001",
                        "stage_approvals": prior_approvals,
                    }
                ),
                encoding="utf-8",
            )

            review_code = workflow_fast.main(
                [
                    "review",
                    "--stage",
                    "5",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(campaigns_dir),
                ]
            )
            self.assertEqual(review_code, 0)
            review_path = pack_dir / "review" / "stage5_review.md"
            review_path.write_text(
                review_path.read_text(encoding="utf-8").replace("Старый заголовок", "Новый заголовок"),
                encoding="utf-8",
            )

            approve_code = workflow_fast.main(
                [
                    "approve",
                    "--stage",
                    "5",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(campaigns_dir),
                    "--context",
                    str(context_path),
                ]
            )

            self.assertEqual(approve_code, 0)
            data = json.loads((pack_dir / "quest_group_choices.json").read_text(encoding="utf-8"))
            context = json.loads(context_path.read_text(encoding="utf-8"))
            quest_group = json.loads((pack_dir / "quest_group.json").read_text(encoding="utf-8"))
            self.assertEqual(data["title"], "Новый заголовок")
            self.assertEqual(quest_group["title"], "Новый заголовок")
            self.assertTrue(context["stage_approvals"]["5"]["approved"])
            self.assertEqual(context["stage_approvals"]["5"]["review_sha256"], file_sha256(review_path))

    def test_stage6_exports_review_workbook_without_csv_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            campaigns_dir = root / "campaigns"
            context_path = root / "active_context.json"
            campaign_dir = campaigns_dir / "Event_2026"
            pack_dir = campaign_dir / "pack_001"
            pack_dir.mkdir(parents=True)
            (campaign_dir / "campaign.json").write_text(
                json.dumps(
                    {
                        "version": 1,
                        "campaign_id": "Event_2026",
                        "packs": [{"pack_id": "pack_001"}],
                    }
                ),
                encoding="utf-8",
            )
            (campaign_dir / "campaign_memory.json").write_text(
                json.dumps({"version": 1, "campaign_id": "Event_2026", "packs": {}}),
                encoding="utf-8",
            )
            (pack_dir / "filled_tasks.json").write_text(
                json.dumps(
                    {
                        "quests": [
                            {
                                "classname_quests": "Event_2026_Story_1",
                                "title_quest": "Проверка",
                                "tasks": [],
                            },
                            {
                                "classname_quests": "Event_2026_Story_2",
                                "title_quest": "Вторая проверка",
                                "tasks": [],
                            },
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (pack_dir / "context_pack.json").write_text(json.dumps({"quests": []}), encoding="utf-8")
            (pack_dir / "filled_tasks.validation.json").write_text(
                json.dumps({"summary": {"errors": 0, "warnings": 0}}),
                encoding="utf-8",
            )
            (pack_dir / "quest_group.json").write_text(
                json.dumps(
                    {
                        "input": "/quest_group/event/Night_2026_Story.proto.js",
                        "output": "/quest_group/event/Event_2026_Story.proto.js",
                        "title": "Проверочная группа",
                        "id": 125028,
                        "find": "Night_2026",
                        "replace": "Event_2026",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (pack_dir / "quest_group.validation.json").write_text(
                json.dumps({"summary": {"errors": 0, "warnings": 0}}),
                encoding="utf-8",
            )
            (campaign_dir / "interactive_objects.json").write_text(
                json.dumps(
                    {
                        "version": 1,
                        "selected_objects": [
                            {"template_id": "chest_1"},
                            {"template_id": "help_1"},
                            {"template_id": "exchanger"},
                            {
                                "template_id": "friend_action_1",
                                "action_start_time": "2026-01-01 00:00:00",
                                "action_end_time": "2026-12-31 23:59:59",
                            },
                            {"template_id": "story_random_recipe"},
                            {"template_id": "mixer_1"},
                        ],
                    }
                ),
                encoding="utf-8",
            )
            for stale_csv in (
                pack_dir / "generated_quests.csv",
                pack_dir / "generated_actions.csv",
                campaign_dir / "resource_table.csv",
                campaign_dir / "generated_interactive_objects_chest_1.csv",
                campaign_dir / "generated_interactive_objects_help_1.csv",
                campaign_dir / "generated_interactive_objects_exchanger.csv",
                campaign_dir / "generated_interactive_objects_story_friendaction_1.csv",
                campaign_dir / "generated_interactive_objects_story_randomrecipe_1.csv",
                campaign_dir / "generated_interactive_objects_mixer_1.csv",
            ):
                stale_csv.write_text("stale", encoding="utf-8")
            for stage in range(1, 6):
                review_path = pack_dir / "review" / f"stage{stage}_review.md"
                write_usable_review(review_path, stage)
            stage5_review_path = pack_dir / "review" / "stage5_review.md"
            context_path.write_text(
                json.dumps(
                    {
                        "campaign_id": "Event_2026",
                        "pack_id": "pack_001",
                        "stage_approvals": {
                            str(stage): approved_stage_record(
                                str(stage),
                                "Event_2026",
                                "pack_001",
                                pack_dir / "review" / f"stage{stage}_review.md",
                            )
                            for stage in range(1, 6)
                        },
                    }
                ),
                encoding="utf-8",
            )

            exit_code = workflow_fast.main(
                [
                    "stage6",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(campaigns_dir),
                    "--context",
                    str(context_path),
                ]
            )

            self.assertEqual(exit_code, 0)
            workbook_path = pack_dir / "review" / "stage6_review.xlsx"
            accomplished_path = pack_dir / "quest_Event_2026_accomplished.xlsx"
            self.assertTrue(workbook_path.exists())
            self.assertTrue(accomplished_path.exists())
            self.assertFalse((pack_dir / "generated_quests.csv").exists())
            self.assertFalse((pack_dir / "generated_actions.csv").exists())
            self.assertFalse((campaign_dir / "generated_interactive_objects_chest_1.csv").exists())
            self.assertFalse((campaign_dir / "generated_interactive_objects_help_1.csv").exists())
            self.assertFalse((campaign_dir / "generated_interactive_objects_exchanger.csv").exists())
            self.assertFalse((campaign_dir / "generated_interactive_objects_story_friendaction_1.csv").exists())
            self.assertFalse((campaign_dir / "generated_interactive_objects_story_randomrecipe_1.csv").exists())
            self.assertFalse((campaign_dir / "generated_interactive_objects_mixer_1.csv").exists())
            self.assertFalse((campaign_dir / "resource_table.csv").exists())
            self.assertTrue((pack_dir / "generated_actions.summary.json").exists())
            self.assertTrue((campaign_dir / "generated_interactive_objects.summary.json").exists())
            self.assertTrue((campaign_dir / "resource_table.summary.json").exists())
            from openpyxl import load_workbook

            workbook = load_workbook(workbook_path, data_only=False)
            accomplished = load_workbook(accomplished_path, data_only=False)
            self.assertEqual(accomplished.sheetnames, ["conf"])
            accomplished_sheet = accomplished["conf"]
            self.assertIsNone(accomplished_sheet.freeze_panes)
            self.assertEqual(accomplished_sheet.max_row, 5)
            self.assertEqual(accomplished_sheet.cell(1, 2).value, "quest_Event_2026_Story_accomplished")
            self.assertEqual(accomplished_sheet.cell(2, 1).value, "sl")
            self.assertEqual(
                [accomplished_sheet.cell(row, 3).value for row in (4, 5)],
                [
                    "/post_action/quest_Event_2026_Story_1_accomplished.proto.js",
                    "/post_action/quest_Event_2026_Story_2_accomplished.proto.js",
                ],
            )
            self.assertTrue(all(accomplished_sheet.cell(row, 13).value is None for row in (4, 5)))
            self.assertTrue(all(isinstance(accomplished_sheet.cell(row, 11).value, int) for row in (4, 5)))
            self.assertTrue(all(isinstance(accomplished_sheet.cell(row, 12).value, int) for row in (4, 5)))
            self.assertEqual(
                workbook.sheetnames,
                [
                    "КВЕСТЫ",
                    "ЭКШЕНЫ",
                    "РЕСУРСЫ",
                    "ИНТЕРАКТИВ Chest",
                    "ИНТЕРАКТИВ HELP",
                    "ИНТЕРАКТИВ Exchanger",
                    "ИНТЕРАКТИВ Story_FriendAction",
                    "ИНТЕРАКТИВ Story_RandomRecipe",
                    "ИНТЕРАКТИВ Mixer",
                ],
            )
            self.assertEqual(workbook["КВЕСТЫ"].freeze_panes, "A4")
            self.assertEqual(workbook["ЭКШЕНЫ"].freeze_panes, "A4")
            self.assertEqual(workbook["РЕСУРСЫ"].freeze_panes, "A76")
            self.assertEqual(workbook["ИНТЕРАКТИВ Story_RandomRecipe"].freeze_panes, "A22")
            self.assertTrue(
                all(
                    sheet.freeze_panes is None
                    for sheet in workbook.worksheets[3:]
                    if sheet.title != "ИНТЕРАКТИВ Story_RandomRecipe"
                )
            )
            quest_values = [str(value) for row in workbook["КВЕСТЫ"].iter_rows(values_only=True) for value in row if value]
            self.assertIn("Проверочная группа", quest_values)

    def test_stage6_refuses_incomplete_review_set(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            pack_dir = root / "campaigns" / "Event_2026" / "pack_001"
            pack_dir.mkdir(parents=True)
            exit_code = workflow_fast.main(
                [
                    "stage6",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(root / "campaigns"),
                    "--context",
                    str(root / "active_context.json"),
                ]
            )
            self.assertEqual(exit_code, 1)

    def test_stage6_preserves_existing_workbook_without_force(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            pack_dir = root / "campaigns" / "Event_2026" / "pack_001"
            review_dir = pack_dir / "review"
            review_dir.mkdir(parents=True)
            for stage in range(1, 6):
                write_usable_review(review_dir / f"stage{stage}_review.md", stage)
            workbook_path = review_dir / "stage6_review.xlsx"
            workbook_path.write_bytes(b"manual workbook edits")

            exit_code = workflow_fast.main(
                [
                    "stage6",
                    "--campaign",
                    "Event_2026",
                    "--pack",
                    "pack_001",
                    "--campaigns-dir",
                    str(root / "campaigns"),
                    "--context",
                    str(root / "active_context.json"),
                ]
            )
            self.assertEqual(exit_code, 1)
            self.assertEqual(workbook_path.read_bytes(), b"manual workbook edits")


if __name__ == "__main__":
    unittest.main()
