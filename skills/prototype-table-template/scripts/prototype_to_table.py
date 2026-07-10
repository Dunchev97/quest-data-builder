from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_proto(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def rel_or_raw(path: Path, root: Path | None) -> str:
    if root:
        try:
            return path.resolve().relative_to(root.resolve()).as_posix()
        except ValueError:
            pass
    return path.as_posix()


def is_scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def cell_type(value: Any) -> str:
    if isinstance(value, int) and not isinstance(value, bool):
        return "int"
    return "string"


def json_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def write_title(ws, row: int, title: str) -> int:
    ws.cell(row, 2, title)
    ws.cell(row, 2).fill = TITLE_FILL
    ws.cell(row, 2).font = Font(bold=True)
    return row + 1


def write_top_block(ws, row: int, title: str, input_path: str, output_path: str, proto: dict[str, Any]) -> int:
    flat_items: list[tuple[str, str, Any]] = []
    for key, value in proto.items():
        if isinstance(value, list) and all(is_scalar(v) for v in value):
            flat_items.append(("array_of_values", key, value))
        elif is_scalar(value):
            flat_items.append((cell_type(value), key, value))

    row = write_title(ws, row, title)
    ws.cell(row, 1, "ml")
    ws.cell(row, 2, "string")
    ws.cell(row, 3, "string")
    for offset, (typ, _key, _value) in enumerate(flat_items, 4):
        ws.cell(row, offset, typ)
    row += 1
    ws.cell(row, 2, "input")
    ws.cell(row, 3, "output")
    for offset, (_typ, key, _value) in enumerate(flat_items, 4):
        ws.cell(row, offset, key)
    row += 1
    ws.cell(row, 2, input_path)
    ws.cell(row, 3, output_path)
    max_extra = 0
    for offset, (typ, key, value) in enumerate(flat_items, 4):
        if key == "id":
            value = None
        if typ == "array_of_values":
            values = value
            max_extra = max(max_extra, len(values) - 1)
            for i, item in enumerate(values):
                ws.cell(row + i, offset, json_value(item))
        else:
            ws.cell(row, offset, json_value(value))
    return row + max_extra + 2


def write_object_block(ws, row: int, title: str, input_path: str, output_path: str, key_path: str, obj: dict[str, Any]) -> int:
    row = write_title(ws, row, title)
    ws.cell(row, 1, "ml")
    ws.cell(row, 2, "string")
    ws.cell(row, 3, "string")
    ws.cell(row, 4, "object")
    row += 1
    ws.cell(row, 2, "input")
    ws.cell(row, 3, "output")
    ws.cell(row, 4, key_path)
    row += 1
    first = True
    for key, value in obj.items():
        if first:
            ws.cell(row, 2, input_path)
            ws.cell(row, 3, output_path)
            first = False
        ws.cell(row, 4, key)
        ws.cell(row, 5, "" if key in {"id", "identifier"} else json_value(value))
        row += 1
    if first:
        ws.cell(row, 2, input_path)
        ws.cell(row, 3, output_path)
        row += 1
    return row + 1


def walk_objects(value: Any, path: str = "") -> list[tuple[str, dict[str, Any]]]:
    blocks: list[tuple[str, dict[str, Any]]] = []
    if isinstance(value, dict):
        if path:
            blocks.append((path, value))
        for key, nested in value.items():
            nested_path = f"{path}.{key}" if path else key
            if isinstance(nested, (dict, list)):
                blocks.extend(walk_objects(nested, nested_path))
    elif isinstance(value, list):
        for i, item in enumerate(value):
            nested_path = f"{path}.{i}" if path else str(i)
            if isinstance(item, dict):
                blocks.append((nested_path, item))
                for key, nested in item.items():
                    if isinstance(nested, (dict, list)):
                        blocks.extend(walk_objects(nested, f"{nested_path}.{key}"))
    return blocks


def style(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column == 1 and cell.value in {"ml", "sl", "temp_01", "temp_02"}:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 8
        for cell in ws[letter]:
            if cell.value is not None:
                parts = str(cell.value).splitlines() or [""]
                width = max(width, min(60, max(len(part) for part in parts) + 2))
        ws.column_dimensions[letter].width = width


def build(input_path: Path, output_xlsx: Path, target_proto: str, root: Path | None) -> None:
    proto = load_proto(input_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "conf"

    donor = rel_or_raw(input_path, root)
    row = 1
    identifier = proto.get("identifier", input_path.stem)
    row = write_top_block(ws, row, f"Прототип {identifier}", donor, target_proto, proto)

    for key_path, obj in walk_objects(proto):
        if key_path in {"tasks"}:
            continue
        if "." not in key_path and key_path in proto and is_scalar(proto[key_path]):
            continue
        row = write_object_block(ws, row, key_path, donor, target_proto, key_path, obj)

    style(ws)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a radmirxan xlsx-to-json compatible template from a JSON .proto.js file.")
    parser.add_argument("input", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    parser.add_argument("--target-proto", required=True, help="Target output path to write into the table output column.")
    parser.add_argument("--root", type=Path, default=None, help="Optional root for making donor input path relative.")
    args = parser.parse_args()
    build(args.input, args.output_xlsx, args.target_proto, args.root)


if __name__ == "__main__":
    main()
