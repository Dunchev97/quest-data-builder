from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook


DIRECTIVES = {"temp_01", "temp_02", "sl", "ml"}
TASK_RE = re.compile(r"^tasks\.\d+$")


def cell(ws, row: int, col: int):
    return ws.cell(row, col).value


def validate(path: Path) -> list[str]:
    errors: list[str] = []
    wb = load_workbook(path, data_only=False)
    ws = wb.worksheets[0]

    if ws.title != "conf":
        errors.append(f"First worksheet should be 'conf', got '{ws.title}'")

    text_values: list[str] = []
    for row in ws.iter_rows(values_only=True):
        text_values.extend(str(v) for v in row if v is not None)

    for marker in ["????", "Рџ", "Рє", "Рё"]:
        if any(marker in v for v in text_values):
            errors.append(f"Encoding/mojibake marker found: {marker}")

    for r in range(1, ws.max_row + 1):
        a = cell(ws, r, 1)
        if a is not None and a != "":
            if a not in DIRECTIVES:
                errors.append(f"Row {r}: column A contains non-directive value {a!r}")
                continue

            if cell(ws, r, 2) != "string" or cell(ws, r, 3) != "string":
                errors.append(f"Row {r}: directive block must have string/string types for input/output in B/C")

            if cell(ws, r + 1, 2) != "input" or cell(ws, r + 1, 3) != "output":
                errors.append(f"Row {r}: next row must have input/output in B/C")

            for c in range(4, ws.max_column + 1):
                key = cell(ws, r + 1, c)
                typ = cell(ws, r, c)
                if key == "id":
                    first_data = cell(ws, r + 2, c)
                    if first_data not in (None, ""):
                        errors.append(f"Row {r}: id value should be blank unless supplied")

                if isinstance(key, str) and TASK_RE.match(key):
                    if a != "ml" or typ != "object":
                        errors.append(f"Row {r}: task block {key} must use ml + object")
                    first_key = cell(ws, r + 2, c)
                    if first_key in (None, ""):
                        errors.append(f"Row {r}: task block {key} has no object key/value data")

    return errors


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate a radmirxan xlsx-to-json compatible workbook.")
    parser.add_argument("xlsx", type=Path)
    args = parser.parse_args()
    errors = validate(args.xlsx)
    if errors:
        for error in errors:
            print(error)
        raise SystemExit(1)
    print(f"OK: {args.xlsx}")


if __name__ == "__main__":
    main()
