from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from posixpath import basename
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PAIR_TYPES = {"object", "array_of_objects", "replace"}
MULTILINE_TYPES = {"array", "array_of_values", "array_of_objects", "object", "replace"}
DEFAULT_SKIP_FIELDS = {"class", "group", "subgroup", "currency", "permit_sell", "tags", "meta_info", "price"}
DEFAULT_FLATTEN_FIELDS = {"special_icon"}


@dataclass
class Field:
    typ: str
    key: str
    values: list[Any]
    width: int = 1


def load_proto(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def rel_or_raw(path: Path, root: Path | None) -> str:
    if root:
        try:
            return "/" + path.resolve().relative_to(root.resolve()).as_posix()
        except ValueError:
            pass
    return path.as_posix()


def split_csv(value: str | None) -> set[str]:
    if not value:
        return set()
    return {part.strip() for part in value.split(",") if part.strip()}


def parse_replace(values: list[str]) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    for value in values:
        if "=" not in value:
            raise ValueError(f"Replace must be FIND=REPLACE, got {value!r}")
        find, replace = value.split("=", 1)
        pairs.append((find, replace))
    return pairs


def is_scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def cell_type(value: Any) -> str:
    if isinstance(value, int) and not isinstance(value, bool):
        return "int"
    return "string"


def json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def blank_if_generated(key: str, value: Any) -> Any:
    if key == "id":
        return None
    return json_value(value)


def target_stem(target_proto: str) -> str:
    name = basename(target_proto.replace("\\", "/"))
    if name.endswith(".proto.js"):
        return name[: -len(".proto.js")]
    return Path(name).stem


def flatten_scalars(prefix: str, value: dict[str, Any]) -> list[Field]:
    fields: list[Field] = []
    for key, nested in value.items():
        path = f"{prefix}.{key}"
        if is_scalar(nested):
            fields.append(Field(cell_type(nested), path, [blank_if_generated(path, nested)]))
        elif isinstance(nested, dict):
            fields.extend(flatten_scalars(path, nested))
        elif isinstance(nested, list):
            fields.append(Field("array", path, [json_value(item) for item in nested]))
    return fields


def object_pairs(value: dict[str, Any]) -> list[tuple[str, Any]]:
    pairs: list[tuple[str, Any]] = []
    for key, nested in value.items():
        pairs.append((key, blank_if_generated(key, nested)))
    return pairs


def field_from_value(key: str, value: Any, flatten_fields: set[str], compact_extra: bool) -> list[Field]:
    if is_scalar(value):
        return [Field(cell_type(value), key, [blank_if_generated(key, value)])]

    if isinstance(value, dict):
        if key == "extra" and compact_extra and "preparing_title" in value:
            preparing_title = value["preparing_title"]
            return [Field(cell_type(preparing_title), "extra.preparing_title", [json_value(preparing_title)])]
        if key in flatten_fields:
            return flatten_scalars(key, value)
        if all(is_scalar(nested) for nested in value.values()):
            return [Field("object", key, object_pairs(value), width=2)]
        return flatten_scalars(key, value)

    if isinstance(value, list):
        if all(is_scalar(item) for item in value):
            return [Field("array", key, [json_value(item) for item in value])]
        return [Field("array", key, [json_value(item) for item in value])]

    return [Field("string", key, [json_value(value)])]


def collect_fields(
    proto: dict[str, Any],
    target_proto: str,
    skip_fields: set[str],
    include_fields: set[str],
    flatten_fields: set[str],
    replace_pairs: list[tuple[str, str]],
    compact_extra: bool,
) -> list[Field]:
    fields: list[Field] = []
    stem = target_stem(target_proto)
    for key, value in proto.items():
        if include_fields and key not in include_fields:
            continue
        if key in skip_fields:
            continue
        if key in {"classname", "identifier"}:
            value = stem
        fields.extend(field_from_value(key, value, flatten_fields, compact_extra))

    for find, replace in replace_pairs:
        fields.append(Field("replace", "find", [(find, replace)], width=2))
    return fields


def choose_directive(fields: list[Field], requested: str) -> str:
    if requested != "auto":
        return requested
    if any(field.typ in MULTILINE_TYPES or len(field.values) > 1 for field in fields):
        return "ml"
    return "sl"


def write_title(ws, row: int, title: str) -> int:
    ws.cell(row, 2, title)
    ws.cell(row, 2).fill = TITLE_FILL
    ws.cell(row, 2).font = Font(bold=True)
    return row + 1


def write_field_headers(ws, row: int, fields: list[Field]) -> list[int]:
    cols: list[int] = []
    col = 4
    for field in fields:
        cols.append(col)
        ws.cell(row, col, field.typ)
        if field.width > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + field.width - 1)
            if field.typ == "replace":
                ws.cell(row + 1, col, "find")
                ws.cell(row + 1, col + 1, "replace")
            else:
                ws.cell(row + 1, col, field.key)
                ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + field.width - 1)
        else:
            ws.cell(row + 1, col, field.key)
        col += field.width
    return cols


def write_field_values(ws, row: int, col: int, field: Field) -> int:
    if field.typ in PAIR_TYPES:
        rows = field.values or [(None, None)]
        for offset, pair in enumerate(rows):
            left, right = pair
            ws.cell(row + offset, col, left)
            ws.cell(row + offset, col + 1, right)
        return len(rows)

    rows = field.values or [None]
    for offset, value in enumerate(rows):
        ws.cell(row + offset, col, value)
    return len(rows)


def write_compact_block(
    ws,
    row: int,
    title: str,
    input_path: str,
    output_path: str,
    fields: list[Field],
    directive: str,
) -> int:
    row = write_title(ws, row, title)

    ws.cell(row, 1, directive)
    ws.cell(row, 2, "string")
    ws.cell(row, 3, "string")
    cols = write_field_headers(ws, row, fields)
    row += 1

    ws.cell(row, 2, "input")
    ws.cell(row, 3, "output")
    row += 1

    ws.cell(row, 2, input_path)
    ws.cell(row, 3, output_path)
    max_rows = 1
    for col, field in zip(cols, fields):
        max_rows = max(max_rows, write_field_values(ws, row, col, field))
    return row + max_rows + 2


def auto_include_fields(proto: dict[str, Any]) -> set[str]:
    proto_class = proto.get("class")
    group = proto.get("group")

    if proto_class == "CollectionPrototype":
        return {"identifier", "title", "reward", "conditions", "special_icon", "req_assets", "id"}

    if proto_class == "AssetPrototype" and group == "collection":
        return {"classname", "title", "id"}

    if proto_class == "AssetPrototype" and group == "bush_seeds":
        return {
            "classname",
            "title",
            "description",
            "extra",
            "shop_conditions",
            "rand_reward",
            "rand_reward_in_guest",
            "id",
        }

    return set()


def style(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column == 1 and cell.value in {"ml", "sl", "temp_01", "temp_02"}:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 8
        for cell in ws[letter]:
            if cell.value is None:
                continue
            parts = str(cell.value).splitlines() or [""]
            width = max(width, min(70, max(len(part) for part in parts) + 2))
        ws.column_dimensions[letter].width = width


def build(
    input_path: Path,
    output_xlsx: Path,
    target_proto: str,
    root: Path | None,
    title: str | None,
    directive: str,
    skip_fields: set[str],
    include_fields: set[str],
    flatten_fields: set[str],
    replace_pairs: list[tuple[str, str]],
    preset: str,
) -> None:
    proto = load_proto(input_path)
    compact_extra = False
    if preset == "auto" and not include_fields:
        include_fields = auto_include_fields(proto)
        if proto.get("class") == "AssetPrototype" and proto.get("group") == "bush_seeds":
            compact_extra = True

    fields = collect_fields(
        proto=proto,
        target_proto=target_proto,
        skip_fields=skip_fields,
        include_fields=include_fields,
        flatten_fields=flatten_fields,
        replace_pairs=replace_pairs,
        compact_extra=compact_extra,
    )
    selected_directive = choose_directive(fields, directive)

    wb = Workbook()
    ws = wb.active
    ws.title = "conf"

    donor = rel_or_raw(input_path, root)
    section_title = title or str(proto.get("identifier") or proto.get("classname") or input_path.stem)
    write_compact_block(ws, 1, section_title, donor, target_proto, fields, selected_directive)

    style(ws)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(description="Create a compact radmirxan xlsx-to-json compatible proto table.")
    parser.add_argument("input", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    parser.add_argument("--target-proto", required=True, help="Target output path to write into the table output column.")
    parser.add_argument("--root", type=Path, default=None, help="Optional server/prototypes root for relative donor paths.")
    parser.add_argument("--title", default=None, help="Human section title for column B.")
    parser.add_argument("--directive", choices=["auto", "sl", "ml", "temp_01", "temp_02"], default="auto")
    parser.add_argument(
        "--skip-fields",
        default="class,group,subgroup,currency,permit_sell,tags,meta_info,price",
        help="Comma-separated top-level fields to omit.",
    )
    parser.add_argument("--include-fields", default="", help="Comma-separated top-level fields to include; empty means include all except skipped.")
    parser.add_argument("--flatten-fields", default="special_icon", help="Comma-separated object fields to flatten into dot paths.")
    parser.add_argument("--replace", action="append", default=[], help="Add a replace pair, formatted as FIND=REPLACE. Can be repeated.")
    parser.add_argument("--preset", choices=["auto", "generic"], default="auto", help="Auto applies compact known proto-family field sets.")
    args = parser.parse_args()

    build(
        input_path=args.input,
        output_xlsx=args.output_xlsx,
        target_proto=args.target_proto,
        root=args.root,
        title=args.title,
        directive=args.directive,
        skip_fields=split_csv(args.skip_fields) or DEFAULT_SKIP_FIELDS,
        include_fields=split_csv(args.include_fields),
        flatten_fields=split_csv(args.flatten_fields) or DEFAULT_FLATTEN_FIELDS,
        replace_pairs=parse_replace(args.replace),
        preset=args.preset,
    )


if __name__ == "__main__":
    main()
