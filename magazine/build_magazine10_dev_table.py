from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
OUT_PATH = ROOT / "Magazine10_dev_table.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TYPE_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="E2F0D9")
TODO_FILL = PatternFill("solid", fgColor="FFF2CC")
COPY_FILL = PatternFill("solid", fgColor="EADCF8")
NEW_FILL = PatternFill("solid", fgColor="DDEBF7")
REMOVE_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def make_wb() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    return wb


def style_range(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF", size=13)
                cell.fill = HEADER_FILL
            elif cell.row == 2:
                cell.font = Font(italic=True, color="666666")
                cell.fill = TYPE_FILL
            elif cell.row == 3:
                cell.font = Font(bold=True)
                cell.fill = TITLE_FILL
            if isinstance(cell.value, str):
                if "TODO_DEV" in cell.value or "уточнить" in cell.value.lower():
                    cell.fill = TODO_FILL
                elif cell.value.startswith("COPY"):
                    cell.fill = COPY_FILL
                elif cell.value.startswith("REMOVE"):
                    cell.fill = REMOVE_FILL


def autosize(ws, max_width: int = 70):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        best = 8
        for cell in ws[letter]:
            if cell.value is None:
                continue
            value = str(cell.value)
            longest_line = max((len(part) for part in value.splitlines()), default=0)
            best = max(best, min(max_width, longest_line + 2))
        ws.column_dimensions[letter].width = best


def write_table(wb: Workbook, sheet_name: str, title: str, columns: list[str], rows: list[dict], types: list[str] | None = None):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    ws.delete_rows(1, ws.max_row)
    ws.append([title] + [""] * (len(columns) - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.append(types or ["string"] * len(columns))
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(columns))}{max(3, ws.max_row)}"
    style_range(ws)
    autosize(ws)
    return ws


def join(items: list[str]) -> str:
    return "\n".join(items)


def page_rows() -> list[dict]:
    return [
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 0, "template": "cover", "label_title": "", "short_description": "", "label_color": "", "quest_group": "", "quest_groups": "", "view": "Magazine10Cover", "source_page_9": 0, "status": "COPY_RENAME", "content_spec": "Обложка выпуска 10", "dev_notes": "Обновить арт обложки и номер выпуска."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 1, "template": "content", "label_title": "Содержание", "short_description": "", "label_color": "content", "quest_group": "", "quest_groups": "", "view": "Magazine10Content1", "source_page_9": 1, "status": "COPY_UPDATE", "content_spec": "Оглавление, часть 1", "dev_notes": "Обновить номера страниц после удаления color/test/domovestnik."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 2, "template": "content", "label_title": "Содержание", "short_description": "", "label_color": "content", "quest_group": "", "quest_groups": "", "view": "Magazine10Content2", "source_page_9": 2, "status": "COPY_UPDATE", "content_spec": "Оглавление, часть 2", "dev_notes": "Обновить номера страниц."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 3, "template": "stone", "label_title": "Каменная сказка", "short_description": "Алмаз - камень ясности, блеска и крепкой воли.", "label_color": "stone", "quest_group": "Magazine10Stone", "quest_groups": "", "view": "", "source_page_9": 3, "status": "NEW_CONTENT", "content_spec": "Камень номера: алмаз", "dev_notes": "Заменить help_asset, piles, icon, hidden asset на Magazine10."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 4, "template": "flower", "label_title": "Лютики-цветочки", "short_description": "Цветочные губы - тропический цветок с улыбкой.", "label_color": "flower", "quest_group": "Magazine10Flower", "quest_groups": "", "view": "", "source_page_9": 4, "status": "NEW_CONTENT", "content_spec": "Цветок: Цветочные губы, психотрия возвышенная", "dev_notes": "Нужны collection_id, recipe_id, asset_seed, asset_for_sale."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 5, "template": "tree", "label_title": "В тени дерев", "short_description": "Кизил - садовый лекарь с солнечными ягодами.", "label_color": "tree", "quest_group": "Magazine10Tree", "quest_groups": "", "view": "", "source_page_9": 5, "status": "NEW_CONTENT", "content_spec": "Дерево: кизил", "dev_notes": "Reward tree asset: TODO_DEV."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 6, "template": "achieve", "label_title": "Уши, лапы, хвост", "short_description": "Окапи - лесной тихоня в полосатых чулочках.", "label_color": "achieve", "quest_group": "Magazine10Animal", "quest_groups": "", "view": "", "source_page_9": "6 + 22", "status": "NEW_TEMPLATE_MIX", "content_spec": "Животное окапи в template achieve", "dev_notes": "Старый animal template не использовать. Факты о животном положить в steps."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 7, "template": "photo", "label_title": "Все о фотографии!", "short_description": "COPY_FROM_MAGAZINE9", "label_color": "photo", "quest_group": "Magazine10Photo", "quest_groups": "", "view": "", "source_page_9": 7, "status": "COPY_RENAME", "content_spec": "Пока копия фото-страницы", "dev_notes": "Если старые объекты фото уже недоступны, заменить params."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 8, "template": "clothes", "label_title": "Шитье и рукоделие", "short_description": "Северный уют, теплые узоры и костюм для большой ярмарочной прогулки.", "label_color": "clothes", "quest_group": "", "quest_groups": "Magazine10ClothesScandiBoy, Magazine10ClothesScandiGirl", "view": "", "source_page_9": 8, "status": "NEW_INSTANCE", "content_spec": "Костюм в скандинавском стиле: мальчик + девочка", "dev_notes": "Отдельный hidden: Magazine10_ClothesScandi_Hidden."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 9, "template": "clothes", "label_title": "Шитье и рукоделие", "short_description": "Ярмарочный наряд для домовят, которые любят ленты, тесьму и кармашки для гостинцев.", "label_color": "clothes", "quest_group": "", "quest_groups": "Magazine10ClothesFairBoy, Magazine10ClothesFairGirl", "view": "", "source_page_9": 8, "status": "NEW_INSTANCE", "content_spec": "Костюм ярмарочный: мальчик + девочка", "dev_notes": "Отдельный hidden: Magazine10_ClothesFair_Hidden."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 10, "template": "clothes", "label_title": "Шитье и рукоделие", "short_description": "Подводное ателье нашило перламутровые детали, волны и пуговицы-ракушки.", "label_color": "clothes", "quest_group": "", "quest_groups": "Magazine10ClothesAtelierBoy, Magazine10ClothesAtelierGirl", "view": "", "source_page_9": 8, "status": "NEW_INSTANCE", "content_spec": "Костюм в стиле подводного ателье: мальчик + девочка", "dev_notes": "Отдельный hidden: Magazine10_ClothesAtelier_Hidden."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 11, "template": "diy", "label_title": "Советы Самоделкина", "short_description": "Сыщик начинается не с лупы, а с уютного уголка для важных снимков.", "label_color": "diy", "quest_group": "Magazine10Diy", "quest_groups": "", "view": "", "source_page_9": 9, "status": "NEW_CONTENT", "content_spec": "Фотозона \"Детективное агентство котиков\", 3 шага", "dev_notes": "В пользовательском ТЗ было \"агенство\"; рекомендовано публичное написание \"агентство\"."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 12, "template": "interior_title", "label_title": "Интерьер", "short_description": "BOHO собирает в доме тепло, узоры и вольное настроение.", "label_color": "interior_shop", "quest_group": "", "quest_groups": "", "view": "", "source_page_9": 11, "status": "NEW_CONTENT", "content_spec": "Титульная страница интерьера BOHO", "dev_notes": "Page color/shop skin можно брать из Magazine9."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 13, "template": "interior_shop", "label_title": "Интерьер", "short_description": "", "label_color": "interior_shop", "quest_group": "", "quest_groups": "", "view": "", "source_page_9": "12-14", "status": "NEW_ASSETS", "content_spec": "12 декоров в стиле BOHO", "dev_notes": "Проверить вместимость 12 assets на одной странице interior_shop."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 14, "template": "beauty_hair", "label_title": "Коса - девичья краса", "short_description": "Самые эффектные прически для самых эффектных домовяток.", "label_color": "beauty_hair", "quest_group": "", "quest_groups": "", "view": "", "source_page_9": 17, "status": "NEW_ASSETS", "content_spec": "Шегги с челкой-шторкой; Небрежный пучок; Боб-каре; Кукольные бабл-локоны", "dev_notes": "Оставить правило: покупается только прическа."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 15, "template": "beauty_hair", "label_title": "Причешись-ка!", "short_description": "Стрижка задает настроение быстрее, чем соседушка успеет спросить, где расческа.", "label_color": "beauty_hair", "quest_group": "", "quest_groups": "", "view": "", "source_page_9": 18, "status": "NEW_ASSETS", "content_spec": "Текстурный кроп; Мини-маллет; Корейский two-block; Помпадур", "dev_notes": "Оставить sex: boy."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 16, "template": "beauty_makeup", "label_title": "Уроки преображения", "short_description": "Три образа для тех, кто любит менять настроение одним взмахом кисточки.", "label_color": "beauty_makeup", "quest_group": "", "quest_groups": "", "view": "", "source_page_9": 19, "status": "NEW_ASSETS", "content_spec": "Звездная фантазия; Славянская красавица; Скандинавский рунолог; Забавный стилист", "dev_notes": "payment_service_id и asset_sale classnames: TODO_DEV."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 17, "template": "cook", "label_title": "Пальчики оближешь!", "short_description": "Запеченный ролл-пирог - когда роллы решили стать домашним угощением.", "label_color": "cook", "quest_group": "", "quest_groups": "", "view": "", "source_page_9": 20, "status": "NEW_CONTENT", "content_spec": "Блюдо: запеченный ролл-пирог", "dev_notes": "reward_classname: TODO_DEV."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 18, "template": "interview", "label_title": "В гостях...", "short_description": "Кощей Бессмертный рассказывает, почему ключи не теряются, а проходят проверку.", "label_color": "interview", "quest_group": "", "quest_groups": "", "view": "", "source_page_9": 21, "status": "NEW_TEXT", "content_spec": "Доброе сказочное интервью с Кощеем", "dev_notes": "Главный icon заменить на Кощея."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 19, "template": "achieve", "label_title": "Профессия - Кукольный модельер", "short_description": "Кукольный модельер знает, как оживить образ одним бантиком.", "label_color": "achieve", "quest_group": "Magazine10Achieve", "quest_groups": "", "view": "", "source_page_9": 22, "status": "NEW_CONTENT", "content_spec": "Профессия: кукольный модельер", "dev_notes": "Второе использование template achieve, отдельно от окапи."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 20, "template": "search", "label_title": "Расследование начинается!", "short_description": "COPY_FROM_MAGAZINE9", "label_color": "search", "quest_group": "Magazine10Search", "quest_groups": "", "view": "", "source_page_9": 23, "status": "COPY_RENAME", "content_spec": "Пока копия поиска из выпуска 9", "dev_notes": "При желании заменить сюжет поиска под детективных котиков."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 21, "template": "tiny_horoscope", "label_title": "12 подвигов героя", "short_description": "Двенадцать добрых дел для домовенка, который готов к приключениям.", "label_color": "tiny_horoscope", "quest_group": "Magazine10HeroFeats", "quest_groups": "", "view": "Magazine10HeroFeats", "source_page_9": 26, "status": "NEW_TEXT_SAME_MECHANIC", "content_spec": "12 ролей + финальный питомец лисенок Луи", "dev_notes": "Механику tiny_horoscope оставить, публичный label заменить."},
        {"input": "magazine/magazine/Magazine9.proto.js", "output": "magazine/magazine/Magazine10.proto.js", "page_number": 22, "template": "surprise", "label_title": "Суперприз", "short_description": "Фасад для дома \"Детективное агентство котиков\".", "label_color": "surprise", "quest_group": "Magazine10", "quest_groups": "", "view": "Magazine10PageSurprise", "source_page_9": 28, "status": "NEW_REWARD", "content_spec": "Главный приз - фасад", "dev_notes": "take_magazine_surprise action на Magazine10."},
    ]


def removed_rows() -> list[dict]:
    return [
        {"source_page_9": 10, "template": "color_furniture", "label_title": "Квартирный вопрос", "decision": "REMOVE", "notes": "Страницу полностью убрать из Magazine10."},
        {"source_page_9": 15, "template": "color_clothes", "label_title": "Модный приговор", "decision": "REMOVE", "notes": "Обе страницы color_clothes убрать."},
        {"source_page_9": 16, "template": "color_clothes", "label_title": "Модный приговор", "decision": "REMOVE", "notes": "Обе страницы color_clothes убрать."},
        {"source_page_9": 24, "template": "domovestnik", "label_title": "ДомоВестник", "decision": "REMOVE", "notes": "Обе страницы ДомоВестника убрать."},
        {"source_page_9": 25, "template": "domovestnik", "label_title": "ДомоВестник", "decision": "REMOVE", "notes": "Обе страницы ДомоВестника убрать."},
        {"source_page_9": 27, "template": "test", "label_title": "Тест", "decision": "REMOVE", "notes": "Полностью убрать."},
        {"source_page_9": 6, "template": "animal", "label_title": "Уши, лапы, хвост", "decision": "REPLACE", "notes": "Заменить на template achieve с окапи."},
    ]


def quest_file_rows() -> list[dict]:
    rows = [
        {"input": "magazine/quest/Magazine9_1.proto.js", "output": "magazine/quest/Magazine10_1.proto.js", "page_number": 22, "quest_identifier": "Magazine10_1", "group_identifier": "Magazine10", "title": "Журнал №10", "conditions": "level>10", "reward": "asset=Magazine10_FinalReward:1", "opens": "", "completion_flag": "Magazine10_FinalReward", "source_file": "Magazine9_1.proto.js", "status": "COPY_UPDATE", "notes": "Root quest должен проверять hidden всех рабочих страниц и финальное действие surprise."},
        {"input": "magazine/quest/Magazine9Stone_1.proto.js", "output": "magazine/quest/Magazine10Stone_1.proto.js", "page_number": 3, "quest_identifier": "Magazine10Stone_1", "group_identifier": "Magazine10Stone", "title": "Алмаз", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "", "opens": "", "completion_flag": "Magazine10_Stone_Hidden", "source_file": "Magazine9Stone_1.proto.js", "status": "NEW_CONTENT", "notes": "Нужны diamond piles и help_asset."},
        {"input": "magazine/quest/Magazine9Flower_1.proto.js", "output": "magazine/quest/Magazine10Flower_1.proto.js", "page_number": 4, "quest_identifier": "Magazine10Flower_1", "group_identifier": "Magazine10Flower", "title": "Цветочные губы", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "", "opens": "", "completion_flag": "Magazine10_Flower_Hidden", "source_file": "Magazine9Flower_1.proto.js", "status": "NEW_CONTENT", "notes": "Финальный collection condition: FlowerMagazine10_PsychotriaCondition."},
        {"input": "magazine/quest/Magazine9Tree_1.proto.js", "output": "magazine/quest/Magazine10Tree_1.proto.js", "page_number": 5, "quest_identifier": "Magazine10Tree_1", "group_identifier": "Magazine10Tree", "title": "Кизил", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "asset=Magazine10_Tree_Dogwood:1", "opens": "", "completion_flag": "Magazine10_Tree_Hidden", "source_file": "Magazine9Tree_1.proto.js", "status": "NEW_CONTENT", "notes": "Проверить classname дерева."},
        {"input": "magazine/quest/Magazine9Animal_1.proto.js", "output": "magazine/quest/Magazine10Animal_1.proto.js", "page_number": 6, "quest_identifier": "Magazine10Animal_1", "group_identifier": "Magazine10Animal", "title": "Окапи", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "asset=Magazine10_AnimalReward:1", "opens": "", "completion_flag": "Magazine10_Animal_Hidden", "source_file": "Magazine9Animal_1.proto.js", "status": "NEW_CONTENT", "notes": "Quest остается отдельной группой, страница использует template achieve."},
        {"input": "magazine/quest/Magazine9Photo_1.proto.js", "output": "magazine/quest/Magazine10Photo_1.proto.js", "page_number": 7, "quest_identifier": "Magazine10Photo_1", "group_identifier": "Magazine10Photo", "title": "Фотографии", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "asset=Magazine10_PhotoReward:1", "opens": "", "completion_flag": "Magazine10_Photo_Hidden", "source_file": "Magazine9Photo_1.proto.js", "status": "COPY_RENAME", "notes": "Пока копия 10 задач post_photo."},
    ]
    costumes = [
        ("Scandi", 8, "Костюм в скандинавском стиле", "Magazine10_ClothesScandi_Hidden"),
        ("Fair", 9, "Костюм ярмарочный", "Magazine10_ClothesFair_Hidden"),
        ("Atelier", 10, "Костюм в стиле подводного ателье", "Magazine10_ClothesAtelier_Hidden"),
    ]
    for key, page, title, hidden in costumes:
        for sex, ru, source, condition in [
            ("Boy", "мальчик", "Magazine9ClothesBoy_1.proto.js", "quest_active_or_finished=Magazine10_1+personage_sex=male"),
            ("Girl", "девочка", "Magazine9ClothesGirl_1.proto.js", "quest_active_or_finished=Magazine10_1+personage_sex=female"),
        ]:
            rows.append({
                "input": f"magazine/quest/{source}",
                "output": f"magazine/quest/Magazine10Clothes{key}{sex}_1.proto.js",
                "page_number": page,
                "quest_identifier": f"Magazine10Clothes{key}{sex}_1",
                "group_identifier": f"Magazine10Clothes{key}{sex}",
                "title": f"{title} ({ru})",
                "conditions": condition,
                "reward": "",
                "opens": "",
                "completion_flag": hidden,
                "source_file": source,
                "status": "NEW_INSTANCE",
                "notes": "Сохраняет механику покупки/получения частей костюма из Magazine9.",
            })
    rows.extend([
        {"input": "magazine/quest/Magazine9Diy_1.proto.js", "output": "magazine/quest/Magazine10Diy_1.proto.js", "page_number": 11, "quest_identifier": "Magazine10Diy_1", "group_identifier": "Magazine10Diy", "title": "Детективное агентство котиков 1", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "", "opens": "Magazine10Diy_2", "completion_flag": "", "source_file": "Magazine9Diy_1.proto.js", "status": "NEW_CONTENT", "notes": "Шаг 1: установка основы фотозоны."},
        {"input": "magazine/quest/Magazine9Diy_2.proto.js", "output": "magazine/quest/Magazine10Diy_2.proto.js", "page_number": 11, "quest_identifier": "Magazine10Diy_2", "group_identifier": "Magazine10Diy", "title": "Детективное агентство котиков 2", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "", "opens": "Magazine10Diy_3", "completion_flag": "", "source_file": "Magazine9Diy_2.proto.js", "status": "NEW_CONTENT", "notes": "Шаг 2: сборка декора."},
        {"input": "magazine/quest/Magazine9Diy_3.proto.js", "output": "magazine/quest/Magazine10Diy_3.proto.js", "page_number": 11, "quest_identifier": "Magazine10Diy_3", "group_identifier": "Magazine10Diy", "title": "Детективное агентство котиков 3", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "asset=Magazine10_DiyReward:1", "opens": "", "completion_flag": "Magazine10_Diy_Hidden", "source_file": "Magazine9Diy_3.proto.js", "status": "NEW_CONTENT", "notes": "Финал выдает фотозону."},
        {"input": "magazine/quest/Magazine9Achieve_1.proto.js", "output": "magazine/quest/Magazine10Achieve_1.proto.js", "page_number": 19, "quest_identifier": "Magazine10Achieve_1", "group_identifier": "Magazine10Achieve", "title": "Кукольный модельер", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "asset=Magazine10_AchieveReward:1", "opens": "", "completion_flag": "Magazine10_Achieve_Hidden", "source_file": "Magazine9Achieve_1.proto.js", "status": "NEW_CONTENT", "notes": "Профессия через template achieve."},
    ])
    for i in range(1, 6):
        rows.append({
            "input": f"magazine/quest/Magazine9Search_{i}.proto.js",
            "output": f"magazine/quest/Magazine10Search_{i}.proto.js",
            "page_number": 20,
            "quest_identifier": f"Magazine10Search_{i}",
            "group_identifier": "Magazine10Search",
            "title": f"Подсказка №{i}",
            "conditions": "quest_active_or_finished=Magazine10_1",
            "reward": "asset=Magazine10_SearchReward:1" if i == 5 else "",
            "opens": f"Magazine10Search_{i + 1}" if i < 5 else "",
            "completion_flag": "Magazine10_Search_Hidden" if i == 5 else "",
            "source_file": f"Magazine9Search_{i}.proto.js",
            "status": "COPY_RENAME",
            "notes": "Пока копия цепочки поиска.",
        })
    rows.append({"input": "magazine/quest/Magazine9TinyHoroscope_start.proto.js", "output": "magazine/quest/Magazine10HeroFeats_start.proto.js", "page_number": 21, "quest_identifier": "Magazine10HeroFeats_start", "group_identifier": "Magazine10HeroFeats", "title": "12 подвигов героя", "conditions": "quest_active_or_finished=Magazine10_1", "reward": "", "opens": "Magazine10HeroFeats_1..12 + Magazine10HeroFeats_final", "completion_flag": "", "source_file": "Magazine9TinyHoroscope_start.proto.js", "status": "NEW_TEXT_SAME_MECHANIC", "notes": "Autostart-служебный quest."})
    roles = hero_roles()
    for idx, role in enumerate(roles, 1):
        rows.append({"input": f"magazine/quest/Magazine9TinyHoroscope_{idx}.proto.js", "output": f"magazine/quest/Magazine10HeroFeats_{idx}.proto.js", "page_number": 21, "quest_identifier": f"Magazine10HeroFeats_{idx}", "group_identifier": "Magazine10HeroFeats", "title": role["quest_title"], "conditions": "stuff=Magazine10HeroFeats_Memory:1+stuff_extra=Magazine10HeroFeats_Memory:1+quest_active_or_finished=Magazine10_1", "reward": f"asset=Magazine10_HeroFeats_Badge_{idx}:1", "opens": "", "completion_flag": "", "source_file": f"Magazine9TinyHoroscope_{idx}.proto.js", "status": "NEW_TEXT_SAME_MECHANIC", "notes": role["role"]})
    rows.append({"input": "magazine/quest/Magazine9TinyHoroscope_final.proto.js", "output": "magazine/quest/Magazine10HeroFeats_final.proto.js", "page_number": 21, "quest_identifier": "Magazine10HeroFeats_final", "group_identifier": "Magazine10HeroFeats", "title": "Лисенок Луи", "conditions": "done_quest=Magazine10HeroFeats_1+...+done_quest=Magazine10HeroFeats_12", "reward": "asset=PetFoxLouie:1", "opens": "", "completion_flag": "Magazine10_HeroFeats_Hidden", "source_file": "Magazine9TinyHoroscope_final.proto.js", "status": "NEW_REWARD", "notes": "Pet classname временный: TODO_DEV."})
    return rows


def hero_roles() -> list[dict]:
    raw = [
        ("Пекарь", "Подвиг пекаря", "Испечь угощение для дороги."),
        ("Лесник-следопыт", "Подвиг лесника-следопыта", "Найти верную тропинку и не растерять шишки."),
        ("Мастер игрушек", "Подвиг мастера игрушек", "Починить игрушку, которая поднимает настроение."),
        ("Плотник", "Подвиг плотника", "Сколотить крепкую полку для приключений."),
        ("Собиратель", "Подвиг собирателя", "Собрать полезные мелочи без лишней суеты."),
        ("Хранитель сада", "Подвиг хранителя сада", "Ухаживать за садом и защитить ростки."),
        ("Картограф", "Подвиг картографа", "Нарисовать карту, где даже поворот за печкой отмечен."),
        ("Портной", "Подвиг портного", "Сшить походный наряд с тайным карманом."),
        ("Фонарщик", "Подвиг фонарщика", "Зажечь огоньки, чтобы никто не заблудился."),
        ("Смотритель маяка", "Подвиг смотрителя маяка", "Подать световой знак с самой высокой точки."),
        ("Охотник за сокровищами", "Подвиг охотника за сокровищами", "Найти блестящую находку и не зазнаться."),
        ("Почтальон", "Подвиг почтальона", "Доставить важную весточку точно в срок."),
    ]
    return [{"role": role, "quest_title": title, "meaning": meaning} for role, title, meaning in raw]


def task_rows() -> list[dict]:
    rows: list[dict] = []

    def add(q, group, page, no, ttype, title, action="", classname="", param="", amount="TODO_DEV", price="TODO_DEV", hint="", source="", notes=""):
        rows.append({
            "quest_identifier": q,
            "group_identifier": group,
            "page_number": page,
            "task_no": no,
            "task_identifier": "TODO_DEV",
            "type": ttype,
            "action": action,
            "classname": classname,
            "param": param,
            "amount": amount,
            "price": price,
            "title": title,
            "hint": hint,
            "source_task": source,
            "notes": notes,
        })

    add("Magazine10Stone_1", "Magazine10Stone", 3, 1, "action", "Найди алмаз", action="clean_debris", param="TODO_DEV_MagazineStoneDiamond", source="Magazine9Stone_1/e3320", notes="Нужны diamond pile assets и on_activate.")

    flower_titles = [
        ("FlowerMagazine10_PsychotriaCollection1", "Найди алые лепестки"),
        ("FlowerMagazine10_PsychotriaCollection2", "Найди тропические листочки"),
        ("FlowerMagazine10_PsychotriaCollection3", "Найди лесные капельки"),
        ("FlowerMagazine10_PsychotriaCollection4", "Найди цветочные улыбки"),
        ("FlowerMagazine10_PsychotriaCollection5", "Найди солнечную пыльцу"),
        ("FlowerMagazine10_PsychotriaCondition", "Собери цветочные губы"),
    ]
    for i, (classname, title) in enumerate(flower_titles, 1):
        add("Magazine10Flower_1", "Magazine10Flower", 4, i, "get_asset", title, classname=classname, source=f"Magazine9Flower_1/e332{i}", notes="Classname/amount/price уточнить по новой flower collection.")

    add("Magazine10Tree_1", "Magazine10Tree", 5, 1, "have_asset", "Получи леечки для кизила", classname="Magazine10_WateringCan", source="Magazine9Tree_1/e3327")
    add("Magazine10Tree_1", "Magazine10Tree", 5, 2, "get_asset", "Найди садовые бирки", classname="TODO_DEV_DogwoodCollectionItem", source="Magazine9Tree_1/e3328")
    add("Magazine10Tree_1", "Magazine10Tree", 5, 3, "get_asset", "Получи ведра воды", classname="WaterBucket", source="Magazine9Tree_1/e3329")

    add("Magazine10Animal_1", "Magazine10Animal", 6, 1, "get_asset", "Найди полосатые чулочки", classname="TODO_DEV_OkapiCollectionItem", source="Magazine9Animal_1/e3330", notes="Смысловая замена под окапи.")
    add("Magazine10Animal_1", "Magazine10Animal", 6, 2, "action", "Вырасти листья для окапи", action="take_crop", param="TODO_DEV_VALID_CROP", source="Magazine9Animal_1/e3331")
    add("Magazine10Animal_1", "Magazine10Animal", 6, 3, "action", "Подари угощения друзьям", action="send_free_gift", source="Magazine9Animal_1/e3332")

    photo_params = [
        ("MysteryBoxMini_Kitchen", "Сфотографируй Ларчик повара"),
        ("FlowerPot_Magazine3", "Сфотографируй Горшок \"Эльфийский\""),
        ("NY20_Reward_Decor_Lamp", "Сфотографируй Торшер \"Гном\""),
        ("BD15_Reward_Decor_Marionette", "Сфотографируй Марионетки"),
        ("BlackBox45Toy3", "Сфотографируй Трубокота"),
        ("Pet4House", "Сфотографируй домик ушанов"),
        ("Maslenitsa19Sale_Mirror", "Сфотографируй Зеркало \"Подсолнушек\""),
        ("Harpy", "Сфотографируй Гарпию"),
        ("BD13_BouquetFromYaga", "Сфотографируй Букет из мухоморов"),
        ("CupFishermanDayIron", "Сфотографируй Кубок участника"),
    ]
    for i, (param, title) in enumerate(photo_params, 1):
        add("Magazine10Photo_1", "Magazine10Photo", 7, i, "action", title, action="post_photo", param=param, amount=1, price=3, source=f"Magazine9Photo_1/task_{i}", notes="COPY_FROM_MAGAZINE9")

    costume_defs = [
        ("Scandi", 8, "Костюм в скандинавском стиле"),
        ("Fair", 9, "Костюм ярмарочный"),
        ("Atelier", 10, "Костюм в стиле подводного ателье"),
    ]
    boy_parts = [("Body", "Получи основу костюма"), ("Head", "Получи головной убор"), ("Pants", "Получи брюки"), ("Boot", "Получи обувь"), ("Glove", "Получи перчатки"), ("Wings", "Получи декоративную деталь"), ("Neck", "Получи украшение")]
    girl_parts = [("Body", "Получи основу костюма"), ("Head", "Получи головной убор"), ("Earring", "Получи сережки"), ("Boot", "Получи обувь"), ("Glove", "Получи перчатки"), ("Wings", "Получи декоративную деталь"), ("Neck", "Получи украшение")]
    for key, page, title in costume_defs:
        for sex, parts, group_suffix, source_name in [
            ("Boy", boy_parts, "Boy", "Magazine9ClothesBoy_1"),
            ("Girl", girl_parts, "Girl", "Magazine9ClothesGirl_1"),
        ]:
            q = f"Magazine10Clothes{key}{sex}_1"
            group = f"Magazine10Clothes{key}{sex}"
            for idx, (part, part_title) in enumerate(parts, 1):
                add(q, group, page, idx, "have_asset", f"{part_title}: {title}", classname=f"Magazine10Clothes{key}{sex}_{part}", amount=1, price="", source=f"{source_name}/asset_part_{part}", notes="Часть костюма. Реальный classname подтвердить.")
            for service_no in range(8, 11):
                add(q, group, page, service_no, "have_location", "Служебное задание", classname="", amount="", price="", source=f"{source_name}/service_{service_no}", notes="Служебная строка как в Magazine9 clothes.")

    diy_tasks = [
        ("Magazine10Diy_1", 1, "get_asset", "Найди лупы для сыщиков", "", "TODO_DEV_DetectiveMagnifier", ""),
        ("Magazine10Diy_1", 2, "action", "Вырасти кошачью мяту", "take_crop", "", "TODO_DEV_VALID_CROP"),
        ("Magazine10Diy_1", 3, "garbage", "Убери следы лапок в гостях", "", "TODO_DEV_PawPrintGarbage", ""),
        ("Magazine10Diy_2", 1, "garbage", "Убери коробки с уликами", "", "TODO_DEV_EvidenceBoxes", ""),
        ("Magazine10Diy_2", 2, "get_asset", "Найди таблички агентства", "", "TODO_DEV_AgencySignCollection", ""),
        ("Magazine10Diy_2", 3, "action", "Собери ткань для ширмы в гостях", "take_crop_in_guest", "", "TODO_DEV_VALID_CROP"),
        ("Magazine10Diy_3", 1, "action", "Собери цветы для фотозоны в гостях", "take_crop_in_guest", "", "TODO_DEV_VALID_FLOWER"),
        ("Magazine10Diy_3", 2, "action", "Вырасти угощение для котиков", "take_crop", "", "TODO_DEV_VALID_CROP"),
        ("Magazine10Diy_3", 3, "get_asset", "Найди фонарики для витрины", "", "TODO_DEV_PhotozoneLanterns", ""),
    ]
    for q, no, ttype, title, action, classname, param in diy_tasks:
        add(q, "Magazine10Diy", 11, no, ttype, title, action=action, classname=classname, param=param, source=q.replace("10", "9") + f"/task_{no}", notes="Смысловая замена под фотозону.")

    achieve_tasks = [
        ("garbage", "Убери обрезки ткани в мастерской", "", "TODO_DEV_FabricScraps", ""),
        ("action", "Вырасти хлопок для кукол", "take_crop", "", "TODO_DEV_VALID_CROP"),
        ("garbage", "Убери клубки ниток в гостях", "", "TODO_DEV_ThreadTangle", ""),
        ("action", "Покорми любого крохоньку", "feed_any_pet", "", ""),
        ("get_asset", "Найди миниатюрные пуговицы", "", "TODO_DEV_DollButtonsCollection", ""),
        ("action", "Прогони вредителя от тканей", "kill_pest", "", ""),
        ("action", "Помойся перед примеркой", "clean_personage", "", ""),
        ("action", "Получи бесплатные подарки", "receive_free_gift", "", ""),
        ("get_asset", "Найди кукольные выкройки", "", "TODO_DEV_DollPatternCollection", ""),
        ("action", "Подари подарки друзьям", "send_free_gift", "", ""),
    ]
    for i, (ttype, title, action, classname, param) in enumerate(achieve_tasks, 1):
        add("Magazine10Achieve_1", "Magazine10Achieve", 19, i, ttype, title, action=action, classname=classname, param=param, source=f"Magazine9Achieve_1/task_{i}", notes="Смысловая замена под профессию.")

    search_titles = [
        "Найди Кастрюля с черпаком",
        "Найди Подковы",
        "Найди Банка с крышкой",
        "Найди Сережки",
        "Найди Коробка шурупов",
    ]
    for i, title in enumerate(search_titles, 1):
        add(f"Magazine10Search_{i}", "Magazine10Search", 20, 1, "action", title, action="clean_debris", param=f"Magazine10SearchClue_{i}", amount=1, price=15, source=f"Magazine9Search_{i}", notes="COPY_FROM_MAGAZINE9; можно заменить clue titles под новый сюжет.")

    hero_task_sets = [
        [("action", "Вырасти пшеницу", "take_crop", "", "TODO_DEV_VALID_CROP"), ("get_asset", "Найди формочки для пирогов", "", "TODO_DEV_BakerMolds", ""), ("action", "Подари угощения друзьям", "send_free_gift", "", "")],
        [("garbage", "Убери ветки с тропинки", "", "TODO_DEV_ForestBranches", ""), ("get_asset", "Найди следопытские метки", "", "TODO_DEV_TrackerMarks", ""), ("action", "Собери лесные цветы", "take_crop", "", "TODO_DEV_VALID_FLOWER")],
        [("get_asset", "Найди деревянные колесики", "", "TODO_DEV_ToyWheels", ""), ("action", "Вырасти яркие ленточки", "take_crop", "", "TODO_DEV_VALID_CROP"), ("action", "Подари игрушки друзьям", "send_free_gift", "", "")],
        [("garbage", "Убери щепки после работы", "", "TODO_DEV_WoodChips", ""), ("get_asset", "Найди крепкие гвоздики", "", "TODO_DEV_TinyNails", ""), ("action", "Прогони вредителя от досок", "kill_pest", "", "")],
        [("get_asset", "Найди дорожные мешочки", "", "TODO_DEV_GathererBags", ""), ("garbage", "Убери лишние корзинки", "", "TODO_DEV_EmptyBaskets", ""), ("action", "Получи бесплатные подарки", "receive_free_gift", "", "")],
        [("action", "Вырасти садовые цветы", "take_crop", "", "TODO_DEV_VALID_FLOWER"), ("get_asset", "Найди садовые таблички", "", "TODO_DEV_GardenTags", ""), ("garbage", "Убери сухие листья", "", "TODO_DEV_DryLeaves", "")],
        [("get_asset", "Найди кусочки карты", "", "TODO_DEV_MapPieces", ""), ("action", "Собери чернила у соседушек", "receive_free_gift", "", ""), ("garbage", "Убери потерянные компасы", "", "TODO_DEV_LostCompasses", "")],
        [("get_asset", "Найди катушки ниток", "", "TODO_DEV_ThreadSpools", ""), ("action", "Вырасти лен для ткани", "take_crop", "", "TODO_DEV_VALID_CROP"), ("action", "Подари лоскутки друзьям", "send_free_gift", "", "")],
        [("get_asset", "Найди стеклышки для фонарей", "", "TODO_DEV_LanternGlass", ""), ("garbage", "Убери погасшие фитильки", "", "TODO_DEV_OldWicks", ""), ("action", "Зажги помощь соседям", "send_free_gift", "", "")],
        [("garbage", "Убери морскую пену у маяка", "", "TODO_DEV_SeaFoam", ""), ("get_asset", "Найди маячные линзы", "", "TODO_DEV_LighthouseLens", ""), ("action", "Собери сигнальные цветы", "take_crop", "", "TODO_DEV_VALID_FLOWER")],
        [("get_asset", "Найди старые ключики", "", "TODO_DEV_TreasureKeys", ""), ("garbage", "Убери песок с сундуков", "", "TODO_DEV_ChestSand", ""), ("action", "Подари находки друзьям", "send_free_gift", "", "")],
        [("get_asset", "Найди почтовые марки", "", "TODO_DEV_PostStamps", ""), ("action", "Получай весточки от друзей", "receive_free_gift", "", ""), ("garbage", "Убери рассыпанные конверты", "", "TODO_DEV_Envelopes", "")],
    ]
    for idx, tasks in enumerate(hero_task_sets, 1):
        q = f"Magazine10HeroFeats_{idx}"
        for no, (ttype, title, action, classname, param) in enumerate(tasks, 1):
            add(q, "Magazine10HeroFeats", 21, no, ttype, title, action=action, classname=classname, param=param, source=f"Magazine9TinyHoroscope_{idx}/task_{no}", notes="Новые подвиги, экономика TODO_DEV.")
    add("Magazine10HeroFeats_final", "Magazine10HeroFeats", 21, 1, "action", "Забрать лисенка Луи", action="action_Magazine10_HeroFeats_TakeReward", amount=1, price="", source="Magazine9TinyHoroscope_final/e3397", notes="Финальная кнопка награды. Pet classname TODO_DEV.")
    return rows


def root_check_rows() -> list[dict]:
    checks = [
        (1, 3, "have_asset", "Magazine10_Stone_Hidden", 1, "3", "Magazine10Stone_1", ""),
        (2, 4, "have_asset", "Magazine10_Flower_Hidden", 1, "4", "Magazine10Flower_1", ""),
        (3, 5, "have_asset", "Magazine10_Tree_Hidden", 1, "5", "Magazine10Tree_1", ""),
        (4, 6, "have_asset", "Magazine10_Animal_Hidden", 1, "6", "Magazine10Animal_1", "Окапи, хотя page template achieve."),
        (5, 7, "have_asset", "Magazine10_Photo_Hidden", 1, "7", "Magazine10Photo_1", ""),
        (6, 8, "have_asset", "Magazine10_ClothesScandi_Hidden", 1, "8", "Magazine10ClothesScandiBoy/Girl_1", ""),
        (7, 9, "have_asset", "Magazine10_ClothesFair_Hidden", 1, "9", "Magazine10ClothesFairBoy/Girl_1", ""),
        (8, 10, "have_asset", "Magazine10_ClothesAtelier_Hidden", 1, "10", "Magazine10ClothesAtelierBoy/Girl_1", ""),
        (9, 11, "have_asset", "Magazine10_Diy_Hidden", 1, "11", "Magazine10Diy_3", ""),
        (10, 19, "have_asset", "Magazine10_Achieve_Hidden", 1, "19", "Magazine10Achieve_1", ""),
        (11, 20, "have_asset", "Magazine10_Search_Hidden", 1, "20", "Magazine10Search_5", ""),
        (12, 21, "have_asset", "PetFoxLouie", 1, "21", "Magazine10HeroFeats_final", "Можно заменить на Magazine10_HeroFeats_Hidden, если pet asset не подходит для проверки."),
        (13, 22, "action", "Magazine10 / take_magazine_surprise", 1, "", "surprise page", "Финальная кнопка."),
    ]
    return [
        {"check_no": no, "page_number": page, "task_type": ttype, "classname_or_action": target, "amount": amount, "root_task_title": title, "source_completion": source, "notes": notes}
        for no, page, ttype, target, amount, title, source, notes in checks
    ]


def text_rows() -> list[dict]:
    interview = (
        "Д - Кощей Бессмертный, спасибо, что заглянули в \"Домовиту\". Говорят, вы храните столько сундуков, что сами иногда теряете ключи. Это правда?\n\n"
        "Кощей - Клевета с легким звоном правды. Ключи я не теряю, я их прячу в надежные места, а потом проверяю, насколько места надежные. Иногда проверка длится триста лет, но порядок есть порядок.\n\n"
        "Д - А почему вас все считают таким суровым?\n\n"
        "Кощей - Вид у меня деловой. Если стоять у окна в плаще и задумчиво считать монеты, соседи сразу решают: \"Ох, опять что-то замышляет\". А я, может, пирог остужаю и думаю, хватит ли корицы.\n\n"
        "Д - Пирог? Неожиданно!\n\n"
        "Кощей - Бессмертие длинное, дружок. За это время и пироги научишься печь, и носки штопать, и спорить с чайником. Самое трудное - не колдовство, а достать из духовки противень и не забыть рукавички.\n\n"
        "Д - Что бы вы пожелали нашим читателям?\n\n"
        "Кощей - Берегите свои иголочки, ключики и добрые слова. Первые две вещи помогают найти сокровища, а третья - тех, с кем сокровищами приятно делиться."
    )
    return [
        {"page_number": 3, "template": "stone", "field_id": "field1", "target": "text_fields", "text": "Алмаз рождается глубоко в земле, где темно, жарко и совсем не до пустяков. Зато потом он сияет так, будто собрал в себе все солнечные искры. Домовята верят: если держать алмаз рядом с важным делом, мысль становится тверже, а сердце смелее.", "notes": ""},
        {"page_number": 4, "template": "flower", "field_id": "field1", "target": "text_fields", "text": "Психотрия возвышенная словно подмигивает путнику алыми губками. В тропическом лесу такой цветок легко принять за смешную записку от самой природы: не хмурься, дружок, даже листья умеют улыбаться.", "notes": ""},
        {"page_number": 5, "template": "tree", "field_id": "field1", "target": "text_fields", "text": "Кизил не любит суеты: весной тихо раскрывает желтые цветы, а к осени прячет в листве рубиновые ягодки. В домовячьем саду он считается деревцем запасливых хозяев: и красиво, и полезно, и варенье выходит такое, что ложка сама просится в банку.", "notes": ""},
        {"page_number": 6, "template": "achieve", "field_id": "steps", "target": "steps", "text": join(["родственник жирафа, хотя похож на сказочную лошадку", "живет в густых лесах Конго и не любит шум", "полоски на ногах помогают прятаться среди теней", "длинным языком достает листья и чистит мордочку", "у каждого окапи свой неповторимый узор"]), "notes": "Животное в achieve template."},
        {"page_number": 11, "template": "diy", "field_id": "field1", "target": "text_fields", "text": "Мастерим фотозону, где каждый домовенок сможет выглядеть как главный следователь по делу пропавших клубков. Немного дерева, капля тайны, пара кошачьих следов - и детективное агентство готово принимать первых посетителей.", "notes": ""},
        {"page_number": 13, "template": "interior_shop", "field_id": "assets", "target": "assets", "text": join(["Обои \"Теплый лен\"", "Пол \"Светлая доска\"", "Арка \"Макраме\"", "Диван \"Мягкий закат\"", "Кресло \"Ротанг\"", "Столик \"Плетеный круг\"", "Ковер \"Песочные узоры\"", "Торшер \"Тихий вечер\"", "Полка \"Домашние травы\"", "Панно \"Солнечный узел\"", "Пуф \"Пряная охра\"", "Ширма \"Ловец снов\""]), "notes": "12 BOHO-декоров."},
        {"page_number": 17, "template": "cook", "field_id": "field1", "target": "text_fields", "text": "Нарежешь как ролл, подашь как пирог, а довольные соседушки попросят еще кусочек.", "notes": ""},
        {"page_number": 17, "template": "cook", "field_id": "field3", "target": "text_fields", "text": "Подожди 30 минут и получи +400 Силушки.\nГотовить можно неограниченное количество раз!", "notes": "Сохраняет механику Magazine9 cook."},
        {"page_number": 18, "template": "interview", "field_id": "field2", "target": "text_fields", "text": interview, "notes": "Доброе интервью с Кощеем."},
        {"page_number": 19, "template": "achieve", "field_id": "steps", "target": "steps", "text": join(["придумывать характер будущей куклы", "подбирать ткани, ленточки и крошечные пуговицы", "шить аккуратно, даже если иголка больше ботинка", "выбирать прическу, обувь и настроение", "добавлять последнюю деталь, после которой кукла будто улыбается"]), "notes": "Профессия."},
        {"page_number": 21, "template": "tiny_horoscope", "field_id": "rules", "target": "rules", "text": "Выполни 12 добрых подвигов героя и собери все памятные значки. Когда все дела будут закончены, лисенок Луи отправится с тобой в новые приключения.", "notes": "Заменяет Крохо-Гороскоп."},
        {"page_number": 22, "template": "surprise", "field_id": "field1", "target": "text_fields", "text": "Теперь у самого серьезного дела будет самый усатый адрес.", "notes": ""},
        {"page_number": 22, "template": "surprise", "field_id": "field2", "target": "text_fields", "text": "За этим фасадом легко представить шуршащие папки, карту с ниточками и важного котика, который делает вид, что все раскрыл еще утром.", "notes": ""},
        {"page_number": 22, "template": "surprise", "field_id": "field3", "target": "text_fields", "text": "Пусть дом выглядит так, будто в нем каждую пропажу находят по теплому следу и доброму нюху.", "notes": ""},
    ]


def hero_rows() -> list[dict]:
    rows = []
    for idx, role in enumerate(hero_roles(), 1):
        rows.append({
            "position": idx,
            "quest_identifier": f"Magazine10HeroFeats_{idx}",
            "role": role["role"],
            "quest_title": role["quest_title"],
            "badge_asset": f"Magazine10_HeroFeats_Badge_{idx}",
            "meaning": role["meaning"],
            "suggested_task_theme": "см. лист QuestTasks",
            "notes": "На базе Magazine9TinyHoroscope_N, но текст и награда новые.",
        })
    rows.append({"position": "final", "quest_identifier": "Magazine10HeroFeats_final", "role": "Финал", "quest_title": "Лисенок Луи", "badge_asset": "PetFoxLouie", "meaning": "За все 12 подвигов игрок получает питомца лисенка Луи, помощника в приключениях.", "suggested_task_theme": "Забрать награду", "notes": "Pet classname TODO_DEV."})
    return rows


def asset_rows() -> list[dict]:
    assets = [
        ("stone", "TODO_DEV_MagazineStoneDiamond", "Алмаз", "page 3, Magazine10Stone_1", "Новый камень, нужны pile/help assets."),
        ("flower", "FlowerMagazine10_Psychotria", "Цветочные губы / психотрия возвышенная", "page 4, Magazine10Flower_1", "Нужны seed, condition, collection items."),
        ("tree", "Magazine10_Tree_Dogwood", "Кизил", "page 5, Magazine10Tree_1", "Reward tree asset."),
        ("animal_reward", "Magazine10_AnimalReward", "Окапи", "page 6, Magazine10Animal_1", "Reward icon/asset для achieve-страницы."),
        ("photo_reward", "Magazine10_PhotoReward", "Фотографии", "page 7", "Можно копировать механику Magazine9."),
        ("clothes", "Magazine10ClothesScandiBoy/Girl_*", "Костюм в скандинавском стиле", "page 8", "2 sex quest groups, parts classnames TODO_DEV."),
        ("clothes", "Magazine10ClothesFairBoy/Girl_*", "Костюм ярмарочный", "page 9", "2 sex quest groups, parts classnames TODO_DEV."),
        ("clothes", "Magazine10ClothesAtelierBoy/Girl_*", "Костюм в стиле подводного ателье", "page 10", "2 sex quest groups, parts classnames TODO_DEV."),
        ("diy_reward", "Magazine10_DiyReward", "Фотозона \"Детективное агентство котиков\"", "page 11, Magazine10Diy_3", "Финальная награда DIY."),
        ("interior", "Magazine10_BOHO_*", "12 декоров BOHO", "page 13", "См. лист Texts для списка display titles."),
        ("beauty_hair", "Magazine10BeautyHair1..8", "8 причесок", "pages 14-15", "Женские 1-4, мужские 5-8."),
        ("beauty_makeup", "Magazine10BeautyMakeup1..3", "3 образа", "page 16", "Звездная фантазия, Славянская красавица, Скандинавский рунолог."),
        ("beauty_extra", "Magazine10FunnyStylist", "Забавный стилист", "page 16", "Хомяк барбер-стилист вместо органайзера аксессуаров."),
        ("cook", "Magazine10CookFood", "Запеченный ролл-пирог", "page 17", "reward_classname TODO_DEV."),
        ("achieve_reward", "Magazine10_AchieveReward", "Кукольный модельер", "page 19", "Reward icon/asset профессии."),
        ("search", "Magazine10SearchClue_1..5", "Подсказки поиска", "page 20", "Пока копия с новым prefix."),
        ("hero_badges", "Magazine10_HeroFeats_Badge_1..12", "Значки 12 подвигов", "page 21", "Аналог плакатов TinyHoroscope."),
        ("pet", "PetFoxLouie", "Лисенок Луи", "page 21 final", "Реальный pet classname уточнить."),
        ("surprise", "Magazine10_FinalReward", "Фасад \"Детективное агентство котиков\"", "page 22, Magazine10_1", "Главный приз журнала."),
    ]
    return [{"category": c, "placeholder_classname": cl, "display_title": t, "used_in": used, "status": "TODO_DEV" if "TODO" in note or "уточнить" in note else "READY_FOR_DEV_FILL", "notes": note} for c, cl, t, used, note in assets]


def todo_rows() -> list[dict]:
    return [
        {"priority": "P0", "area": "Все quest-файлы", "question": "Выдать уникальные numeric id и task identifier e....", "suggested_default": "Не переносить id из Magazine9.", "owner": "dev"},
        {"priority": "P0", "area": "Assets", "question": "Подтвердить реальные classnames всех новых ассетов.", "suggested_default": "Использовать prefix Magazine10_*.", "owner": "dev/content"},
        {"priority": "P0", "area": "Flower", "question": "Нужны collection_id, recipe_id, asset_seed, asset_for_sale для психотрии.", "suggested_default": "FlowerMagazine10_Psychotria*", "owner": "dev"},
        {"priority": "P0", "area": "HeroFeats", "question": "Какой реальный classname питомца лисенка Луи?", "suggested_default": "PetFoxLouie как временный placeholder.", "owner": "dev/art"},
        {"priority": "P1", "area": "Interior", "question": "Помещает ли template interior_shop 12 assets на одной странице?", "suggested_default": "Если нет, разбить на 2 страницы и перенумеровать дальше.", "owner": "dev"},
        {"priority": "P1", "area": "Animal/Achieve", "question": "Корректно ли template achieve отображает животное окапи с label_title \"Уши, лапы, хвост\"?", "suggested_default": "Использовать steps как факты о животном.", "owner": "dev/ui"},
        {"priority": "P1", "area": "Beauty", "question": "Нужен payment_service_id для makeup и, возможно, Забавного стилиста.", "suggested_default": "Не копировать без проверки старый 1018.", "owner": "dev"},
        {"priority": "P1", "area": "Copy pages", "question": "Photo/Search точно копируем по старым объектам?", "suggested_default": "Пока копия из Magazine9, но проверить доступность params.", "owner": "content/dev"},
        {"priority": "P2", "area": "Spelling", "question": "В публичных текстах писать \"Детективное агентство котиков\"?", "suggested_default": "Да, исправить пользовательское \"агенство\" на нормативное \"агентство\".", "owner": "content"},
        {"priority": "P2", "area": "Economy", "question": "Amounts/prices в новых задачах.", "suggested_default": "Заполнить деву по балансу, в таблице стоит TODO_DEV.", "owner": "dev/balance"},
    ]


def readme_rows() -> list[dict]:
    return [
        {"section": "Назначение", "value": "Рабочая xlsx-таблица для сборки журнала Домовита, выпуск 10."},
        {"section": "Основа", "value": "Изучены magazine/magazine/Magazine9.proto.js и magazine/quest/Magazine9*.proto.js."},
        {"section": "Формат", "value": "В рабочих листах row 1 - название блока, row 2 - типы, row 3 - имена полей, как в местных xlsx-шаблонах."},
        {"section": "Важно", "value": "TODO_DEV означает неизвестные разработческие значения: classnames, id, prices, payment_service_id, collection_id."},
        {"section": "Главная логика", "value": "Magazine10 убирает color_furniture, color_clothes, domovestnik, test; animal заменен на achieve; clothes повторен 3 раза; Крохо-Гороскоп заменен на 12 подвигов героя."},
        {"section": "Соседний документ", "value": "magazine/Magazine10_GDD.md содержит текстовую версию ГДД."},
    ]


def build() -> None:
    wb = make_wb()
    write_table(wb, "README", "Magazine10 dev table: README", ["section", "value"], readme_rows(), ["string", "string"])
    write_table(wb, "Pages", "Magazine10.proto.js: pages", ["input", "output", "page_number", "template", "label_title", "short_description", "label_color", "quest_group", "quest_groups", "view", "source_page_9", "status", "content_spec", "dev_notes"], page_rows(), ["string", "string", "int", "string", "string", "string", "string", "string", "array", "string", "string", "string", "string", "string"])
    write_table(wb, "RemovedPages", "Magazine9 pages removed or replaced in Magazine10", ["source_page_9", "template", "label_title", "decision", "notes"], removed_rows(), ["int", "string", "string", "string", "string"])
    write_table(wb, "QuestFiles", "Magazine10 quest prototype files", ["input", "output", "page_number", "quest_identifier", "group_identifier", "title", "conditions", "reward", "opens", "completion_flag", "source_file", "status", "notes"], quest_file_rows(), ["string", "string", "int", "string", "string", "string", "string", "string", "array", "string", "string", "string", "string"])
    write_table(wb, "QuestTasks", "Magazine10 suggested quest tasks", ["quest_identifier", "group_identifier", "page_number", "task_no", "task_identifier", "type", "action", "classname", "param", "amount", "price", "title", "hint", "source_task", "notes"], task_rows(), ["string", "string", "int", "int", "string", "string", "string", "string", "string", "int", "int", "string", "string", "string", "string"])
    write_table(wb, "RootChecks", "Magazine10_1 root quest checks", ["check_no", "page_number", "task_type", "classname_or_action", "amount", "root_task_title", "source_completion", "notes"], root_check_rows(), ["int", "int", "string", "string", "int", "string", "string", "string"])
    write_table(wb, "Texts", "Magazine10 page text fields", ["page_number", "template", "field_id", "target", "text", "notes"], text_rows(), ["int", "string", "string", "string", "string", "string"])
    write_table(wb, "HeroFeats", "12 подвигов героя", ["position", "quest_identifier", "role", "quest_title", "badge_asset", "meaning", "suggested_task_theme", "notes"], hero_rows(), ["string", "string", "string", "string", "string", "string", "string", "string"])
    write_table(wb, "Assets", "Magazine10 asset placeholders", ["category", "placeholder_classname", "display_title", "used_in", "status", "notes"], asset_rows(), ["string", "string", "string", "string", "string", "string"])
    write_table(wb, "TODO", "Questions and TODO for developers", ["priority", "area", "question", "suggested_default", "owner"], todo_rows(), ["string", "string", "string", "string", "string"])

    wb.save(OUT_PATH)


def smoke_test() -> None:
    wb = load_workbook(OUT_PATH, data_only=False)
    expected = {"README", "Pages", "RemovedPages", "QuestFiles", "QuestTasks", "RootChecks", "Texts", "HeroFeats", "Assets", "TODO"}
    missing = expected.difference(wb.sheetnames)
    if missing:
        raise AssertionError(f"Missing sheets: {sorted(missing)}")
    text = ""
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            text += " ".join("" if v is None else str(v) for v in row) + "\n"
    bad_markers = ["????", "Рџ", "Рє", "Рё"]
    found = [m for m in bad_markers if m in text]
    if found:
        raise AssertionError(f"Encoding markers found: {found}")
    if "Окапи" not in text or "Кощей" not in text or "PetFoxLouie" not in text:
        raise AssertionError("Key content missing")
    print(f"saved: {OUT_PATH}")
    print(f"sheets: {wb.sheetnames}")
    print(f"QuestTasks rows: {wb['QuestTasks'].max_row - 3}")
    print(f"Pages rows: {wb['Pages'].max_row - 3}")


if __name__ == "__main__":
    build()
    smoke_test()
